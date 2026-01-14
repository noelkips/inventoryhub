from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Q, F, Case, When, IntegerField, Count
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseRedirect
from django.template.loader import get_template
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta, datetime
from io import TextIOWrapper

# Models
from devices.models import CustomUser, DeviceUserHistory, Import, Centre, Notification, PendingUpdate, Department
from it_operations.models import BackupRegistry, WorkPlan, IncidentReport, MissionCriticalAsset, WorkPlanTask
from devices.forms import ClearanceForm
from ppm.models import PPMTask, PPMPeriod, PPMActivity

# Third-party & Standard Library
import csv
import logging
from io import BytesIO

# Excel (openpyxl)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# PDF (ReportLab)
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    Frame, PageTemplate
)

# Django Shortcuts and HTTP
from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404

# Logging
logger = logging.getLogger(__name__)

#
# ==============================================================================
#  1. CORRECTED FILE HANDLING LOGIC
# ==============================================================================
#


def handle_uploaded_file(file, user, centre, department, category):
    """
    Processes an uploaded CSV file to bulk-create Import records.
    Centre, department, and category are provided from the form, NOT from CSV.
    
    Args:
        file: The uploaded CSV file
        user: The user uploading the file
        centre: Centre object (pre-selected from form)
        department: Department object (pre-selected from form)
        category: Category string (pre-selected from form)
    
    Returns:
        Dictionary with import statistics
    """
    
    # Simplified header mapping - NO centre, department, or category in CSV
    header_mapping = {
        'hardware': 'hardware',
        'system_model': 'system_model',
        'processor': 'processor',
        'ram_gb': 'ram_gb',
        'hdd_gb': 'hdd_gb',
        'serial_number': 'serial_number',
        'assignee_first_name': 'assignee_first_name',
        'assignee_last_name': 'assignee_last_name',
        'assignee_email_address': 'assignee_email_address',
        'device_condition': 'device_condition',
        'status': 'status',
        'date': 'date',
    }

    stats = {
        'total_rows': 0,
        'created_count': 0,
        'skipped_existing': 0,
        'skipped_validation': 0,
    }

    try:
        file.seek(0)
        decoded_file = TextIOWrapper(file.file, encoding='utf-8-sig')
        reader = csv.reader(decoded_file)
        headers = next(reader, None)

        if not headers:
            raise ValueError("CSV file is empty or invalid.")

        headers = [h.lower().strip() for h in headers]

        # Only serial_number is required in CSV
        if 'serial_number' not in headers:
            raise ValueError("Missing required header: 'serial_number'")

        serial_number_index = headers.index('serial_number')
        
        import_instances = []
        admins = CustomUser.objects.filter(is_superuser=True, is_trainer=False)

        for row in reader:
            if not any(row):  # Skip empty rows
                continue
            
            stats['total_rows'] += 1

            # Get serial number
            serial_number = row[serial_number_index].strip() if serial_number_index < len(row) else None
            if not serial_number:
                logger.warning("Skipping row with empty serial number.")
                stats['skipped_validation'] += 1
                continue 
            
            # Check if serial number already exists
            if Import.objects.filter(serial_number=serial_number).exists():
                logger.info(f"Skipping existing serial number: {serial_number}")
                stats['skipped_existing'] += 1
                continue 

            # Create instance with form-provided values
            import_instance = Import(
                added_by=user,
                serial_number=serial_number,
                centre=centre,  # From form
                department=department,  # From form
                category=category  # From form
            )

            # Map all other fields from CSV
            for header, value in zip(headers, row):
                value = value.strip() if value else ''
                field_name = header_mapping.get(header)

                if not field_name or field_name == 'serial_number':
                    continue
                
                if field_name == 'date' and value:
                    date_value = None
                    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                        try:
                            date_value = datetime.strptime(value, fmt).date()
                            break
                        except ValueError:
                            continue
                    setattr(import_instance, field_name, date_value)
                else:
                    setattr(import_instance, field_name, value or None)

            # Set approval status
            import_instance.is_approved = not user.is_trainer
            import_instance.approved_by = user if not user.is_trainer and user.is_superuser else None
            
            # Set date if not provided
            if not import_instance.date:
                import_instance.date = timezone.now().date()
            
            import_instances.append(import_instance)

        # Bulk create
        if import_instances:
            with transaction.atomic():
                created_imports = Import.objects.bulk_create(
                    import_instances, 
                    batch_size=500, 
                    ignore_conflicts=True
                )
                stats['created_count'] = len(created_imports)
                
                # Create notifications for trainers
                if user.is_trainer:
                    created_serials = [imp.serial_number for imp in created_imports]
                    newly_created_imports = Import.objects.filter(serial_number__in=created_serials)
                    
                    for import_obj in newly_created_imports:
                        for admin in admins:
                            Notification.objects.create(
                                user=admin,
                                message=f"New device added by {user.username} with serial number {import_obj.serial_number} awaiting approval.",
                                content_type=ContentType.objects.get_for_model(Import),
                                object_id=import_obj.pk
                            )
        
        return stats

    except ValueError as ve:
        logger.warning(f"Validation error during CSV import for user {user.username}: {str(ve)}")
        raise
    except Exception as e:
        logger.error(f"Error processing CSV file for user {user.username}: {str(e)}", exc_info=True)
        raise ValueError(f"An unexpected error occurred: {str(e)}")
    finally:
        if 'decoded_file' in locals() and decoded_file:
            decoded_file.detach()

@login_required
def download_csv_template(request):
    """
    Generates a CSV template file for device bulk upload.
    The headers match the keys in the 'header_mapping' of handle_uploaded_file.
    """
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="device_upload_template.csv"'},
    )

    # These headers are the *keys* from your header_mapping
    # We put serial_number first as it's the most critical
    headers = [
        'serial_number',
        'hardware',
        'system_model',
        'processor',
        'ram_gb',
        'hdd_gb',
        'assignee_first_name',
        'assignee_last_name',
        'assignee_email_address',
        'device_condition',
        'status',
        'date',
    ]

    writer = csv.writer(response)
    writer.writerow(headers)  # Write only the header row

    return response


@login_required
def import_add(request):
    user = request.user
    """Add new device - single or bulk upload"""
    if request.method == 'POST':
        
        # === FILE UPLOAD (BULK) ===
        if 'file' in request.FILES:
            file = request.FILES['file']
            if not file.name.lower().endswith('.csv'):
                messages.error(request, "Only CSV files are accepted.")
                return redirect('import_add')
            
            try:
                # Get centre, department, and category from form
                centre_id = request.POST.get('bulk_centre')
                department_id = request.POST.get('bulk_department')
                category = request.POST.get('bulk_category')

                # Validate form inputs
                if not department_id:
                    messages.error(request, "Please select a department before uploading.")
                    return redirect('import_add')
                
                if not category:
                    messages.error(request, "Please select a device category before uploading.")
                    return redirect('import_add')

                # Get centre object
                if user.is_trainer:
                    if not request.user.centre:
                        messages.error(request, "Your account is not assigned to a centre. Please contact an administrator.")
                        return redirect('import_add')
                    centre = request.user.centre
                else:
                    if not centre_id:
                        messages.error(request, "Please select a centre before uploading.")
                        return redirect('import_add')
                    try:
                        centre = Centre.objects.get(id=centre_id)
                    except Centre.DoesNotExist:
                        messages.error(request, "Invalid centre selected.")
                        return redirect('import_add')

                # Get department object
                try:
                    department = Department.objects.get(id=department_id)
                except Department.DoesNotExist:
                    messages.error(request, "Invalid department selected.")
                    return redirect('import_add')

                # Process CSV with form-provided values
                upload_stats = handle_uploaded_file(file, request.user, centre, department, category)

                created_count = upload_stats.get('created_count', 0)
                skipped_existing = upload_stats.get('skipped_existing', 0)
                skipped_validation = upload_stats.get('skipped_validation', 0)

                # Build success/warning messages
                if created_count > 0:
                    approval_msg = " Pending approval." if request.user.is_trainer else ""
                    messages.success(
                        request, 
                        f"Successfully imported {created_count} new devices to {centre.name} - {department.name} ({category}).{approval_msg}"
                    )

                if skipped_existing > 0:
                    messages.warning(
                        request, 
                        f"Skipped {skipped_existing} devices because their serial numbers already exist."
                    )
                
                if skipped_validation > 0:
                    messages.warning(
                        request, 
                        f"Skipped {skipped_validation} rows due to validation errors (e.g., missing serial number)."
                    )

                if created_count == 0 and skipped_existing == 0:
                    messages.info(request, "No devices were imported. Please check your CSV file.")

                return redirect('display_approved_imports')

            except ValueError as ve:
                messages.error(request, str(ve))
                return redirect('import_add')
            except Exception as e:
                logger.error(f"Error during bulk upload: {str(e)}", exc_info=True)
                messages.error(request, f"Error processing CSV file: {str(e)}")
                return redirect('import_add')
        
        # === MANUAL SINGLE DEVICE SUBMISSION ===
        else:
            try:
                with transaction.atomic():
                    # Get centre
                    if request.user.is_trainer:
                        if not request.user.centre:
                            messages.error(request, "Your account is not assigned to a centre.")
                            return redirect('import_add')
                        centre = request.user.centre
                    else:
                        centre_id = request.POST.get('centre')
                        if not centre_id:
                            messages.error(request, "Centre is required.")
                            return redirect('import_add')
                        centre = Centre.objects.get(id=centre_id)
                    
                    # Get department
                    department_id = request.POST.get('department')
                    if not department_id:
                        messages.error(request, "Department is required.")
                        return redirect('import_add')
                    department = Department.objects.get(id=department_id)
                    
                    # Get category
                    category = request.POST.get('category')
                    if not category:
                        messages.error(request, "Device category is required.")
                        return redirect('import_add')
                    
                    # Get serial number
                    serial_number = request.POST.get('serial_number')
                    if not serial_number:
                        messages.error(request, "Serial Number is required.")
                        return redirect('import_add')
                    
                    # Check for duplicate serial number
                    if Import.objects.filter(serial_number=serial_number).exists():
                        messages.error(request, f"Serial number {serial_number} already exists.")
                        return redirect('import_add')

                    # Create device instance
                    import_instance = Import(
                        added_by=request.user,
                        centre=centre,
                        department=department,
                        category=category,
                        hardware=request.POST.get('hardware'),
                        system_model=request.POST.get('system_model'),
                        processor=request.POST.get('processor'),
                        ram_gb=request.POST.get('ram_gb'),
                        hdd_gb=request.POST.get('hdd_gb'),
                        serial_number=serial_number,
                        assignee_first_name=request.POST.get('assignee_first_name'),
                        assignee_last_name=request.POST.get('assignee_last_name'),
                        assignee_email_address=request.POST.get('assignee_email_address'),
                        device_condition=request.POST.get('device_condition'),
                        status=request.POST.get('status'),
                        date=timezone.now().date(),
                        is_approved=not request.user.is_trainer,
                        approved_by=request.user if not request.user.is_trainer and request.user.is_superuser else None
                    )
                    import_instance.save()
                    
                    # Create notification for trainers
                    if request.user.is_trainer:
                        admins = CustomUser.objects.filter(is_superuser=True, is_trainer=False)
                        for admin in admins:
                            Notification.objects.create(
                                user=admin,
                                message=f"New device added by {request.user.username} with serial number {import_instance.serial_number} awaiting approval.",
                                content_type=ContentType.objects.get_for_model(Import),
                                object_id=import_instance.pk
                            )
                    
                    approval_msg = " Pending approval." if request.user.is_trainer else ""
                    messages.success(request, f"Device added successfully.{approval_msg}")
                    return redirect('display_approved_imports')
            
            except Centre.DoesNotExist:
                messages.error(request, "Invalid centre selected.")
                return redirect('import_add')
            except Department.DoesNotExist:
                messages.error(request, "Invalid department selected.")
                return redirect('import_add')
            except Exception as e:
                logger.error(f"Error adding manual device: {str(e)}", exc_info=True)
                messages.error(request, f"Error adding device: {str(e)}")
                return redirect('import_add')

    # === GET REQUEST - Show Form ===
    all_centres = Centre.objects.all().order_by('name')
    all_departments = Department.objects.all().order_by('name')
    
    context = {
        'centres': all_centres,
        'departments': all_departments
    }
    return render(request, 'import/add.html', context)


@login_required
def import_update(request, pk):
    """Update existing device"""
    import_instance = get_object_or_404(Import, pk=pk)
    
    # Permission check for trainers
    if request.user.is_trainer and import_instance.centre != request.user.centre:
        messages.error(request, "You can only update records for your own centre.")
        return redirect('display_approved_imports')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get form values
                department_id = request.POST.get('department', '').strip()
                category = request.POST.get('category', '').strip()
                serial_number = request.POST.get('serial_number', '').strip()
                
                # Handle centre (trainers can't change it)
                if request.user.is_trainer:
                    centre = request.user.centre
                else:
                    centre_id = request.POST.get('centre', '').strip()
                    centre = Centre.objects.get(id=centre_id) if centre_id else None

                # Validate department
                department = Department.objects.get(id=department_id) if department_id else None
                if not department:
                    messages.error(request, "Department is required.")
                    return redirect('display_approved_imports')

                # Validate category
                if not category:
                    messages.error(request, "Device category is required.")
                    return redirect('display_approved_imports')

                # Validate serial number
                if serial_number and Import.objects.filter(serial_number=serial_number).exclude(id=pk).exists():
                    messages.error(request, f"Serial number {serial_number} already exists.")
                    return redirect('display_approved_imports')

                # Parse date
                date_str = request.POST.get('date', '').strip()
                date_value = None
                if date_str:
                    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                        try:
                            date_value = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue

                # Collect changed fields
                fields_to_update = {}
                form_data = {
                    'centre': centre,
                    'department': department,
                    'category': category,
                    'hardware': request.POST.get('hardware', '').strip(),
                    'system_model': request.POST.get('system_model', '').strip(),
                    'processor': request.POST.get('processor', '').strip(),
                    'ram_gb': request.POST.get('ram_gb', '').strip(),
                    'hdd_gb': request.POST.get('hdd_gb', '').strip(),
                    'serial_number': serial_number,
                    'assignee_first_name': request.POST.get('assignee_first_name', '').strip(),
                    'assignee_last_name': request.POST.get('assignee_last_name', '').strip(),
                    'assignee_email_address': request.POST.get('assignee_email_address', '').strip(),
                    'device_condition': request.POST.get('device_condition', '').strip(),
                    'status': request.POST.get('status', '').strip(),
                    'reason_for_update': request.POST.get('reason_for_update', '').strip(),
                    'date': date_value,
                }

                # Compare with current values
                for field, new_value in form_data.items():
                    current_value = getattr(import_instance, field, None)
                    
                    if field in ['centre', 'department']:
                        if new_value != current_value:
                            fields_to_update[field] = new_value
                    elif field in ['date']:
                        if new_value and new_value != current_value:
                            fields_to_update[field] = new_value
                    else:
                        current_str = str(current_value) if current_value else ''
                        new_str = str(new_value) if new_value else ''
                        if new_str and new_str != current_str:
                            fields_to_update[field] = new_value

                if not fields_to_update:
                    messages.info(request, "No changes detected.")
                    return redirect('display_approved_imports')

                # Handle trainer updates (requires approval)
                if request.user.is_trainer:
                    reason_for_update = request.POST.get('reason_for_update', '').strip()
                    if not reason_for_update:
                        messages.error(request, "Reason for update is required for trainers.")
                        return redirect('display_approved_imports')
                    
                    pending_update = PendingUpdate.objects.create(
                        import_record=import_instance,
                        **fields_to_update,
                        updated_by=request.user
                    )
                    import_instance.is_approved = False
                    import_instance.approved_by = None
                    import_instance.save()
                    
                    # Notify admins
                    admins = CustomUser.objects.filter(is_superuser=True, is_trainer=False)
                    for admin in admins:
                        Notification.objects.create(
                            user=admin,
                            message=f"Update to device {import_instance.serial_number} by {request.user.username} awaiting approval.",
                            content_type=ContentType.objects.get_for_model(PendingUpdate),
                            object_id=pending_update.pk
                        )
                    messages.success(request, "Update submitted for approval.")
                    return redirect('notifications_view')
                else:
                    # Direct update for non-trainers
                    for field, value in fields_to_update.items():
                        setattr(import_instance, field, value)
                    
                    import_instance.is_approved = True if request.user.is_superuser else import_instance.is_approved
                    import_instance.approved_by = request.user if request.user.is_superuser else import_instance.approved_by
                    import_instance.save()
                    
                    messages.success(request, "Device updated successfully.")
                    return redirect('display_approved_imports')
                    
        except Department.DoesNotExist:
            messages.error(request, "Invalid department selected.")
            return redirect('display_approved_imports')
        except Centre.DoesNotExist:
            messages.error(request, "Invalid centre selected.")
            return redirect('display_approved_imports')
        except Exception as e:
            logger.error(f"Error updating device {import_instance.serial_number}: {str(e)}", exc_info=True)
            messages.error(request, f"Error updating device: {str(e)}")
            return redirect('display_approved_imports')
    
    # GET request - show edit form
    return render(request, 'import/edit.html', {
        'import_instance': import_instance,
        'centres': Centre.objects.all().order_by('name'),
        'departments': Department.objects.all().order_by('name')
    })




@login_required
@user_passes_test(lambda u: u.is_superuser and not u.is_trainer)
def import_approve(request, pk):
    import_instance = get_object_or_404(Import, pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            pending_update = PendingUpdate.objects.filter(import_record=import_instance).order_by('-created_at').first()
            if pending_update:
                # Save pk before deleting
                pending_update_id = pending_update.pk  

                # Apply updates
                import_instance.centre = pending_update.centre
                import_instance.department = pending_update.department
                import_instance.hardware = pending_update.hardware
                import_instance.system_model = pending_update.system_model
                import_instance.processor = pending_update.processor
                import_instance.ram_gb = pending_update.ram_gb
                import_instance.hdd_gb = pending_update.hdd_gb
                import_instance.serial_number = pending_update.serial_number
                import_instance.assignee_first_name = pending_update.assignee_first_name
                import_instance.assignee_last_name = pending_update.assignee_last_name
                import_instance.assignee_email_address = pending_update.assignee_email_address
                import_instance.device_condition = pending_update.device_condition
                import_instance.status = pending_update.status
                import_instance.date = pending_update.date if pending_update.date else timezone.now().date()
                import_instance.reason_for_update = pending_update.reason_for_update
                import_instance.is_approved = True
                import_instance.approved_by = request.user
                import_instance.save()

                # Delete pending update after saving
                pending_update.delete()

                # Mark related notifications as read (for admins only)
                content_type = ContentType.objects.get_for_model(PendingUpdate)
                Notification.objects.filter(
                    content_type=content_type,
                    object_id=pending_update_id,
                    user__is_superuser=True,
                    user__is_trainer=False,
                    is_read=False
                ).update(is_read=True, responded_by=request.user)

                messages.success(request, f"Device {import_instance.serial_number} update approved.")
            else:
                import_instance.is_approved = True
                import_instance.approved_by = request.user
                import_instance.save()

                # Mark related notifications as read (for admins only)
                content_type = ContentType.objects.get_for_model(Import)
                Notification.objects.filter(
                    content_type=content_type,
                    object_id=import_instance.pk,
                    user__is_superuser=True,
                    user__is_trainer=False,
                    is_read=False
                ).update(is_read=True, responded_by=request.user)

                messages.success(request, f"Device {import_instance.serial_number} approved.")

            return redirect('display_unapproved_imports')

    return redirect('display_unapproved_imports')



@login_required
@user_passes_test(lambda u: u.is_superuser and not u.is_trainer)
def import_reject(request, pk):
    import_instance = get_object_or_404(Import, pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            pending_update = PendingUpdate.objects.filter(import_record=import_instance).order_by('-created_at').first()
            if pending_update:
                pending_update.pending_clarification = True
                pending_update.save()
                # Notify trainer for clarification (single notification)
                trainer = pending_update.updated_by
                if trainer:
                    content_type = ContentType.objects.get_for_model(PendingUpdate)
                    if not Notification.objects.filter(user=trainer, content_type=content_type, object_id=pending_update.pk).exists():
                        Notification.objects.create(
                            user=trainer,
                            message=f"Your update for device {pending_update.serial_number} was rejected. Please provide clarification.",
                            content_type=content_type,
                            object_id=pending_update.pk
                        )
                # Mark notifications as read for admins only
                content_type = ContentType.objects.get_for_model(PendingUpdate)
                Notification.objects.filter(
                    content_type=content_type,
                    object_id=pending_update.pk,
                    user__is_superuser=True,
                    user__is_trainer=False,
                    is_read=False
                ).update(is_read=True, responded_by=request.user)
                messages.success(request, f"Update for device {pending_update.serial_number} sent back for clarification.")
            else:
                # For new import requests, mark as pending clarification
                import_instance.pending_clarification = True
                import_instance.save()
                # Notify trainer for clarification (single notification)
                trainer = import_instance.added_by
                if trainer:
                    content_type = ContentType.objects.get_for_model(Import)
                    if not Notification.objects.filter(user=trainer, content_type=content_type, object_id=import_instance.pk).exists():
                        Notification.objects.create(
                            user=trainer,
                            message=f"Your import request for device {import_instance.serial_number} was rejected. Please provide clarification.",
                            content_type=content_type,
                            object_id=import_instance.pk
                        )
                # Mark notifications as read for admins only
                content_type = ContentType.objects.get_for_model(Import)
                Notification.objects.filter(
                    content_type=content_type,
                    object_id=import_instance.pk,
                    user__is_superuser=True,
                    user__is_trainer=False,
                    is_read=False
                ).update(is_read=True, responded_by=request.user)
                messages.success(request, f"Import request for device {import_instance.serial_number} sent back for clarification.")
            return redirect('display_unapproved_imports')
    return redirect('display_unapproved_imports')

@login_required
@user_passes_test(lambda u: u.is_superuser and not u.is_trainer)
def import_approve_all(request):
    if request.method == 'POST':
        page_number = request.GET.get('page', '1')
        items_per_page = request.GET.get('items_per_page', '10')
        search_query = request.GET.get('search', '')
        
        try:
            items_per_page = int(items_per_page)
            if items_per_page not in [10, 25, 50, 100, 500]:
                items_per_page = 10
        except ValueError:
            items_per_page = 10
        try:
            page_number = int(page_number) if page_number else 1
        except ValueError:
            page_number = 1

        data = Import.objects.filter(is_approved=False, is_disposed=False)
        if search_query:
            query = (
                Q(centre__name__icontains=search_query) |
                Q(centre__centre_code__icontains=search_query) |
                Q(department__name__icontains=search_query) |
                Q(hardware__icontains=search_query) |
                Q(system_model__icontains=search_query) |
                Q(processor__icontains=search_query) |
                Q(ram_gb__icontains=search_query) |
                Q(hdd_gb__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(assignee_first_name__icontains=search_query) |
                Q(assignee_last_name__icontains=search_query) |
                Q(assignee_email_address__icontains=search_query) |
                Q(device_condition__icontains=search_query) |
                Q(status__icontains=search_query) |
                Q(date__icontains=search_query) |
                Q(reason_for_update__icontains=search_query)
            )
            data = data.filter(query)

        paginator = Paginator(data, items_per_page)
        try:
            data_on_page = paginator.page(page_number)
        except PageNotAnInteger:
            data_on_page = paginator.page(1)
        except EmptyPage:
            data_on_page = paginator.page(paginator.num_pages)

        approved_count = 0
        with transaction.atomic():
            for item in data_on_page:
                pending_update = PendingUpdate.objects.filter(import_record=item).order_by('-created_at').first()
                if pending_update:
                    item.centre = pending_update.centre
                    item.department = pending_update.department
                    item.hardware = pending_update.hardware
                    item.system_model = pending_update.system_model
                    item.processor = pending_update.processor
                    item.ram_gb = pending_update.ram_gb
                    item.hdd_gb = pending_update.hdd_gb
                    item.serial_number = pending_update.serial_number
                    item.assignee_first_name = pending_update.assignee_first_name
                    item.assignee_last_name = pending_update.assignee_last_name
                    item.assignee_email_address = pending_update.assignee_email_address
                    item.device_condition = pending_update.device_condition
                    item.status = pending_update.status
                    item.date = pending_update.date if pending_update.date else item.date or timezone.now().date()
                    item.reason_for_update = pending_update.reason_for_update
                    item.is_approved = True
                    item.approved_by = request.user
                    item.save()
                    pending_update.delete()
                    content_type = ContentType.objects.get_for_model(PendingUpdate)
                    Notification.objects.filter(
                        content_type=content_type,
                        object_id=pending_update.pk,
                        user__is_superuser=True,
                        user__is_trainer=False,
                        is_read=False
                    ).update(is_read=True, responded_by=request.user)
                    approved_count += 1
                elif not item.is_approved:
                    item.is_approved = True
                    item.approved_by = request.user
                    item.save()
                    content_type = ContentType.objects.get_for_model(Import)
                    Notification.objects.filter(
                        content_type=content_type,
                        object_id=item.pk,
                        user__is_superuser=True,
                        user__is_trainer=False,
                        is_read=False
                    ).update(is_read=True, responded_by=request.user)
                    approved_count += 1

        if approved_count > 0:
            messages.success(request, f"{approved_count} device(s) approved successfully.")
        else:
            messages.info(request, "No unapproved devices to approve on this page.")
        
        redirect_url = reverse('display_unapproved_imports')
        query_params = [f"page={page_number}", f"items_per_page={items_per_page}"]
        if search_query:
            query_params.append(f"search={search_query}")
        redirect_url += "?" + "&".join(query_params)
        return redirect(redirect_url)
    return redirect('display_unapproved_imports')

def _can_delete(user):
    """Only IT Manager or Senior IT Officer can delete."""
    return user.is_it_manager or user.is_senior_it_officer


@login_required
@user_passes_test(_can_delete, login_url='display_approved_imports')
def import_delete(request, pk):
    import_instance = get_object_or_404(Import, pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            serial_number = import_instance.serial_number
            import_instance.delete()

            # Mark any related notifications as read
            content_type = ContentType.objects.get_for_model(Import)
            Notification.objects.filter(
                content_type=content_type,
                object_id=pk,
                is_read=False
            ).update(is_read=True, responded_by=request.user)

            messages.success(request, f"Device {serial_number} deleted successfully.")
        return redirect('display_approved_imports')

    # If GET → just redirect (no form shown)
    return redirect('display_approved_imports')

@login_required
def notifications_view(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    # For admins, exclude notifications already responded to unless it's an unresponded approval request
    if request.user.is_superuser and not request.user.is_trainer:
        content_types = [ContentType.objects.get_for_model(Import), ContentType.objects.get_for_model(PendingUpdate)]
        notifications = notifications.exclude(
            responded_by__isnull=False
        ).filter(
            content_type__in=content_types,
            is_read=False
        ) | notifications.filter(
            responded_by__isnull=True,
            content_type__in=content_types,
            is_read=False
        )
    return render(request, 'notifications.html', {'notifications': notifications})


@login_required
def clear_all_notifications(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        messages.success(request, "All notifications cleared.")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/dashboard/'))
    return HttpResponseRedirect('/dashboard/')


@login_required
def device_history(request, pk):
    device = get_object_or_404(Import, pk=pk)
    history = device.history.all().order_by('-history_date')
    history_data = []
    
    for record in history:
        diff = {}
        if record.prev_record:
            changes = record.diff_against(record.prev_record)
            for change in changes.changes:
                if hasattr(change, 'field') and hasattr(change, 'old') and hasattr(change, 'new'):
                    # Skip if old and new are the same or both None/N/A/empty
                    old_str = str(change.old) if change.old is not None else ''
                    new_str = str(change.new) if change.new is not None else ''
                    if old_str == new_str or (old_str in ['', 'N/A', 'None'] and new_str in ['', 'N/A', 'None']):
                        continue
                    # Resolve human-readable values
                    if change.field == 'centre':
                        old_value = Centre.objects.get(pk=change.old).name if change.old and Centre.objects.filter(pk=change.old).exists() else 'N/A'
                        new_value = Centre.objects.get(pk=change.new).name if change.new and Centre.objects.filter(pk=change.new).exists() else 'N/A'
                    elif change.field == 'approved_by':
                        old_value = CustomUser.objects.get(pk=change.old).username if change.old and CustomUser.objects.filter(pk=change.old).exists() else 'N/A'
                        new_value = CustomUser.objects.get(pk=change.new).username if change.new and CustomUser.objects.filter(pk=change.new).exists() else 'N/A'
                    elif change.field == 'added_by':
                        old_value = CustomUser.objects.get(pk=change.old).username if change.old and CustomUser.objects.filter(pk=change.old).exists() else 'N/A'
                        new_value = CustomUser.objects.get(pk=change.new).username if change.new and CustomUser.objects.filter(pk=change.new).exists() else 'N/A'
                    elif change.field == 'department':
                        old_value = Department.objects.get(pk=change.old).name if change.old and Department.objects.filter(pk=change.old).exists() else 'N/A'
                        new_value = Department.objects.get(pk=change.new).name if change.new and Department.objects.filter(pk=change.new).exists() else 'N/A'
                    elif change.field == 'is_approved':
                        old_value = 'Yes' if change.old == 'True' else 'No' if change.old == 'False' else 'N/A'
                        new_value = 'Yes' if change.new == 'True' else 'No' if change.new == 'False' else 'N/A'
                    else:
                        old_value = change.old if change.old is not None else 'N/A'
                        new_value = change.new if change.new is not None else 'N/A'
                    # Use human-readable field names
                    field_names = {
                        'centre': 'Centre',
                        'department': 'Department',
                        'hardware': 'Hardware',
                        'system_model': 'System Model',
                        'processor': 'Processor',
                        'ram_gb': 'RAM (GB)',
                        'hdd_gb': 'HDD (GB)',
                        'serial_number': 'Serial Number',
                        'assignee_first_name': 'Assignee First Name',
                        'assignee_last_name': 'Assignee Last Name',
                        'assignee_email_address': 'Assignee Email Address',
                        'device_condition': 'Device Condition',
                        'status': 'Status',
                        'date': 'Date',
                        'added_by': 'Added By',
                        'approved_by': 'Approved By',
                        'is_approved': 'Is Approved',
                        'reason_for_update': 'Reason for Update',
                        'disposal_reason': 'Disposal Reason',
                    }
                    field_name = field_names.get(change.field, change.field.replace('_', ' ').title())
                    diff[field_name] = {'old': old_value, 'new': new_value}
        
        # Determine the user who made the change
        user = record.history_user.username if record.history_user else None
        if not user:
            # If history_user is not set, log a warning and use a fallback (should be rare with proper save)
            import logging
            logging.warning(f"No history_user found for record ID {record.id} on device {device.serial_number} at {timezone.now()}")
            user = request.user.username if request.user.is_authenticated else 'Unknown'
        else:
            user = record.history_user.username

        history_data.append({
            'record': record,
            'diff': diff,
            'change_type': record.get_history_type_display() or record.history_type,
            'user': user
        })

    # Fetch user history
    user_history = device.user_history.all().order_by('assigned_date').values(
        'assignee_first_name',
        'assignee_last_name',
        'assignee_email_address',
        'assigned_by__username',
        'assigned_date',
        'cleared_date'
    )
    user_history_data = [
        {
            'assignee_name': f"{entry['assignee_first_name'] or ''} {entry['assignee_last_name'] or ''}".strip() or 'N/A',
            'email': entry['assignee_email_address'] or 'N/A',
            'assigned_by': entry['assigned_by__username'] or 'N/A',
            'assigned_date': entry['assigned_date'],
            'cleared_date': entry['cleared_date']
        } for entry in user_history
    ]

    return render(request, 'import/device_history.html', {
        'device': device,
        'history': history_data,
        'user_history': user_history_data
    })
# Make sure you have these imports at the top of your views.py

@login_required
def export_to_excel(request):
    # === GET PARAMETERS ===
    scope = request.GET.get('scope', 'page')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', '1')
    items_per_page = request.GET.get('items_per_page', '10')
    
    # ***** NEW: Get the view context *****
    view_context = request.GET.get('view_context', 'display_approved_imports')

    # ---- pagination / validation -------------------------------------------------
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10
    try:
        page_number = int(page_number) if page_number else 1
    except ValueError:
        page_number = 1

    # ---- base queryset ----------------------------------------------------------
    if request.user.is_superuser:
        base_qs = Import.objects.all()
    elif request.user.is_trainer:
        base_qs = Import.objects.filter(centre=request.user.centre)
    else:
        base_qs = Import.objects.none()

    # ***** NEW: Apply filtering based on view context *****
    if view_context == 'display_unapproved_imports':
        data = base_qs.filter(is_approved=False, is_disposed=False)
    elif view_context == 'display_disposed_imports':
        data = base_qs.filter(is_disposed=True)
    else: # Default to 'display_approved_imports'
        data = base_qs.filter(is_approved=True, is_disposed=False)

    # ---- apply search -----------------------------------------------------------
    if search_query:
        query = (
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(category__icontains=search_query) |  # ***** ADDED: category search *****
            Q(hardware__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(assignee_first_name__icontains=search_query) |
            Q(assignee_last_name__icontains=search_query) |
            Q(assignee_email_address__icontains=search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(date__icontains=search_query) |
            Q(reason_for_update__icontains=search_query)
        )
        data = data.filter(query)

    if scope == 'page':
        paginator = Paginator(data, items_per_page)
        try:
            data = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            data = paginator.page(1)

    # ---- workbook ---------------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IT Inventory"

    # ---- headers ----------------------------------------------------------------
    headers = [
        'Centre', 'Department', 'Category', 'Hardware', 'System Model', # ***** ADDED: Category *****
        'Processor', 'RAM (GB)', 'HDD (GB)', 'Serial Number',
        'Assignee First Name', 'Assignee Last Name', 'Assignee Email',
        'Device Condition', 'Status', 'Date', 'Added By',
        'Approved By', 'Is Approved', 'Disposal Reason'
    ]
    ws.append(headers)

    # ---- style for the header ---------------------------------------------------
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- data rows --------------------------------------------------------------
    for item in data:
        row = [
            item.centre.name if item.centre else 'N/A',
            item.department.name if item.department else 'N/A',
            item.get_category_display() or 'N/A',  # ***** ADDED: Category data *****
            item.hardware or 'N/A',
            item.system_model or 'N/A',
            item.processor or 'N/A',
            item.ram_gb or 'N/A',
            item.hdd_gb or 'N/A',
            item.serial_number or 'N/A',
            item.assignee_first_name or 'N/A',
            item.assignee_last_name or 'N/A',
            item.assignee_email_address or 'N/A',
            item.device_condition or 'N/A',
            item.status or 'N/A',
            item.date.strftime('%Y-%m-%d') if item.date else 'N/A',
            item.added_by.username if item.added_by else 'N/A',
            item.approved_by.username if item.approved_by else 'N/A',
            'Yes' if item.is_approved else 'No',
            item.disposal_reason or 'N/A',
        ]
        ws.append(row)

    # ---- wrap text & auto‑adjust column widths ----------------------------------
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap_align

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # ---- response ---------------------------------------------------------------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"IT_Inventory_{'All' if scope == 'all' else 'Page'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_to_pdf(request):
    # === GET PARAMETERS ===
    scope = request.GET.get('scope', 'page')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', '1')
    items_per_page = request.GET.get('items_per_page', '10')

    # ***** NEW: Get the view context *****
    view_context = request.GET.get('view_context', 'display_approved_imports')

    # --- Pagination ---
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10
    try:
        page_number = int(page_number) if page_number else 1
    except ValueError:
        page_number = 1

    # --- Base Queryset ---
    if request.user.is_superuser:
        base_qs = Import.objects.select_related('centre', 'department', 'added_by', 'approved_by')
    elif request.user.is_trainer:
        base_qs = Import.objects.filter(centre=request.user.centre)\
                             .select_related('centre', 'department', 'added_by', 'approved_by')
    else:
        base_qs = Import.objects.none()

    # ***** NEW: Apply filtering based on view context *****
    if view_context == 'display_unapproved_imports':
        qs = base_qs.filter(is_approved=False, is_disposed=False)
    elif view_context == 'display_disposed_imports':
        qs = base_qs.filter(is_disposed=True)
    else: # Default to 'display_approved_imports'
        qs = base_qs.filter(is_approved=True, is_disposed=False)

    # --- Apply Search ---
    if search_query:
        qs = qs.filter(
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(category__icontains=search_query) |  # ***** ADDED: category search *****
            Q(hardware__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(assignee_first_name__icontains=search_query) |
            Q(assignee_last_name__icontains=search_query) |
            Q(assignee_email_address__icontains=search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(date__icontains=search_query) |
            Q(reason_for_update__icontains=search_query)
        )

    if scope == 'page':
        paginator = Paginator(qs, items_per_page)
        try:
            data = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            data = paginator.page(1)
    else:
        data = list(qs.iterator())

    # --- Response ---
    response = HttpResponse(content_type='application/pdf')
    filename = f"IT_Inventory_{'All' if scope == 'all' else 'Page'}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # --- Document ---
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=10*mm, leftMargin=10*mm,
        topMargin=15*mm, bottomMargin=18*mm
    )
    elements = []
    styles = getSampleStyleSheet()

    # --- Custom Styles ---
    styles.add(ParagraphStyle(
        name='ReportTitle', fontSize=18, leading=22,
        textColor=colors.HexColor('#143C50'), alignment=TA_CENTER, spaceAfter=6
    ))
    styles.add(ParagraphStyle(
        name='SubTitle', fontSize=11, leading=13,
        textColor=colors.HexColor('#143C50'), alignment=TA_CENTER, spaceAfter=10
    ))
    styles.add(ParagraphStyle(
        name='Cell', fontSize=7.5, leading=9,
        alignment=TA_LEFT, wordWrap='CJK'
    ))

    # --- Title ---
    title = 'MOHO IT Inventory Report' if request.user.is_superuser else f'{request.user.centre.name} IT Inventory Report'
    elements.append(Paragraph(title, styles['ReportTitle']))
    elements.append(Paragraph(
        f"Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles['SubTitle']
    ))
    elements.append(Spacer(1, 6*mm))

    # --- Table: Column Widths & Headers ***** UPDATED ***** ---
    col_widths = [65, 65, 60, 65, 80, 95, 120, 60, 80, 80]  # (pts) Added width for Category
    headers = [
        'Centre', 'Department', 'Category', 'Hardware', 'Model',
        'Specs & serial number', 'Assignee',
        'Condition', 'Status & Date', 'Disposal Reason'
    ]

    table_data = [headers]
    cell_style = styles['Cell']

    def safe(v):
        return str(v or 'N/A')

    for item in data:
        specs = (
            f"<b>RAM:</b> {safe(item.ram_gb)} GB<br/>"
            f"<b>HDD:</b> {safe(item.hdd_gb)} GB<br/>"
            f"<b>Serial:</b> {safe(item.serial_number)}"
        )
        assignee = (
            f"{safe(item.assignee_first_name)} {safe(item.assignee_last_name)}"
            f"<br/><font size=6>{safe(item.assignee_email_address)}</font>"
        )
        status_date = (
            f"<b>Status:</b> {safe(item.status)}<br/>"
            f"<b>Date:</b> {safe(item.date.strftime('%Y-%m-%d') if item.date else '')}"
        )

        row = [
            Paragraph(safe(item.centre.name if item.centre else ''), cell_style),
            Paragraph(safe(item.department.name if item.department else ''), cell_style),
            Paragraph(safe(item.get_category_display()), cell_style), # ***** ADDED: Category data *****
            Paragraph(safe(item.hardware), cell_style),
            Paragraph(safe(item.system_model), cell_style),
            Paragraph(specs, cell_style),
            Paragraph(assignee, cell_style),
            Paragraph(safe(item.device_condition), cell_style),
            Paragraph(status_date, cell_style),
            Paragraph(safe(item.disposal_reason), cell_style),
        ]
        table_data.append(row)

    if len(table_data) == 1:
        table_data.append([Paragraph('No records found.', cell_style)] * len(headers))

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#143C50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('FONTSIZE', (0,1), (-1,-1), 7.5),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,1), (-1,-1), 3),
        ('BOTTOMPADDING', (0,1), (-1,-1), 3),
    ]))
    elements.append(table)

    # --- WATERMARK + PAGE NUMBERS ---
    def add_page_elements(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        page_text = f"Page {doc.page}"
        canvas.drawCentredString(148.5 * mm, 6 * mm, page_text)
        canvas.restoreState()
        canvas.saveState()
        canvas.setFont("Helvetica", 60)
        canvas.setFillGray(0.9, 0.15)
        canvas.rotate(45)
        canvas.drawCentredString(400, 100, "MOHI IT")
        canvas.restoreState()

    # --- Build PDF ---
    try:
        doc.build(
            elements,
            onFirstPage=add_page_elements,
            onLaterPages=add_page_elements
        )
    except Exception as e:
        logger.error(f"PDF generation failed: {e}")
        return HttpResponse(f"Error generating PDF: {e}", status=500)

    return response



@login_required
def profile(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        user = request.user
        errors = []

        if not username:
            errors.append("Username is required.")
        if CustomUser.objects.exclude(id=user.id).filter(username=username).exists():
            errors.append("Username is already taken.")
        if not email:
            errors.append("Email is required.")
        if CustomUser.objects.exclude(id=user.id).filter(email=email).exists():
            errors.append("Email is already in use.")

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            user.username = username
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            user.save()
            messages.success(request, "Profile updated successfully.")
            return redirect('profile')
    return render(request, 'accounts/profile.html', {'user': request.user, 'centres': Centre.objects.all()})



@login_required
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')
        errors = []

        if not old_password or not new_password1 or not new_password2:
            errors.append("All password fields are required.")
        if new_password1 != new_password2:
            errors.append("New passwords do not match.")
        if len(new_password1) < 8:
            errors.append("New password must be at least 8 characters long.")
        if not request.user.check_password(old_password):
            errors.append("Current password is incorrect.")

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            request.user.set_password(new_password1)
            request.user.save()
            update_session_auth_hash(request, request.user)
            messages.success(request, "Password changed successfully.")
            return redirect('change_password')
    return render(request, 'accounts/change_password.html', {})

from django.db.models import Q, Count # Ensure Count is imported
from .models import Import, PendingUpdate, Centre, Department # (and other models)

@login_required
def display_approved_imports(request):
    if request.user.is_superuser:
        data = Import.objects.filter(is_approved=True, is_disposed=False)
    elif request.user.is_trainer:
        data = (Import.objects.filter(centre=request.user.centre,
                                       is_approved=True, is_disposed=False)
                if request.user.centre else Import.objects.none())
    else:
        data = Import.objects.none()

    # ---------- FILTERS ----------
    centre_filter     = request.GET.get('centre', '').strip()
    department_filter = request.GET.get('department', '').strip()
    search_query      = request.GET.get('search', '').strip()
    show_duplicates   = request.GET.get('show_duplicates', '').strip() # New filter

    if centre_filter:
        data = data.filter(centre__id=centre_filter)
    if department_filter:
        data = data.filter(department__id=department_filter)

    # --- New Duplicate Filter Logic ---
    if show_duplicates == 'on':
        # Find serial numbers that are duplicates *within the current filtered set*
        duplicate_serials = (
            data.values('serial_number')
                .annotate(serial_count=Count('serial_number'))
                .filter(serial_count__gt=1)
                .values_list('serial_number', flat=True)
        )
        # Filter the main queryset to only these serials
        data = data.filter(serial_number__in=duplicate_serials)

    if search_query:
        data = data.filter(
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(hardware__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(assignee_first_name__icontains=search_query) |
            Q(assignee_last_name__icontains=search_query) |
            Q(assignee_email_address__icontains=search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(reason_for_update__icontains=search_query)
        )

    # ---------- pagination ----------
    # (Pagination logic remains the same)
    items_per_page = request.GET.get('items_per_page', '10')
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10

    paginator = Paginator(data, items_per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)


    # ---------- pending updates ----------
    # (This logic remains the same)
    data_with_pending = []
    for item in page_obj:
        pending = PendingUpdate.objects.filter(import_record=item) \
                                       .order_by('-created_at').first()
        data_with_pending.append({'item': item, 'pending_update': pending})

    # ---------- stats ----------
    # (This logic remains the same)
    total_devices = (Import.objects.count() if request.user.is_superuser else
                   (Import.objects.filter(centre=request.user.centre).count()
                    if request.user.is_trainer and request.user.centre else 0))
    unapproved_count = (Import.objects.filter(is_approved=False, is_disposed=False).count()
                        if request.user.is_superuser else
                        (Import.objects.filter(centre=request.user.centre,
                                               is_approved=False, is_disposed=False).count()
                         if request.user.is_trainer and request.user.centre else 0))
    approved_imports = total_devices - unapproved_count


    # ---------- context ----------
    context = {
        'data_with_pending': data_with_pending,
        'paginator': paginator,
        'data': page_obj,
        'report_data': {
            'total_records': paginator.count,
            'search_query': search_query,
            'items_per_page': items_per_page,
        },
        'centres': Centre.objects.all(),
        'departments': Department.objects.all(),
        'centre_filter': centre_filter,
        'department_filter': department_filter,
        'show_duplicates': show_duplicates, # Pass new filter to context
        'items_per_page_options': [10, 25, 50, 100, 500],
        'unapproved_count': unapproved_count,
        'total_devices': total_devices,
        'approved_imports': approved_imports,
        'view_name': 'display_approved_imports',
    }
    return render(request, 'import/displaycsv_approved.html', context)


# ----------------------------------------------------------------------
#  UNAPPROVED IMPORTS (Applying the same changes)
# ----------------------------------------------------------------------
@login_required
def display_unapproved_imports(request):
    if request.user.is_superuser:
        data = Import.objects.filter(is_approved=False, is_disposed=False)
    elif request.user.is_trainer:
        data = (Import.objects.filter(centre=request.user.centre,
                                       is_approved=False, is_disposed=False)
                if request.user.centre else Import.objects.none())
    else:
        data = Import.objects.none()

    centre_filter     = request.GET.get('centre', '').strip()
    department_filter = request.GET.get('department', '').strip()
    search_query      = request.GET.get('search', '').strip()
    show_duplicates   = request.GET.get('show_duplicates', '').strip() # New filter

    if centre_filter:
        data = data.filter(centre__id=centre_filter)
    if department_filter:
        data = data.filter(department__id=department_filter)

    # --- New Duplicate Filter Logic ---
    if show_duplicates == 'on':
        duplicate_serials = (
            data.values('serial_number')
                .annotate(serial_count=Count('serial_number'))
                .filter(serial_count__gt=1)
                .values_list('serial_number', flat=True)
        )
        data = data.filter(serial_number__in=duplicate_serials)

    if search_query:
        data = data.filter(
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(hardware__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(assignee_first_name__icontains=search_query) |
            Q(assignee_last_name__icontains=search_query) |
            Q(assignee_email_address__icontains=search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(reason_for_update__icontains=search_query)
        )

    # (Pagination logic remains the same)
    items_per_page = request.GET.get('items_per_page', '10')
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10

    paginator = Paginator(data, items_per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # (Pending updates logic remains the same)
    data_with_pending = []
    for item in page_obj:
        pending = PendingUpdate.objects.filter(import_record=item) \
                                       .order_by('-created_at').first()
        data_with_pending.append({'item': item, 'pending_update': pending})

    # (Stats logic remains the same)
    total_devices = (Import.objects.count() if request.user.is_superuser else
                   (Import.objects.filter(centre=request.user.centre).count()
                    if request.user.is_trainer and request.user.centre else 0))
    unapproved_count = data.count()
    approved_imports = total_devices - unapproved_count

    context = {
        'data_with_pending': data_with_pending,
        'paginator': paginator,
        'data': page_obj,
        'report_data': {
            'total_records': paginator.count,
            'search_query': search_query,
            'items_per_page': items_per_page,
        },
        'centres': Centre.objects.all(),
        'departments': Department.objects.all(),
        'centre_filter': centre_filter,
        'department_filter': department_filter,
        'show_duplicates': show_duplicates, # Pass new filter to context
        'items_per_page_options': [10, 25, 50, 100, 500],
        'unapproved_count': unapproved_count,
        'total_devices': total_devices,
        'approved_imports': approved_imports,
        'view_name': 'display_unapproved_imports',
    }
    return render(request, 'import/displaycsv_unapproved.html', context)
# ----------------------------------------------------------------------
#  DISPOSED IMPORTS (Applying the same changes)
# ----------------------------------------------------------------------
@login_required
def display_disposed_imports(request):
    # 1. Establish the base queryset of ALL disposed devices relevant to the user
    if request.user.is_superuser:
        # Base queryset includes all disposed devices for superusers
        base_disposed_data = Import.objects.filter(is_disposed=True)
    elif request.user.is_trainer:
        # Base queryset includes only disposed devices for the user's centre
        base_disposed_data = (
            Import.objects.filter(centre=request.user.centre, is_disposed=True)
            if request.user.centre else Import.objects.none()
        )
    else:
        base_disposed_data = Import.objects.none()

    # 2. Calculate the Quick Stats based on the full base queryset (before filters/search)
    total_devices_count = base_disposed_data.count()
    approved_disposed_count = base_disposed_data.filter(is_approved=True).count()
    unapproved_disposed_count = base_disposed_data.filter(is_approved=False).count()

    # 3. Apply GET filters (Centre, Department, Search, Duplicates) to the 'data' queryset
    data = base_disposed_data  # Start applying filters/search on this queryset

    centre_filter = request.GET.get('centre', '').strip()
    department_filter = request.GET.get('department', '').strip()
    search_query = request.GET.get('search', '').strip()
    show_duplicates = request.GET.get('show_duplicates', '').strip()

    if centre_filter:
        data = data.filter(centre__id=centre_filter)
    if department_filter:
        data = data.filter(department__id=department_filter)

    # --- New Duplicate Filter Logic ---
    if show_duplicates == 'on':
        from django.db.models import Count  # Ensure this is imported at the top of views.py
        duplicate_serials = (
            data.values('serial_number')
                .annotate(serial_count=Count('serial_number'))
                .filter(serial_count__gt=1)
                .values_list('serial_number', flat=True)
        )
        data = data.filter(serial_number__in=duplicate_serials)

    if search_query:
        from django.db.models import Q  # Ensure this is imported at the top of views.py
        data = data.filter(
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(hardware__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(assignee_first_name__icontains=search_query) |
            Q(assignee_last_name__icontains=search_query) |
            Q(assignee_email_address__icontains=search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(disposal_reason__icontains=search_query)
        )

    # Pagination logic
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger  # Ensure these are imported
    items_per_page = request.GET.get('items_per_page', '10')

    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10

    paginator = Paginator(data, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Pending updates logic (if applicable)
    data_with_pending = []
    for item in page_obj:
        # If PendingUpdate is not relevant for disposed items, keep None
        data_with_pending.append({'item': item, 'pending_update': None})

    # 4. Update the Context with the Correct Stats
    context = {
        'data_with_pending': data_with_pending,
        'paginator': paginator,
        'data': page_obj,
        'report_data': {
            'total_records': paginator.count,  # This count is for the currently filtered/searched data
            'search_query': search_query,
            'items_per_page': items_per_page,
        },
        'items_per_page_options': [10, 25, 50, 100, 500],

        # CORRECTED STATS
        'unapproved_count': unapproved_disposed_count,
        'total_devices': total_devices_count,
        'approved_imports': approved_disposed_count,
        'view_name': 'display_disposed_imports',

        # Filters and dropdowns
        'centres': Centre.objects.all(),
        'departments': Department.objects.all(),
        'centre_filter': centre_filter,
        'department_filter': department_filter,
        'show_duplicates': show_duplicates,
    }

    return render(request, 'import/displaycsv_disposed.html', context)


# (Your dispose_device function remains unchanged)
@login_required
def dispose_device(request, device_id):
    device = get_object_or_404(Import, id=device_id)
    if device.is_disposed:
        messages.error(request, "This device is already disposed.")
        return redirect('display_approved_imports')
    if request.method == 'POST':
        disposal_reason = request.POST.get('disposal_reason', '').strip()
        if not disposal_reason:
            messages.error(request, "Please provide a valid reason for disposal.")
            return redirect('display_approved_imports')
        with transaction.atomic():
            device.is_disposed = True
            device.disposal_reason = disposal_reason
            device.status = 'Disposed'
            device.reason_for_update = f"Device disposed by {request.user.username}: {disposal_reason}"
            device.save()
            messages.success(request, f"Device {device.serial_number} disposed successfully.")
            return redirect('display_approved_imports')
    return render(request, 'import/dispose_device.html', {'device': device})


    
@login_required
def mark_notification_read(request, pk):
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    if request.method == 'POST':
        notification.is_read = True
        notification.save()
        messages.success(request, "Notification marked as read.")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/dashboard/'))
    return HttpResponseRedirect('/dashboard/')



def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            logger.info(f"Successful login for user: {username}")
            return redirect('dashboard')
        else:
            logger.warning(f"Failed login attempt for username: {username}")
            messages.error(request, 'Invalid username or password.')
    return render(request, 'login.html')

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('login')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def manage_users(request):
    users = CustomUser.objects.all()
    centres = Centre.objects.all()
    groups = Group.objects.all()
    permissions = Permission.objects.all()
    for user in users:
        user.stats = {
            'devices_added': user.imports_added.count(),
            'devices_approved': user.imports_approved.count() if request.user.is_superuser else 0,
            'devices_updated': user.pending_updates.count() if request.user.is_trainer else 0
        }
    return render(request, 'manage_users.html', {
        'users': users,
        'centres': centres,
        'groups': groups,
        'permissions': permissions
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def user_add(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        centre_id = request.POST.get('centre')
        is_trainer = request.POST.get('is_trainer') == 'on'
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        groups = request.POST.getlist('groups')
        errors = []

        if not username:
            errors.append("Username is required.")
        if CustomUser.objects.filter(username=username).exists():
            errors.append("Username is already taken.")
        if not email:
            errors.append("Email is required.")
        if CustomUser.objects.filter(email=email).exists():
            errors.append("Email is already in use.")
        if not password:
            errors.append("Password is required.")
        if centre_id and centre_id != '' and not Centre.objects.filter(id=centre_id).exists():
            errors.append("Invalid centre selected.")
        if is_trainer and not centre_id:
            errors.append("Centre is required for trainers.")
        if is_superuser:
            centre_id = None

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            with transaction.atomic():
                centre = Centre.objects.get(id=centre_id) if centre_id and centre_id != '' else None
                user = CustomUser.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    centre=centre,
                    is_trainer=is_trainer,
                    is_staff=is_staff,
                    is_superuser=is_superuser
                )
                if groups:
                    user.groups.set(groups)
                messages.success(request, "User added successfully.")
                return redirect('manage_users')
    return redirect('manage_users')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def user_update(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        centre_id = request.POST.get('centre')
        is_trainer = request.POST.get('is_trainer') == 'on'
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        groups = request.POST.getlist('groups')
        errors = []

        if not username:
            errors.append("Username is required.")
        if CustomUser.objects.filter(username=username).exclude(id=pk).exists():
            errors.append("Username is already taken.")
        if not email:
            errors.append("Email is required.")
        if CustomUser.objects.filter(email=email).exclude(id=pk).exists():
            errors.append("Email is already in use.")
        if centre_id and centre_id != '' and not Centre.objects.filter(id=centre_id).exists():
            errors.append("Invalid centre selected.")
        if is_trainer and not centre_id:
            errors.append("Centre is required for trainers.")
        if is_superuser:
            centre_id = None

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            with transaction.atomic():
                centre = Centre.objects.get(id=centre_id) if centre_id and centre_id != '' else None
                user.username = username
                user.email = email
                if password:
                    user.set_password(password)
                user.first_name = first_name
                user.last_name = last_name
                user.centre = centre
                user.is_trainer = is_trainer
                user.is_staff = is_staff
                user.is_superuser = is_superuser
                user.save()
                user.groups.clear()
                if groups:
                    user.groups.set(groups)
                messages.success(request, "User updated successfully.")
            return redirect('manage_users')
    return redirect('manage_users')
def _can_delete_user(user):
    """Only IT Manager or Senior IT Officer can delete users."""
    return user.is_it_manager or user.is_senior_it_officer

@login_required
@user_passes_test(_can_delete_user, login_url='manage_users')
def user_delete(request, pk):
    user_to_delete = get_object_or_404(CustomUser, pk=pk)

    if request.method == 'POST':
        if user_to_delete == request.user:
            messages.error(request, "You cannot delete your own account.")
            return redirect('manage_users')

        with transaction.atomic():
            username = user_to_delete.username
            user_to_delete.delete()
            messages.success(request, f"User '{username}' deleted successfully.")
        return redirect('manage_users')

    return redirect('manage_users')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def manage_groups(request):
    if request.method == 'POST':
        group_name = request.POST.get('group_name')
        if group_name:
            if not Group.objects.filter(name=group_name).exists():
                Group.objects.create(name=group_name)
                messages.success(request, f"Group '{group_name}' created successfully.")
            else:
                messages.error(request, "Group name already exists.")
        return redirect('manage_users')
    return render(request, 'manage_users.html', {'groups': Group.objects.all()})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_group(request):
    if request.method == 'POST':
        group_id = request.POST.get('group_id')
        group = get_object_or_404(Group, id=group_id)
        group.delete()
        messages.success(request, "Group deleted successfully.")
    return redirect('manage_users')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def update_group_permissions(request):
    if request.method == 'POST':
        group_id = request.POST.get('group_id')
        permission_ids = request.POST.getlist('permissions')
        group = get_object_or_404(Group, id=group_id)
        group.permissions.clear()
        if permission_ids:
            group.permissions.set(permission_ids)
        messages.success(request, "Permissions updated successfully.")
        return redirect('manage_users')
    return redirect('manage_users')


@login_required
def clear_user(request, device_id):
    device = get_object_or_404(Import, id=device_id)
    if request.method == 'POST':
        form = ClearanceForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                # Capture previous user details for history
                if device.assignee_first_name or device.assignee_last_name or device.assignee_email_address:
                    DeviceUserHistory.objects.create(
                        device=device,
                        assignee_first_name=device.assignee_first_name,
                        assignee_last_name=device.assignee_last_name,
                        assignee_email_address=device.assignee_email_address,
                        assigned_by=device.added_by or request.user,  # Use added_by or current user if none
                        cleared_date=timezone.now()
                    )

                clearance = form.save(commit=False)
                clearance.device = device
                clearance.cleared_by = request.user
                clearance.remarks = form.cleaned_data['remarks'] or "Device cleared"
                clearance.save()
                
                # Update device status without losing user info in history
                device.status = 'Available'
                device.department_id = 1  # Default department
                device.assignee_first_name = None
                device.assignee_last_name = None
                device.assignee_email_address = None
                device.reason_for_update = f"Device cleared by {request.user.username} at {timezone.now().date()}"
                device.save()
                
                messages.success(request, f"Device {device.serial_number} cleared successfully.")
                return redirect('display_approved_imports')
    else:
        form = ClearanceForm()
    return render(request, 'import/clear_user.html', {'form': form, 'device': device})


@login_required
def download_clearance_form(request, device_id):
    device = get_object_or_404(Import, id=device_id)
    clearance = device.clearances.order_by('-created_at').first()
    if not clearance:
        messages.error(request, "No clearance record found for this device.")
        return redirect('display_approved_imports')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="clearance_form_{device.serial_number}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    normal_style = styles['Normal']
    footer_style = ParagraphStyle(
        name='FooterStyle',
        parent=normal_style,
        fontSize=10,
        alignment=1  # Center alignment
    )
    remarks_style = ParagraphStyle(
        name='RemarksStyle',
        parent=normal_style,
        fontSize=10,
        wordWrap='CJK',  # Enables word wrapping
        leading=12,  # Line spacing
        alignment=0  # Left alignment
    )

    # Meaningful Title
    elements.append(Paragraph(f'Clearance Form for Device {device.serial_number} - MOHI IT Inventory', title_style))
    elements.append(Spacer(1, 12))

    # Device Details Table with Wrapped Remarks
    data = [
        ['Field', 'Value'],
        ['Device Serial Number', device.serial_number or 'N/A'],
        ['Hardware', device.hardware or 'N/A'],
        ['Centre', device.centre.name if device.centre else 'N/A'],
        ['Department', device.department.name if device.department else 'N/A'],
        ['Status', device.status or 'N/A'],
        ['Date', device.date.strftime("%Y-%m-%d") if device.date else 'N/A'],
        ['Cleared By', clearance.cleared_by.username],
        ['Clearance Date', clearance.created_at.strftime("%Y-%m-%d")],
        ['Approved By', device.approved_by.username if device.approved_by else 'N/A'],
    ]
    # Add Remarks as a Paragraph for wrapping
    remarks = device.reason_for_update or clearance.remarks or 'N/A'
    data.append(['Remarks', Paragraph(remarks, remarks_style)])

    table = Table(data, colWidths=[100*mm, 100*mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align content to top to handle multi-line remarks
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # User History Table
    user_history = device.user_history.all().order_by('assigned_date')
    if user_history.exists():
        history_data = [['Assignee Name', 'Email', 'Assigned By', 'Assigned Date', 'Cleared Date']]
        for history in user_history:
            assignee_name = f"{history.assignee_first_name or ''} {history.assignee_last_name or ''}".strip() or 'N/A'
            history_data.append([
                assignee_name,
                history.assignee_email_address or 'N/A',
                history.assigned_by.username if history.assigned_by else 'N/A',
                history.assigned_date.strftime("%Y-%m-%d") if history.assigned_date else 'N/A',
                history.cleared_date.strftime("%Y-%m-%d") if history.cleared_date else 'N/A',
            ])
        history_table = Table(history_data, colWidths=[40*mm, 40*mm, 40*mm, 40*mm, 40*mm])
        history_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(Paragraph('User History', normal_style))
        elements.append(Spacer(1, 6))
        elements.append(history_table)
        elements.append(Spacer(1, 12))

    # Footer with Signature Section
    elements.append(Paragraph('Signature Section', footer_style))
    elements.append(Spacer(1, 6))
    signature_data = [
        ['Cleared By Signature:', ''],
        ['Date:', ''],
        ['Approved By Name & Signature:', ''],
        ['Date:', ''],
    ]
    signature_table = Table(signature_data, colWidths=[80*mm, 120*mm])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(signature_table)

    # Watermark function with increased repetition and reduced overlap
    def add_watermark(canvas, doc):
        watermark_text = "MOHI IT"
        canvas.saveState()
        canvas.setFont("Helvetica", 20)  # Reduced font size
        canvas.setFillGray(0.95, 0.95)  # Reduced clarity
        page_width, page_height = doc.pagesize
        grid_size = 80  # Reduced grid size for more instances
        placed_positions = []  # Track placed positions to avoid overlap

        for x in range(0, int(page_width), grid_size):
            for y in range(0, int(page_height), grid_size):
                # Add random offset within grid cell
                offset_x = random.randint(-40, 40)
                offset_y = random.randint(-40, 40)
                adjusted_x = x + offset_x
                adjusted_y = y + offset_y
                # Check if position is within bounds and not too close to existing positions
                if (10 <= adjusted_x <= page_width - 10 and 
                    10 <= adjusted_y <= page_height - 10 and 
                    not any(abs(adjusted_x - px) < 50 or abs(adjusted_y - py) < 50 for px, py in placed_positions)):
                    canvas.rotate(45)  # Diagonal effect
                    canvas.drawString(adjusted_x, adjusted_y, watermark_text)
                    canvas.rotate(-45)  # Reset rotation
                    placed_positions.append((adjusted_x, adjusted_y))

        canvas.restoreState()

    # Build the PDF with watermark on all pages
    doc.build(elements, onFirstPage=add_watermark, onLaterPages=add_watermark)

    return response



HARDWARE_CATEGORIES = {
    'Laptop': Q(hardware__icontains='laptop'),
    'Monitor': Q(hardware__icontains='monitor'),
    'System Unit': Q(hardware__icontains='system unit'),
    'Printer': Q(hardware__icontains='printer'),
    'Television': Q(hardware__icontains='television'),
    'Routers/Switch/Server': (
        Q(hardware__icontains='router') |
        Q(hardware__icontains='switch') |
        Q(hardware__icontains='server')
    ),
    'Unknown': (Q(hardware__isnull=True) | Q(hardware='')),
    # You can add 'Endcomputing' if needed
}

@login_required
def dashboard_view(request):
    user = request.user
    
    # Base queries based on user role
    if user.is_superuser and not user.is_trainer:
        device_query = Import.objects.all()
        ppm_query = PPMTask.objects.all()
        # IT Ops queries for superuser
        incident_query = IncidentReport.objects.all()
        workplan_query = WorkPlan.objects.all()
        asset_query = MissionCriticalAsset.objects.all()
        backup_query = BackupRegistry.objects.all()
        user_scope = "all"
    elif user.is_trainer and user.centre:
        device_query = Import.objects.filter(centre=user.centre)
        ppm_query = PPMTask.objects.filter(device__centre=user.centre)
        # IT Ops queries for trainer (scoped)
        incident_query = IncidentReport.objects.filter(reported_by=user) # Example: only see their own
        workplan_query = WorkPlan.objects.filter(user=user) # Only see their own
        asset_query = MissionCriticalAsset.objects.all() # Assuming trainers can see all assets
        backup_query = BackupRegistry.objects.filter(centre=user.centre) # Only see their centre's backups
        user_scope = "centre"
    else:
        device_query = Import.objects.none()
        ppm_query = PPMTask.objects.none()
        # IT Ops queries for none
        incident_query = IncidentReport.objects.none()
        workplan_query = WorkPlan.objects.none()
        asset_query = MissionCriticalAsset.objects.none()
        backup_query = BackupRegistry.objects.none()
        user_scope = "none"
    
    # Get active PPM period
    active_period = PPMPeriod.objects.filter(is_active=True).first()
    
    # === DEVICE STATISTICS ===
    total_devices = device_query.count()
    approved_devices = device_query.filter(is_approved=True, is_disposed=False).count()
    pending_approvals = device_query.filter(is_approved=False, is_disposed=False).count()
    disposed_devices = device_query.filter(is_disposed=True).count()
    
    device_status_breakdown = device_query.filter(is_disposed=False).values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    all_centres = Centre.objects.all()
    devices_by_centre = []
    for centre in all_centres:
        count = device_query.filter(centre=centre, is_approved=True, is_disposed=False).count()
        if user_scope == "centre" and centre != user.centre:
            continue
        devices_by_centre.append({'centre__name': centre.name, 'count': count, 'centre_id': centre.id})
    devices_by_centre = sorted(devices_by_centre, key=lambda x: x['count'], reverse=True)
    
    hardware_counts = device_query.filter(is_approved=True, is_disposed=False).values('hardware').annotate(count=Count('id'))
    grouped_hardware = {}
    categories = {
        'laptop': 'Laptop',
        'monitor': 'Monitor',
        'system unit': 'System Unit',
        'printer': 'Printer',
        'router': 'Routers/Switch/Server',
        'switch': 'Routers/Switch/Server',
        'server': 'Routers/Switch/Server',
        'endcomputing': 'Endcomputing',
        'television': 'Television'
    }
    for item in hardware_counts:
        hw = item['hardware'].lower() if item['hardware'] else "unknown"
        key = "Unknown"
        for cat, label in categories.items():
            if cat in hw:
                key = label
                break
        grouped_hardware[key] = grouped_hardware.get(key, 0) + item['count']
    devices_by_hardware = [{'hardware': k, 'count': v} for k, v in sorted(grouped_hardware.items(), key=lambda x: x[1], reverse=True)]
    
    device_condition_breakdown = device_query.filter(is_approved=True, is_disposed=False).values(
        'device_condition'
    ).annotate(count=Count('id')).order_by('-count')
    
    thirty_days_ago = timezone.now().date() - timedelta(days=30)
    recent_devices_count = device_query.filter(date__gte=thirty_days_ago).count()
    recent_devices = device_query.order_by('-date')[:10]
    
    # === PPM STATISTICS ===
    if active_period:
        period = active_period
        is_active_period = True
    else:
        period = PPMPeriod.objects.order_by('-end_date').first()
        is_active_period = False
    
    total_ppm_tasks = 0
    devices_with_ppm = 0
    devices_without_ppm = 0
    ppm_completion_rate = 0
    ppm_status_labels = []
    ppm_status_data = []
    ppm_status_colors = []
    ppm_tasks_by_activity = []
    ppm_by_centre = []
    period_name = None
    period_id = None
    
    if period:
        period_name = period.name
        period_id = period.id
        ppm_query_period = ppm_query.filter(period=period)
        total_ppm_tasks = ppm_query_period.count()
        devices_with_ppm = ppm_query_period.values('device').distinct().count()
        devices_without_ppm = approved_devices - devices_with_ppm
        ppm_completion_rate = round((devices_with_ppm / approved_devices * 100) if approved_devices > 0 else 0, 1)
        
        if is_active_period:
            ppm_status_labels = ['PPM Done', 'PPM Not Done']
            ppm_status_data = [devices_with_ppm, devices_without_ppm]
            ppm_status_colors = ['#10B981', '#F59E0B']
        else:
            ppm_status_labels = ['PPM Done', 'PPM Overdue']
            ppm_status_data = [devices_with_ppm, devices_without_ppm]
            ppm_status_colors = ['#10B981', '#EF4444']
        
        ppm_tasks_by_activity = ppm_query_period.values(
            'activities__name'
        ).annotate(count=Count('id')).order_by('-count')
        
        ppm_by_centre = []
        for centre in all_centres:
            if user_scope == "centre" and centre != user.centre:
                continue
            centre_approved = device_query.filter(centre=centre, is_approved=True, is_disposed=False).count()
            centre_with_ppm = ppm_query_period.filter(device__centre=centre).values('device').distinct().count()
            ppm_by_centre.append({
                'device__centre__name': centre.name,
                'centre_id': centre.id,
                'total': centre_approved,
                'completed': centre_with_ppm
            })
        ppm_by_centre = sorted(ppm_by_centre, key=lambda x: x['completed'], reverse=True)
    
    overdue_ppm_tasks = ppm_query.filter(
        period__end_date__lt=timezone.now().date(),
        completed_date__isnull=True
    ).count()
    
    seven_days_ahead = timezone.now().date() + timedelta(days=7)
    tasks_due_soon = ppm_query.filter(
        period__end_date__lte=seven_days_ahead,
        period__end_date__gte=timezone.now().date(),
        completed_date__isnull=True
    ).count()
    
    recent_ppm_completions = ppm_query.filter(
        completed_date__isnull=False
    ).order_by('-completed_date')[:5]
    
    # === USER & SYSTEM STATISTICS ===
    total_users = CustomUser.objects.count() if user.is_superuser else 0
    active_users = CustomUser.objects.filter(is_active=True).count() if user.is_superuser else 0
    total_centres = Centre.objects.count() if user.is_superuser else 0
    pending_updates = PendingUpdate.objects.count() if user.is_superuser else (
        PendingUpdate.objects.filter(import_record__centre=user.centre).count() if user.centre else 0
    )
    
    # === NOTIFICATIONS ===
    notifications = Notification.objects.filter(user=user).order_by('-created_at')[:5]
    unread_count = Notification.objects.filter(user=user, is_read=False).count()
    
    # === MONTHLY TRENDS ===
    # (Your existing monthly trends logic...)
    devices_monthly = []
    ppm_completed_monthly = []
    # (Loop to populate them...)
    
    # --- NEW QUERIES: IT OPERATIONS ---
    
    # Incident Reports
    recent_incidents = incident_query.order_by('-date_of_report')[:5]
    open_incidents_count = incident_query.filter(status__in=['Open', 'In Progress']).count() # Use status field

    # Work Plans
    today = timezone.now().date()
    current_work_plan = WorkPlan.objects.filter(
        user=user, 
        week_start_date__lte=today, 
        week_end_date__gte=today
    ).first()
    
    team_work_plans = WorkPlan.objects.filter(
        week_start_date__lte=today, 
        week_end_date__gte=today
    )
    # Get staff users who *should* submit plans (e.g., not superusers, are active)
    total_staff = CustomUser.objects.filter(is_active=True, is_superuser=False, is_trainer=True).count()
    submitted_work_plans = team_work_plans.count()

    # Mission Critical Assets
    critical_assets_count = asset_query.count()
    asset_criticality_breakdown = asset_query.values('criticality_level').annotate(
        count=Count('id')
    ).order_by('criticality_level')

    # Backup Registry
    recent_backups = backup_query.order_by('-date')[:5]
    
    
    # === CONTEXT ===
    context = {
        'user_scope': user_scope,
        
        # Device stats
        'total_devices': total_devices,
        'approved_devices': approved_devices,
        'pending_approvals': pending_approvals,
        'disposed_devices': disposed_devices,
        'recent_devices_count': recent_devices_count,
        'recent_devices': recent_devices,
        'device_status_breakdown': device_status_breakdown,
        'devices_by_centre': devices_by_centre,
        'devices_by_hardware': devices_by_hardware,
        'device_condition_breakdown': device_condition_breakdown,
        
        # PPM stats
        'total_ppm_tasks': total_ppm_tasks,
        'devices_with_ppm': devices_with_ppm,
        'devices_without_ppm': devices_without_ppm,
        'overdue_ppm_tasks': overdue_ppm_tasks,
        'tasks_due_soon': tasks_due_soon,
        'ppm_completion_rate': ppm_completion_rate,
        'ppm_tasks_by_activity': ppm_tasks_by_activity,
        'ppm_by_centre': ppm_by_centre,
        'recent_ppm_completions': recent_ppm_completions,
        'ppm_status_labels': ppm_status_labels,
        'ppm_status_data': ppm_status_data,
        'ppm_status_colors': ppm_status_colors,
        'period_name': period_name,
        'period_id': period_id, # Pass the ID for filtering
        'is_active_period': is_active_period,
        
        # User & system stats
        'total_users': total_users,
        'active_users': active_users,
        'total_centres': total_centres,
        'pending_updates': pending_updates,
        
        # Notifications
        'notifications': notifications,
        'unread_count': unread_count,
        
        # Trends
        'devices_monthly': devices_monthly,
        'ppm_completed_monthly': ppm_completed_monthly,
        
        # --- NEW CONTEXT: IT OPERATIONS ---
        'recent_incidents': recent_incidents,
        'open_incidents_count': open_incidents_count,
        'current_work_plan': current_work_plan,
        'total_staff_for_work_plans': total_staff,
        'submitted_work_plans_count': submitted_work_plans,
        'critical_assets_count': critical_assets_count,
        'asset_criticality_breakdown': asset_criticality_breakdown,
        'recent_backups': recent_backups,
    }
    
    return render(request, 'index.html', context)


@login_required
def filtered_list_view(request, list_type):
    """
    A single, unified view to handle all filtered lists from the dashboard.
    Now includes:
    - Stats based on filtered results
    - Pagination
    - Advanced filters
    """
    user = request.user
    params = request.GET  # Get all URL parameters
    
    # Determine user scope
    user_scope = "none"
    if user.is_superuser:
        user_scope = "all"
    elif user.is_trainer and user.centre:
        user_scope = "centre"

    context = {
        'list_type': list_type,
        'page_title': f'Filtered List: {list_type.title()}',
        'user': user,
        'user_scope': user_scope,
        'params': params.urlencode(), # Pass all params for pagination links
        'filters': params, # Pass GET params to template for filter forms
    }
    
    # Base querysets
    qs = None
    all_centres = Centre.objects.all().order_by('name')
    all_departments = Department.objects.all().order_by('name')
    
    # 1. --- DEVICE LIST ---
    if list_type == 'devices':
        context['page_title'] = 'Filtered Devices'
        context['all_centres'] = all_centres
        context['all_departments'] = all_departments
        context['all_hardware'] = HARDWARE_CATEGORIES.keys()
        context['all_status'] = Import.objects.filter(is_approved=True).values_list('status', flat=True).distinct()
        context['all_conditions'] = Import.objects.filter(is_approved=True).values_list('device_condition', flat=True).distinct()

        if user_scope == "all":
            qs = Import.objects.all()
        elif user_scope == "centre":
            qs = Import.objects.filter(centre=user.centre)
        else:
            qs = Import.objects.none()
            
        filters = Q(is_approved=True)

        if params.get('clear'):
            pass  # No filters
        elif params.get('is_disposed'):
            filters &= Q(is_disposed=True)
        else:
            filters &= Q(is_disposed=False)

        if params.get('centre_id'):
            filters &= Q(centre_id=params.get('centre_id'))
        if params.get('department_id'):
            filters &= Q(department_id=params.get('department_id'))
        if params.get('status'):
            filters &= Q(status=params.get('status'))
        if params.get('device_condition'):
            filters &= Q(device_condition=params.get('device_condition'))
            
        if params.get('hardware'):
            hw_filter = params.get('hardware')
            if hw_filter in HARDWARE_CATEGORIES:
                filters &= HARDWARE_CATEGORIES[hw_filter]
                
        # Apply filters BEFORE calculating stats
        filtered_qs = qs.filter(filters).order_by('centre__name', 'serial_number')
        
        # Calculate stats on the filtered queryset
        context['stats'] = {
            'total': filtered_qs.count(),
            'by_status': filtered_qs.values('status').annotate(count=Count('status')).order_by('-count'),
            'by_condition': filtered_qs.values('device_condition').annotate(count=Count('device_condition')).order_by('-count'),
        }

    # 2. --- PPM LIST ---
    elif list_type == 'ppm':
        context['page_title'] = 'Filtered PPM Tasks'
        context['all_centres'] = all_centres
        context['all_periods'] = PPMPeriod.objects.all().order_by('-start_date')

        if user_scope == "all":
            qs = PPMTask.objects.all()
        elif user_scope == "centre":
            qs = PPMTask.objects.filter(device__centre=user.centre)
        else:
            qs = PPMTask.objects.none()
            
        filters = Q()

        if params.get('centre_id'):
            filters &= Q(device__centre_id=params.get('centre_id'))
        if params.get('period_id'):
            filters &= Q(period_id=params.get('period_id'))
        if params.get('activity'):
            filters &= Q(activities__name=params.get('activity'))
            
        if params.get('ppm_status') == 'done':
            filters &= Q(completed_date__isnull=False)
        elif params.get('ppm_status') == 'pending':
            filters &= Q(completed_date__isnull=True, period__is_active=True)
        elif params.get('ppm_status') == 'overdue':
            filters &= Q(completed_date__isnull=True, period__end_date__lt=timezone.now().date())
        elif params.get('ppm_status') == 'due_soon':
            seven_days_ahead = timezone.now().date() + timedelta(days=7)
            filters &= Q(completed_date__isnull=True, 
                         period__end_date__gte=timezone.now().date(),
                         period__end_date__lte=seven_days_ahead)

        filtered_qs = qs.filter(filters).distinct().order_by('period__name', 'device__serial_number')
        
        # Calculate stats
        total_tasks = filtered_qs.count()
        completed = filtered_qs.filter(completed_date__isnull=False).count()
        pending = total_tasks - completed
        
        context['stats'] = {
            'total': total_tasks,
            'completed': completed,
            'pending': pending,
        }

    # 3. --- ASSET LIST (NEW) ---
    elif list_type == 'assets':
        context['page_title'] = 'Mission Critical Assets'
        context['all_criticality'] = [c[0] for c in MissionCriticalAsset.CRITICALITY_LEVEL_CHOICES]
        context['all_departments'] = all_departments # Already defined

        qs = MissionCriticalAsset.objects.all() # Assets are not centre-scoped in the model
        filters = Q()

        if params.get('department_id'):
            filters &= Q(department_id=params.get('department_id'))
        if params.get('criticality_level'):
            filters &= Q(criticality_level=params.get('criticality_level'))
        
        filtered_qs = qs.filter(filters).order_by('name')
        
        context['stats'] = {
            'total': filtered_qs.count(),
            'by_criticality': filtered_qs.values('criticality_level').annotate(count=Count('id')).order_by(),
        }

    # 4. --- INCIDENT LIST (NEW) ---
    elif list_type == 'incidents':
        context['page_title'] = 'Incident Reports'
        context['all_statuses'] = [s[0] for s in IncidentReport.STATUS_CHOICES]

        if user_scope == "all":
            qs = IncidentReport.objects.all()
        elif user_scope == "centre":
            qs = IncidentReport.objects.filter(reported_by=user) # Only see their own
        else:
            qs = IncidentReport.objects.none()

        filters = Q()

        if params.get('incident_number'):
            filters &= Q(incident_number=params.get('incident_number'))
        if params.get('status'):
            filters &= Q(status=params.get('status'))
        
        filtered_qs = qs.filter(filters).order_by('-date_of_report')
        
        context['stats'] = {
            'total': filtered_qs.count(),
            'by_status': filtered_qs.values('status').annotate(count=Count('id')).order_by(),
        }

    # 5. --- WORKPLAN LIST (NEW) ---
    elif list_type == 'workplans':
        context['page_title'] = 'Work Plans'
        context['all_staff'] = CustomUser.objects.filter(is_active=True, is_trainer=True).order_by('username')

        if user_scope == "all":
            qs = WorkPlan.objects.all()
        elif user_scope == "centre":
            qs = WorkPlan.objects.filter(user=user) # Only see their own
        else:
            qs = WorkPlan.objects.none()
        
        filters = Q()
        
        if params.get('user_id'):
            filters &= Q(user_id=params.get('user_id'))
        
        # Add filter for current week
        if params.get('week') == 'current':
            today = timezone.now().date()
            filters &= Q(week_start_date__lte=today, week_end_date__gte=today)

        filtered_qs = qs.filter(filters).order_by('-week_start_date', 'user__username')
        
        context['stats'] = {
            'total': filtered_qs.count(),
            'users': filtered_qs.values('user__username').distinct().count()
        }

    else:
        raise Http404("Invalid list type specified.")

    # --- PAGINATION ---
    paginator = Paginator(filtered_qs, 25) # Show 25 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context['page_obj'] = page_obj
    context['total_results'] = paginator.count
    context['is_paginated'] = page_obj.has_other_pages()
    
    # Use a single, unified template
    return render(request, 'dashboard/filtering/master_list.html', context)