from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from devices.models import DeviceAgreement, Import


DEFAULT_UAF_REQUIRED_CATEGORIES = {
    "laptop",
    "smart_phone",
    "ipad",
    "tablet",
}

DEFAULT_UAF_KEYWORDS = {
    "laptop",
    "notebook",
    "macbook",
    "ipad",
    "tablet",
    "phone",
    "smartphone",
    "cellphone",
    "mobile",
}

PLACEHOLDER_SERIALS = {
    "",
    "-",
    "n/a",
    "na",
    "none",
    "null",
    "unknown",
}


HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

RED_FILL = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")


@dataclass
class AgreementSummary:
    has_any: bool = False
    has_current: bool = False
    current_issuance_signed: bool = False
    current_clearance_signed: bool = False
    current_uploaded_pdf: bool = False
    latest_issuance_signed: bool = False
    latest_clearance_signed: bool = False
    latest_is_archived: bool = False


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .")


def clean_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def is_blank(value: Any) -> bool:
    return clean_text(value) == ""


def pct_fraction(count: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return count / total


def is_assigned_device(row: dict[str, Any]) -> bool:
    if row.get("assignee_id"):
        return True
    if clean_text(row.get("assignee_cache")):
        return True
    if clean_text(row.get("assignee_first_name")):
        return True
    if clean_text(row.get("assignee_last_name")):
        return True
    if clean_text(row.get("assignee_email_address")):
        return True
    return False


def assignee_display_name(row: dict[str, Any]) -> str:
    if row.get("assignee_id"):
        first = clean_text(row.get("assignee__first_name"))
        last = clean_text(row.get("assignee__last_name"))
        full_name = " ".join(part for part in (first, last) if part).strip()
        if full_name:
            return full_name
    cache_name = clean_text(row.get("assignee_cache"))
    if cache_name:
        return cache_name

    first = clean_text(row.get("assignee_first_name"))
    last = clean_text(row.get("assignee_last_name"))
    legacy_name = " ".join(part for part in (first, last) if part).strip()
    if legacy_name:
        return legacy_name
    return "Unassigned"


def is_agreement_signed_issuance(agreement_row: dict[str, Any]) -> bool:
    return bool(
        agreement_row.get("user_signed_issuance")
        or agreement_row.get("it_approved_issuance")
        or clean_text(agreement_row.get("uploaded_uaf_pdf"))
        or agreement_row.get("issuance_date")
    )


def is_agreement_signed_clearance(agreement_row: dict[str, Any]) -> bool:
    return bool(
        agreement_row.get("user_signed_clearance")
        or agreement_row.get("it_approved_clearance")
        or agreement_row.get("clearance_date")
    )


def is_uaf_eligible(
    row: dict[str, Any],
    *,
    required_categories: set[str],
    portable_keywords: set[str],
) -> bool:
    category_norm = normalize_text(row.get("category"))
    if category_norm in required_categories:
        return True

    device_name_norm = normalize_text(row.get("device_name"))
    system_model_norm = normalize_text(row.get("system_model"))

    if device_name_norm.endswith("- l mohi") or device_name_norm.endswith("-l-mohi"):
        return True

    combined = f"{device_name_norm} {system_model_norm}".strip()
    if not combined:
        return False
    return any(keyword in combined for keyword in portable_keywords)


def is_uaf_required(
    row: dict[str, Any],
    *,
    required_categories: set[str],
    portable_keywords: set[str],
) -> bool:
    return is_assigned_device(row) and is_uaf_eligible(
        row,
        required_categories=required_categories,
        portable_keywords=portable_keywords,
    )


def build_agreement_summary_by_device(device_ids: list[int]) -> dict[int, AgreementSummary]:
    agreement_map: dict[int, AgreementSummary] = {}
    if not device_ids:
        return agreement_map

    rows = (
        DeviceAgreement.objects.filter(device_id__in=device_ids)
        .values(
            "id",
            "device_id",
            "is_archived",
            "user_signed_issuance",
            "it_approved_issuance",
            "issuance_date",
            "user_signed_clearance",
            "it_approved_clearance",
            "clearance_date",
            "uploaded_uaf_pdf",
        )
        .order_by("device_id", "-issuance_date", "-id")
    )

    seen_latest: set[int] = set()
    seen_current: set[int] = set()
    for row in rows:
        device_id = int(row["device_id"])
        summary = agreement_map.setdefault(device_id, AgreementSummary())
        summary.has_any = True

        issuance_signed = is_agreement_signed_issuance(row)
        clearance_signed = is_agreement_signed_clearance(row)

        if device_id not in seen_latest:
            summary.latest_issuance_signed = issuance_signed
            summary.latest_clearance_signed = clearance_signed
            summary.latest_is_archived = bool(row.get("is_archived"))
            seen_latest.add(device_id)

        if not row.get("is_archived") and device_id not in seen_current:
            summary.has_current = True
            summary.current_issuance_signed = issuance_signed
            summary.current_clearance_signed = clearance_signed
            summary.current_uploaded_pdf = bool(clean_text(row.get("uploaded_uaf_pdf")))
            seen_current.add(device_id)

    return agreement_map


def prepare_sheet(worksheet, headers: list[str]) -> dict[str, int]:
    worksheet.append(headers)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=index)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    return {header: index for index, header in enumerate(headers, start=1)}


def autosize_columns(worksheet, *, min_width: int = 10, max_width: int = 55) -> None:
    for column_cells in worksheet.columns:
        column_index = column_cells[0].column
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            cell.alignment = CELL_ALIGNMENT
        adjusted = min(max(max_length + 2, min_width), max_width)
        worksheet.column_dimensions[get_column_letter(column_index)].width = adjusted


def apply_percent_format(worksheet, header_map: dict[str, int], header_names: list[str]) -> None:
    for header in header_names:
        column = header_map.get(header)
        if not column:
            continue
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_idx, column=column).number_format = "0.00%"


def apply_centre_summary_highlights(worksheet, header_map: dict[str, int]) -> None:
    idx_unassigned_pct = header_map.get("Unassigned Percentage")
    idx_approved_pct = header_map.get("Approved Percentage")
    idx_uaf_compliance = header_map.get("UAF Compliance Percentage")
    idx_dup = header_map.get("Duplicate Serial Number Count")
    idx_missing_name = header_map.get("Devices Missing device_name")
    idx_missing_serial = header_map.get("Devices Missing Serial Number")

    for row_idx in range(2, worksheet.max_row + 1):
        if idx_unassigned_pct:
            cell = worksheet.cell(row=row_idx, column=idx_unassigned_pct)
            value = float(cell.value or 0)
            if value >= 0.30:
                cell.fill = RED_FILL
            elif value >= 0.15:
                cell.fill = YELLOW_FILL

        if idx_approved_pct:
            cell = worksheet.cell(row=row_idx, column=idx_approved_pct)
            value = float(cell.value or 0)
            if value >= 0.95:
                cell.fill = GREEN_FILL
            elif value < 0.80:
                cell.fill = RED_FILL
            else:
                cell.fill = YELLOW_FILL

        if idx_uaf_compliance:
            cell = worksheet.cell(row=row_idx, column=idx_uaf_compliance)
            value = float(cell.value or 0)
            if value >= 0.95:
                cell.fill = GREEN_FILL
            elif value < 0.80:
                cell.fill = RED_FILL
            elif value > 0:
                cell.fill = YELLOW_FILL

        if idx_dup:
            cell = worksheet.cell(row=row_idx, column=idx_dup)
            value = int(cell.value or 0)
            if value > 0:
                cell.fill = RED_FILL

        if idx_missing_name:
            cell = worksheet.cell(row=row_idx, column=idx_missing_name)
            if int(cell.value or 0) > 0:
                cell.fill = YELLOW_FILL

        if idx_missing_serial:
            cell = worksheet.cell(row=row_idx, column=idx_missing_serial)
            value = int(cell.value or 0)
            if value > 0:
                cell.fill = RED_FILL


def apply_uaf_compliance_highlights(worksheet, header_map: dict[str, int]) -> None:
    idx_signed = header_map.get("Whether UAF is signed")
    idx_reason = header_map.get("Reason for non-compliance")
    idx_status = header_map.get("Agreement status")

    for row_idx in range(2, worksheet.max_row + 1):
        if idx_signed:
            cell = worksheet.cell(row=row_idx, column=idx_signed)
            if cell.value is True:
                cell.fill = GREEN_FILL
            elif cell.value is False:
                cell.fill = RED_FILL

        if idx_reason:
            cell = worksheet.cell(row=row_idx, column=idx_reason)
            reason = clean_text(cell.value)
            if reason and reason.lower() != "compliant":
                cell.fill = RED_FILL
            elif reason.lower() == "compliant":
                cell.fill = GREEN_FILL

        if idx_status:
            cell = worksheet.cell(row=row_idx, column=idx_status)
            status_text = normalize_text(cell.value)
            if "pending" in status_text or "no agreement" in status_text:
                cell.fill = YELLOW_FILL


def apply_duplicate_highlights(worksheet, header_map: dict[str, int]) -> None:
    idx_type = header_map.get("Duplicate Type")
    if not idx_type:
        return
    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=idx_type)
        type_value = normalize_text(cell.value)
        if "strong" in type_value:
            cell.fill = RED_FILL
        elif "weak" in type_value:
            cell.fill = YELLOW_FILL


def make_centre_stats(centre_name: str, centre_code: str) -> dict[str, Any]:
    return {
        "centre_name": centre_name,
        "centre_code": centre_code,
        "total_devices": 0,
        "total_active_devices": 0,
        "total_disposed_devices": 0,
        "total_assigned_devices": 0,
        "total_unassigned_devices": 0,
        "approved_devices": 0,
        "unapproved_devices": 0,
        "missing_device_name": 0,
        "missing_serial_number": 0,
        "missing_system_model": 0,
        "missing_category": 0,
        "missing_department": 0,
        "duplicate_serial_number_count": 0,
        "duplicate_serial_device_count": 0,
        "total_devices_requiring_uaf": 0,
        "total_devices_with_signed_uaf": 0,
        "total_devices_without_signed_uaf": 0,
        "laptop_devices": 0,
        "laptops_assigned": 0,
        "laptops_without_signed_uaf": 0,
        "assigned_devices_without_signed_uaf": 0,
        "unassigned_laptops": 0,
        "unassigned_devices_without_device_name": 0,
    }


def agreement_status_text(
    *,
    is_assigned: bool,
    agreement_summary: AgreementSummary,
) -> str:
    if agreement_summary.has_current:
        if agreement_summary.current_issuance_signed and agreement_summary.current_clearance_signed:
            return "Current agreement: issuance+clearance signed"
        if agreement_summary.current_issuance_signed:
            return "Current agreement: issuance signed"
        return "Current agreement: issuance pending"

    if agreement_summary.has_any:
        if agreement_summary.latest_issuance_signed and agreement_summary.latest_clearance_signed:
            return "Latest agreement archived and fully signed"
        if agreement_summary.latest_issuance_signed and not agreement_summary.latest_clearance_signed:
            if is_assigned:
                return "Latest agreement signed issuance only"
            return "Latest agreement signed issuance only (clearance pending)"
        return "Agreement exists but incomplete"

    return "No agreement record"


def non_compliance_reasons(
    *,
    row: dict[str, Any],
    is_assigned: bool,
    uaf_required: bool,
    uaf_signed_effective: bool,
    uaf_eligible: bool,
    agreement_summary: AgreementSummary,
) -> list[str]:
    reasons: list[str] = []

    if uaf_required and not agreement_summary.has_current:
        reasons.append("assigned but no agreement")
    if (
        uaf_required
        and agreement_summary.has_current
        and not agreement_summary.current_issuance_signed
    ):
        reasons.append("agreement exists but issuance not signed")
    if (
        not is_assigned
        and agreement_summary.has_any
        and agreement_summary.latest_issuance_signed
        and not agreement_summary.latest_clearance_signed
    ):
        reasons.append("signed issuance but no clearance after return if relevant")
    if (
        is_assigned
        and normalize_text(row.get("category")) == "laptop"
        and not uaf_signed_effective
    ):
        reasons.append("laptop assigned without UAF")
    if is_blank(row.get("device_name")) and uaf_eligible:
        reasons.append("missing device_name so cannot classify well")
    return reasons


def build_inventory_operational_report_workbook(
    *,
    centre_codes: list[str] | None = None,
    include_no_centre: bool = True,
    required_uaf_categories: set[str] | None = None,
    uaf_keywords: set[str] | None = None,
    include_weak_duplicates: bool = True,
) -> Workbook:
    required_categories = {
        normalize_text(item) for item in (required_uaf_categories or DEFAULT_UAF_REQUIRED_CATEGORIES)
    }
    keyword_set = {normalize_text(item) for item in (uaf_keywords or DEFAULT_UAF_KEYWORDS)}
    category_display_map = dict(Import.CATEGORY_CHOICES)

    queryset = Import.objects.select_related("centre", "department", "assignee")
    if centre_codes:
        queryset = queryset.filter(centre__centre_code__in=centre_codes)
    if not include_no_centre:
        queryset = queryset.exclude(centre__isnull=True)

    import_rows = list(
        queryset.values(
            "id",
            "centre_id",
            "centre__name",
            "centre__centre_code",
            "department_id",
            "department__name",
            "category",
            "device_name",
            "system_model",
            "serial_number",
            "processor",
            "ram_gb",
            "hdd_gb",
            "assignee_id",
            "assignee__first_name",
            "assignee__last_name",
            "assignee_cache",
            "assignee_first_name",
            "assignee_last_name",
            "assignee_email_address",
            "status",
            "is_approved",
            "is_disposed",
            "uaf_signed",
        )
    )

    device_ids = [int(row["id"]) for row in import_rows]
    agreements_by_device = build_agreement_summary_by_device(device_ids)

    # Strong duplicate map: same normalized serial (excluding placeholders).
    serial_to_rows: dict[str, list[int]] = defaultdict(list)
    centre_serial_to_rows: dict[tuple[int | None, str, str], dict[str, list[int]]] = defaultdict(
        lambda: defaultdict(list)
    )

    for row in import_rows:
        serial_norm = normalize_text(row.get("serial_number"))
        if not serial_norm or serial_norm in PLACEHOLDER_SERIALS:
            continue
        row_id = int(row["id"])
        centre_key = (
            row.get("centre_id"),
            clean_text(row.get("centre__name"), "No Centre"),
            clean_text(row.get("centre__centre_code"), "NO_CENTRE"),
        )
        serial_to_rows[serial_norm].append(row_id)
        centre_serial_to_rows[centre_key][serial_norm].append(row_id)

    strong_duplicate_ids: set[int] = set()
    for serial_ids in serial_to_rows.values():
        if len(serial_ids) > 1:
            strong_duplicate_ids.update(serial_ids)

    # Optional weak map: same centre + device_name + system_model + assignee.
    weak_group_to_rows: dict[tuple[int | None, str, str, str], list[int]] = defaultdict(list)
    if include_weak_duplicates:
        for row in import_rows:
            device_name_norm = normalize_text(row.get("device_name"))
            model_norm = normalize_text(row.get("system_model"))
            assignee_norm = normalize_text(assignee_display_name(row))
            if not device_name_norm or not model_norm or not assignee_norm:
                continue
            weak_key = (
                row.get("centre_id"),
                device_name_norm,
                model_norm,
                assignee_norm,
            )
            weak_group_to_rows[weak_key].append(int(row["id"]))

    weak_duplicate_ids: set[int] = set()
    for group_ids in weak_group_to_rows.values():
        if len(group_ids) > 1:
            weak_duplicate_ids.update(group_ids)

    centre_stats: dict[tuple[int | None, str, str], dict[str, Any]] = {}
    device_name_stats: dict[tuple[str, str, str], dict[str, Any]] = {}
    missing_data_rows: list[list[Any]] = []
    duplicate_rows: list[list[Any]] = []
    uaf_compliance_rows: list[list[Any]] = []

    for row in import_rows:
        row_id = int(row["id"])
        centre_name = clean_text(row.get("centre__name"), "No Centre")
        centre_code = clean_text(row.get("centre__centre_code"), "NO_CENTRE")
        centre_key = (row.get("centre_id"), centre_name, centre_code)
        stats = centre_stats.setdefault(centre_key, make_centre_stats(centre_name, centre_code))

        is_assigned = is_assigned_device(row)
        is_disposed = bool(row.get("is_disposed"))
        is_approved = bool(row.get("is_approved"))
        category_norm = normalize_text(row.get("category"))
        device_name_display = clean_text(row.get("device_name"), "Unknown")
        category_display = category_display_map.get(
            clean_text(row.get("category")),
            clean_text(row.get("category"), "Unknown"),
        ) or "Unknown"
        assignee_name = assignee_display_name(row)
        agreement_summary = agreements_by_device.get(row_id, AgreementSummary())

        uaf_eligible = is_uaf_eligible(
            row,
            required_categories=required_categories,
            portable_keywords=keyword_set,
        )
        uaf_required = is_uaf_required(
            row,
            required_categories=required_categories,
            portable_keywords=keyword_set,
        )
        uaf_signed_effective = bool(
            row.get("uaf_signed") or (agreement_summary.has_current and agreement_summary.current_issuance_signed)
        )

        stats["total_devices"] += 1
        stats["total_active_devices"] += 0 if is_disposed else 1
        stats["total_disposed_devices"] += 1 if is_disposed else 0
        stats["total_assigned_devices"] += 1 if is_assigned else 0
        stats["total_unassigned_devices"] += 0 if is_assigned else 1
        stats["approved_devices"] += 1 if is_approved else 0
        stats["unapproved_devices"] += 0 if is_approved else 1

        stats["missing_device_name"] += 1 if is_blank(row.get("device_name")) else 0
        stats["missing_serial_number"] += 1 if is_blank(row.get("serial_number")) else 0
        stats["missing_system_model"] += 1 if is_blank(row.get("system_model")) else 0
        stats["missing_category"] += 1 if is_blank(row.get("category")) else 0
        stats["missing_department"] += 1 if row.get("department_id") is None else 0

        if normalize_text(row.get("category")) == "laptop":
            stats["laptop_devices"] += 1
            if is_assigned:
                stats["laptops_assigned"] += 1
            else:
                stats["unassigned_laptops"] += 1

        if uaf_required:
            stats["total_devices_requiring_uaf"] += 1
            if uaf_signed_effective:
                stats["total_devices_with_signed_uaf"] += 1
            else:
                stats["total_devices_without_signed_uaf"] += 1
                stats["assigned_devices_without_signed_uaf"] += 1
                if category_norm == "laptop":
                    stats["laptops_without_signed_uaf"] += 1

        if (not is_assigned) and is_blank(row.get("device_name")):
            stats["unassigned_devices_without_device_name"] += 1

        serial_norm = normalize_text(row.get("serial_number"))
        is_strong_duplicate = row_id in strong_duplicate_ids
        is_weak_duplicate = include_weak_duplicates and (row_id in weak_duplicate_ids) and not is_strong_duplicate

        # Device-name analysis grouping.
        device_name_key = (
            centre_name,
            centre_code,
            device_name_display,
        )
        group = device_name_stats.setdefault(
            device_name_key,
            {
                "centre_name": centre_name,
                "centre_code": centre_code,
                "device_name": device_name_display,
                "total_count": 0,
                "assigned_count": 0,
                "unassigned_count": 0,
                "approved_count": 0,
                "disposed_count": 0,
                "missing_serial_numbers_count": 0,
                "duplicate_serial_numbers_count": 0,
                "uaf_required_count": 0,
                "uaf_signed_count": 0,
                "uaf_unsigned_count": 0,
            },
        )
        group["total_count"] += 1
        group["assigned_count"] += 1 if is_assigned else 0
        group["unassigned_count"] += 0 if is_assigned else 1
        group["approved_count"] += 1 if is_approved else 0
        group["disposed_count"] += 1 if is_disposed else 0
        group["missing_serial_numbers_count"] += 1 if is_blank(row.get("serial_number")) else 0
        group["duplicate_serial_numbers_count"] += 1 if is_strong_duplicate else 0
        if uaf_required:
            group["uaf_required_count"] += 1
            group["uaf_signed_count"] += 1 if uaf_signed_effective else 0
            group["uaf_unsigned_count"] += 0 if uaf_signed_effective else 1

        # Missing-data detail rows.
        missing_fields: list[str] = []
        if row.get("centre_id") is None:
            missing_fields.append("centre")
        if row.get("department_id") is None:
            missing_fields.append("department")
        if is_blank(row.get("category")):
            missing_fields.append("category")
        if is_blank(row.get("device_name")):
            missing_fields.append("device_name")
        if is_blank(row.get("serial_number")):
            missing_fields.append("serial_number")
        if is_blank(row.get("system_model")):
            missing_fields.append("system_model")
        if is_blank(row.get("processor")):
            missing_fields.append("processor")
        if is_blank(row.get("ram_gb")):
            missing_fields.append("ram_gb")
        if is_blank(row.get("hdd_gb")):
            missing_fields.append("hdd_gb")

        if missing_fields:
            missing_data_rows.append(
                [
                    centre_name,
                    centre_code,
                    row_id,
                    clean_text(row.get("serial_number"), "N/A"),
                    device_name_display,
                    clean_text(row.get("system_model"), "N/A"),
                    category_display,
                    clean_text(row.get("department__name"), "N/A"),
                    assignee_name,
                    clean_text(row.get("status"), "N/A"),
                    ", ".join(missing_fields),
                ]
            )

        # Duplicate rows: strong first, then optional weak where strong is absent.
        if is_strong_duplicate:
            duplicate_rows.append(
                [
                    centre_name,
                    centre_code,
                    row_id,
                    clean_text(row.get("serial_number"), "N/A"),
                    device_name_display,
                    clean_text(row.get("system_model"), "N/A"),
                    clean_text(row.get("department__name"), "N/A"),
                    assignee_name,
                    clean_text(row.get("status"), "N/A"),
                    "Strong Duplicate (Serial)",
                    f"Same serial_number: {serial_norm}",
                ]
            )
        elif is_weak_duplicate:
            duplicate_rows.append(
                [
                    centre_name,
                    centre_code,
                    row_id,
                    clean_text(row.get("serial_number"), "N/A"),
                    device_name_display,
                    clean_text(row.get("system_model"), "N/A"),
                    clean_text(row.get("department__name"), "N/A"),
                    assignee_name,
                    clean_text(row.get("status"), "N/A"),
                    "Possible Weak Duplicate",
                    "Same centre + device_name + system_model + assignee",
                ]
            )

        # UAF compliance rows (focus on eligible user-issued devices).
        if uaf_eligible:
            reasons = non_compliance_reasons(
                row=row,
                is_assigned=is_assigned,
                uaf_required=uaf_required,
                uaf_signed_effective=uaf_signed_effective,
                uaf_eligible=uaf_eligible,
                agreement_summary=agreement_summary,
            )
            if not reasons:
                reasons = ["Compliant"] if (uaf_required and uaf_signed_effective) else ["Not currently required"]

            uaf_compliance_rows.append(
                [
                    centre_name,
                    row_id,
                    clean_text(row.get("serial_number"), "N/A"),
                    device_name_display,
                    clean_text(row.get("system_model"), "N/A"),
                    category_display,
                    assignee_name,
                    clean_text(row.get("status"), "N/A"),
                    bool(uaf_required),
                    bool(uaf_signed_effective),
                    agreement_status_text(is_assigned=is_assigned, agreement_summary=agreement_summary),
                    bool(agreement_summary.has_current),
                    "; ".join(reasons),
                ]
            )

    # Centre-level strong duplicate counts.
    for centre_key, serial_groups in centre_serial_to_rows.items():
        stats = centre_stats.setdefault(
            centre_key,
            make_centre_stats(centre_key[1], centre_key[2]),
        )
        dup_groups = 0
        dup_devices = 0
        for serial_ids in serial_groups.values():
            if len(serial_ids) > 1:
                dup_groups += 1
                dup_devices += len(serial_ids)
        stats["duplicate_serial_number_count"] = dup_groups
        stats["duplicate_serial_device_count"] = dup_devices

    workbook = Workbook()
    workbook.remove(workbook.active)

    # Sheet A: Centre Summary
    centre_summary_headers = [
        "Centre name",
        "Centre code",
        "Total devices",
        "Total active devices",
        "Total disposed devices",
        "Total assigned devices",
        "Total unassigned devices",
        "Assigned Percentage",
        "Unassigned Percentage",
        "Approved devices",
        "Unapproved devices",
        "Approved Percentage",
        "Devices Missing device_name",
        "Devices Missing Serial Number",
        "Devices Missing System Model",
        "Devices Missing Category",
        "Devices Missing Department",
        "Duplicate Serial Number Count",
        "Duplicate Serial Device Count",
        "Overall total devices requiring UAF",
        "Total devices with signed UAF",
        "Total devices without signed UAF",
        "UAF Compliance Percentage",
        "Number of laptop devices",
        "Number of laptops assigned",
        "Number of laptops without signed UAF",
        "Number of assigned devices without signed UAF",
        "Number of unassigned laptops",
        "Number of unassigned devices without device name",
    ]
    ws_summary = workbook.create_sheet("Centre Summary")
    summary_header_map = prepare_sheet(ws_summary, centre_summary_headers)

    centre_rows = sorted(centre_stats.values(), key=lambda entry: entry["centre_name"].casefold())
    for entry in centre_rows:
        total = int(entry["total_devices"])
        uaf_required_total = int(entry["total_devices_requiring_uaf"])
        ws_summary.append(
            [
                entry["centre_name"],
                entry["centre_code"],
                total,
                entry["total_active_devices"],
                entry["total_disposed_devices"],
                entry["total_assigned_devices"],
                entry["total_unassigned_devices"],
                pct_fraction(entry["total_assigned_devices"], total),
                pct_fraction(entry["total_unassigned_devices"], total),
                entry["approved_devices"],
                entry["unapproved_devices"],
                pct_fraction(entry["approved_devices"], total),
                entry["missing_device_name"],
                entry["missing_serial_number"],
                entry["missing_system_model"],
                entry["missing_category"],
                entry["missing_department"],
                entry["duplicate_serial_number_count"],
                entry["duplicate_serial_device_count"],
                uaf_required_total,
                entry["total_devices_with_signed_uaf"],
                entry["total_devices_without_signed_uaf"],
                pct_fraction(entry["total_devices_with_signed_uaf"], uaf_required_total),
                entry["laptop_devices"],
                entry["laptops_assigned"],
                entry["laptops_without_signed_uaf"],
                entry["assigned_devices_without_signed_uaf"],
                entry["unassigned_laptops"],
                entry["unassigned_devices_without_device_name"],
            ]
        )

    apply_percent_format(
        ws_summary,
        summary_header_map,
        ["Assigned Percentage", "Unassigned Percentage", "Approved Percentage", "UAF Compliance Percentage"],
    )
    apply_centre_summary_highlights(ws_summary, summary_header_map)
    autosize_columns(ws_summary)

    # Sheet B: Device Name Analysis
    device_name_headers = [
        "Centre",
        "Centre Code",
        "device_name",
        "Total Count",
        "Assigned Count",
        "Unassigned Count",
        "Approved Count",
        "Disposed Count",
        "Missing Serial Numbers Count",
        "Duplicate Serial Numbers Count",
        "UAF Required Count",
        "UAF Signed Count",
        "UAF Unsigned Count",
    ]
    ws_device_name = workbook.create_sheet("Device Name Analysis")
    prepare_sheet(ws_device_name, device_name_headers)

    for item in sorted(
        device_name_stats.values(),
        key=lambda row: (row["centre_name"].casefold(), row["device_name"].casefold()),
    ):
        ws_device_name.append(
            [
                item["centre_name"],
                item["centre_code"],
                item["device_name"],
                item["total_count"],
                item["assigned_count"],
                item["unassigned_count"],
                item["approved_count"],
                item["disposed_count"],
                item["missing_serial_numbers_count"],
                item["duplicate_serial_numbers_count"],
                item["uaf_required_count"],
                item["uaf_signed_count"],
                item["uaf_unsigned_count"],
            ]
        )
    autosize_columns(ws_device_name)

    # Sheet C: UAF Compliance
    uaf_headers = [
        "Centre",
        "Device ID",
        "Serial Number",
        "Device Name",
        "System Model",
        "Category",
        "Assignee",
        "Status",
        "Whether UAF is required",
        "Whether UAF is signed",
        "Agreement status",
        "Current agreement exists",
        "Reason for non-compliance",
    ]
    ws_uaf = workbook.create_sheet("UAF Compliance")
    uaf_header_map = prepare_sheet(ws_uaf, uaf_headers)
    for row in sorted(uaf_compliance_rows, key=lambda entry: (str(entry[0]).casefold(), int(entry[1]))):
        ws_uaf.append(row)
    apply_uaf_compliance_highlights(ws_uaf, uaf_header_map)
    autosize_columns(ws_uaf)

    # Sheet D: Missing Data
    missing_headers = [
        "Centre",
        "Centre Code",
        "Device ID",
        "Serial Number",
        "Device Name",
        "System Model",
        "Category",
        "Department",
        "Assignee",
        "Status",
        "Missing Fields",
    ]
    ws_missing = workbook.create_sheet("Missing Data")
    prepare_sheet(ws_missing, missing_headers)
    for row in sorted(missing_data_rows, key=lambda entry: (str(entry[0]).casefold(), int(entry[2]))):
        ws_missing.append(row)
        ws_missing.cell(row=ws_missing.max_row, column=len(missing_headers)).fill = RED_FILL
    autosize_columns(ws_missing)

    # Sheet E: Duplicate Devices
    duplicate_headers = [
        "Centre",
        "Centre Code",
        "Device ID",
        "Serial Number",
        "Device Name",
        "System Model",
        "Department",
        "Assignee",
        "Status",
        "Duplicate Type",
        "Duplicate reason",
    ]
    ws_dup = workbook.create_sheet("Duplicate Devices")
    dup_header_map = prepare_sheet(ws_dup, duplicate_headers)
    for row in sorted(
        duplicate_rows,
        key=lambda entry: (
            0 if "strong" in normalize_text(entry[9]) else 1,
            str(entry[0]).casefold(),
            int(entry[2]),
        ),
    ):
        ws_dup.append(row)
    apply_duplicate_highlights(ws_dup, dup_header_map)
    autosize_columns(ws_dup)

    # Sheet F: Centre Rankings
    ws_rank = workbook.create_sheet("Centre Rankings")
    ranking_headers = ["Rank", "Centre", "Centre Code", "Value", "Total Devices"]
    prepare_sheet(ws_rank, ranking_headers)

    def append_ranking_section(
        title: str,
        rows: list[dict[str, Any]],
        value_key: str,
        *,
        value_is_percent: bool = False,
    ) -> None:
        ws_rank.append(["", "", "", "", ""])
        ws_rank.append([title, "", "", "", ""])
        title_row = ws_rank.max_row
        ws_rank.cell(row=title_row, column=1).font = Font(bold=True, color="1F4E78")
        ws_rank.cell(row=title_row, column=1).alignment = Alignment(horizontal="left")
        for index, item in enumerate(rows, start=1):
            ws_rank.append(
                [
                    index,
                    item["centre_name"],
                    item["centre_code"],
                    item[value_key],
                    item["total_devices"],
                ]
            )
            if value_is_percent:
                ws_rank.cell(row=ws_rank.max_row, column=4).number_format = "0.00%"

    ranking_payload = []
    for entry in centre_rows:
        total = int(entry["total_devices"])
        uaf_required_total = int(entry["total_devices_requiring_uaf"])
        ranking_payload.append(
            {
                "centre_name": entry["centre_name"],
                "centre_code": entry["centre_code"],
                "total_devices": total,
                "unassigned_devices": entry["total_unassigned_devices"],
                "missing_device_name": entry["missing_device_name"],
                "unsigned_uaf": entry["total_devices_without_signed_uaf"],
                "uaf_compliance": pct_fraction(entry["total_devices_with_signed_uaf"], uaf_required_total),
                "duplicate_serial_count": entry["duplicate_serial_number_count"],
                "missing_serial_count": entry["missing_serial_number"],
                "approval_compliance": pct_fraction(entry["approved_devices"], total),
            }
        )

    top_n = 10
    append_ranking_section(
        "Highest Unassigned Devices",
        sorted(ranking_payload, key=lambda row: row["unassigned_devices"], reverse=True)[:top_n],
        "unassigned_devices",
    )
    append_ranking_section(
        "Highest Missing device_name",
        sorted(ranking_payload, key=lambda row: row["missing_device_name"], reverse=True)[:top_n],
        "missing_device_name",
    )
    append_ranking_section(
        "Highest Devices Without Signed UAF",
        sorted(ranking_payload, key=lambda row: row["unsigned_uaf"], reverse=True)[:top_n],
        "unsigned_uaf",
    )
    append_ranking_section(
        "Lowest UAF Compliance",
        sorted(ranking_payload, key=lambda row: row["uaf_compliance"])[:top_n],
        "uaf_compliance",
        value_is_percent=True,
    )
    append_ranking_section(
        "Highest Duplicate Serial Numbers",
        sorted(ranking_payload, key=lambda row: row["duplicate_serial_count"], reverse=True)[:top_n],
        "duplicate_serial_count",
    )
    append_ranking_section(
        "Highest Missing Serial Numbers",
        sorted(ranking_payload, key=lambda row: row["missing_serial_count"], reverse=True)[:top_n],
        "missing_serial_count",
    )
    append_ranking_section(
        "Best Approval Compliance",
        sorted(ranking_payload, key=lambda row: row["approval_compliance"], reverse=True)[:top_n],
        "approval_compliance",
        value_is_percent=True,
    )
    autosize_columns(ws_rank)

    # Optional metadata sheet.
    ws_meta = workbook.create_sheet("Report Metadata")
    meta_headers = ["Key", "Value"]
    prepare_sheet(ws_meta, meta_headers)
    ws_meta.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_meta.append(["Report Scope", "Per-centre operational inventory quality and UAF accountability"])
    ws_meta.append(["Filter Centre Codes", ", ".join(centre_codes or []) or "All centres"])
    ws_meta.append(["Include No Centre", include_no_centre])
    ws_meta.append(["UAF Required Categories", ", ".join(sorted(required_categories))])
    ws_meta.append(["UAF Keyword Heuristics", ", ".join(sorted(keyword_set))])
    ws_meta.append(
        [
            "UAF Rule Assumption",
            "UAF required = assigned device AND UAF-eligible category/keyword match.",
        ]
    )
    ws_meta.append(
        [
            "Weak Duplicate Rule",
            "Possible weak duplicates use same centre + device_name + system_model + assignee.",
        ]
    )
    ws_meta.append(
        [
            "Strong Duplicate Rule",
            "Strong duplicates use same serial_number excluding placeholders.",
        ]
    )
    autosize_columns(ws_meta)

    return workbook


def export_inventory_operational_excel_report(
    *,
    output_path: str | Path | None = None,
    centre_codes: list[str] | None = None,
    include_no_centre: bool = True,
    required_uaf_categories: set[str] | None = None,
    uaf_keywords: set[str] | None = None,
    include_weak_duplicates: bool = True,
) -> Path:
    workbook = build_inventory_operational_report_workbook(
        centre_codes=centre_codes,
        include_no_centre=include_no_centre,
        required_uaf_categories=required_uaf_categories,
        uaf_keywords=uaf_keywords,
        include_weak_duplicates=include_weak_duplicates,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = Path(output_path) if output_path else Path("reports") / f"inventory_operational_report_{timestamp}.xlsx"
    destination = destination.expanduser()
    if not destination.is_absolute():
        destination = Path.cwd() / destination
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination
