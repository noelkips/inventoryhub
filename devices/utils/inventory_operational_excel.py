from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from devices.models import CustomUser, DeviceAgreement, Import


MISSING_VALUE_TOKENS = {"", "-", "n/a", "na", "none", "null", "unknown"}

SERIAL_IGNORE_TOKENS = {"", "-", "n/a", "na", "none", "null", "unknown"}

SPEC_EXEMPT_CATEGORIES = {
    "monitor",
    "tv",
    "printer",
    "projector",
    "smart_phone",
    "desk_phone",
    "power_backup_equipment",
    "other",
}

HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

RED_FILL = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .")


def clean_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def is_missing_value(value: Any) -> bool:
    normalized = normalize_text(value)
    return normalized in MISSING_VALUE_TOKENS


def is_assigned(row: dict[str, Any]) -> bool:
    if row.get("assignee_id"):
        return True
    if clean_text(row.get("assignee_cache")):
        return True
    if clean_text(row.get("assignee_first_name")):
        return True
    if clean_text(row.get("assignee_last_name")):
        return True
    if clean_text(row.get("assignee_email_address")):
        return True
    return False


def assignee_display(row: dict[str, Any]) -> str:
    if row.get("assignee_id"):
        first = clean_text(row.get("assignee__first_name"))
        last = clean_text(row.get("assignee__last_name"))
        full = " ".join(part for part in (first, last) if part).strip()
        if full:
            return full

    cache_value = clean_text(row.get("assignee_cache"))
    if cache_value:
        return cache_value

    first = clean_text(row.get("assignee_first_name"))
    last = clean_text(row.get("assignee_last_name"))
    legacy_name = " ".join(part for part in (first, last) if part).strip()
    if legacy_name:
        return legacy_name
    return "Unassigned"


def is_uaf_required(row: dict[str, Any]) -> bool:
    """UAF is required ONLY for Laptop category devices."""
    return normalize_text(row.get("category")) == "laptop"


def category_requires_specs(row: dict[str, Any]) -> bool:
    category = normalize_text(row.get("category"))
    return category not in SPEC_EXEMPT_CATEGORIES


def is_signed_uaf(row: dict[str, Any], signed_map: dict[int, bool]) -> bool:
    device_id = int(row["id"])
    return bool(row.get("uaf_signed") or signed_map.get(device_id, False))


def collect_missing_fields(row: dict[str, Any]) -> list[str]:
    missing: list[str] = []

    # Always required for active inventory records.
    if row.get("centre_id") is None:
        missing.append("centre")
    if row.get("department_id") is None:
        missing.append("department")
    if is_missing_value(row.get("device_name")):
        missing.append("device_name")
    if is_missing_value(row.get("serial_number")):
        missing.append("serial_number")
    if is_missing_value(row.get("system_model")):
        missing.append("system_model")
    if is_missing_value(row.get("category")):
        missing.append("category")

    # Conditionally required hardware specs.
    if category_requires_specs(row):
        if is_missing_value(row.get("processor")):
            missing.append("processor")
        if is_missing_value(row.get("ram_gb")):
            missing.append("ram_gb")
        if is_missing_value(row.get("hdd_gb")):
            missing.append("hdd_gb")

    return missing


def prepare_sheet(worksheet, headers: list[str], *, header_row: int = 1) -> dict[str, int]:
    for index, header in enumerate(headers, start=1):
        worksheet.cell(row=header_row, column=index, value=header)
    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"

    for index, _header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=index)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    return {header: index for index, header in enumerate(headers, start=1)}


def autosize_columns(worksheet, *, min_width: int = 10, max_width: int = 55) -> None:
    for column_cells in worksheet.columns:
        col_index = column_cells[0].column
        longest = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            longest = max(longest, len(value))
            cell.alignment = CELL_ALIGNMENT
        worksheet.column_dimensions[get_column_letter(col_index)].width = min(
            max(longest + 2, min_width), max_width
        )


def safe_sheet_title(base_name: str, existing: set[str]) -> str:
    raw = clean_text(base_name, "Missing Data")
    raw = re.sub(r"[:\\/?*\[\]]", "-", raw)
    title = raw[:31] or "Missing Data"
    if title not in existing:
        return title

    suffix = 2
    while True:
        suffix_text = f" ({suffix})"
        max_len = 31 - len(suffix_text)
        candidate = f"{title[:max_len]}{suffix_text}"
        if candidate not in existing:
            return candidate
        suffix += 1


def percent_fraction(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return numerator / denominator


def build_signed_uaf_map(device_ids: list[int]) -> dict[int, bool]:
    signed_map: dict[int, bool] = {}
    if not device_ids:
        return signed_map

    rows = DeviceAgreement.objects.filter(device_id__in=device_ids).values(
        "device_id",
        "user_signed_issuance",
        "it_approved_issuance",
        "issuance_date",
        "uploaded_uaf_pdf",
    )
    for row in rows:
        device_id = int(row["device_id"])
        signed = bool(
            row.get("user_signed_issuance")
            or row.get("it_approved_issuance")
            or row.get("issuance_date")
            or clean_text(row.get("uploaded_uaf_pdf"))
        )
        if signed:
            signed_map[device_id] = True
    return signed_map


def compute_duplicate_serial_ids_by_centre(active_rows: list[dict[str, Any]]) -> dict[int | None, set[int]]:
    grouped: dict[tuple[int | None, str], list[int]] = defaultdict(list)
    duplicates: dict[int | None, set[int]] = defaultdict(set)

    for row in active_rows:
        serial = normalize_text(row.get("serial_number"))
        if not serial or serial in SERIAL_IGNORE_TOKENS:
            continue
        grouped[(row.get("centre_id"), serial)].append(int(row["id"]))

    for (centre_id, _serial), ids in grouped.items():
        if len(ids) > 1:
            duplicates[centre_id].update(ids)

    return duplicates


def build_inventory_operational_workbook(
    *,
    centre_codes: list[str] | None = None,
    include_no_centre: bool = True,
    ranking_limit: int = 10,
) -> Workbook:
    category_display_map = dict(Import.CATEGORY_CHOICES)
    queryset = Import.objects.select_related("centre", "department", "assignee")
    if centre_codes:
        queryset = queryset.filter(centre__centre_code__in=centre_codes)
    if not include_no_centre:
        queryset = queryset.exclude(centre__isnull=True)

    rows = list(
        queryset.values(
            "id",
            "centre_id",
            "centre__name",
            "department_id",
            "department__name",
            "device_name",
            "system_model",
            "serial_number",
            "processor",
            "ram_gb",
            "hdd_gb",
            "category",
            "assignee_id",
            "assignee__first_name",
            "assignee__last_name",
            "assignee_cache",
            "assignee_first_name",
            "assignee_last_name",
            "assignee_email_address",
            "status",
            "is_disposed",
            "uaf_signed",
            "device_condition",
            "disposal_reason",
            "date",
        )
    )

    active_rows = [row for row in rows if not row.get("is_disposed")]
    disposed_rows = [row for row in rows if row.get("is_disposed")]

    active_device_ids = [int(row["id"]) for row in active_rows]
    signed_uaf_map = build_signed_uaf_map(active_device_ids)
    duplicate_ids_by_centre = compute_duplicate_serial_ids_by_centre(active_rows)

    trainer_names_by_centre: dict[int | None, list[str]] = defaultdict(list)
    trainer_rows = CustomUser.objects.filter(is_active=True, is_trainer=True).values(
        "centre_id", "first_name", "last_name", "username"
    )
    for trainer in trainer_rows:
        first = clean_text(trainer.get("first_name"))
        last = clean_text(trainer.get("last_name"))
        full_name = " ".join(part for part in (first, last) if part).strip()
        display_name = full_name or clean_text(trainer.get("username"), "Unknown")
        trainer_names_by_centre[trainer.get("centre_id")].append(display_name)

    for centre_id, names in trainer_names_by_centre.items():
        trainer_names_by_centre[centre_id] = sorted(set(names), key=lambda n: n.casefold())

    # Centre stats and missing data rows (active inventory only).
    stats_by_centre: dict[int | None, dict[str, Any]] = {}
    missing_rows_by_centre: dict[int | None, list[list[Any]]] = defaultdict(list)

    def get_stats(row: dict[str, Any]) -> dict[str, Any]:
        centre_id = row.get("centre_id")
        if centre_id not in stats_by_centre:
            stats_by_centre[centre_id] = {
                "centre_id": centre_id,
                "centre_name": clean_text(row.get("centre__name"), "No Centre"),
                "total_devices": 0,
                "total_active_devices": 0,
                "total_assigned_devices": 0,
                "total_unassigned_devices": 0,
                "missing_device_name": 0,
                "missing_serial_number": 0,
                "missing_system_model": 0,
                "missing_category": 0,
                "missing_department": 0,
                "duplicate_serial_number_count": 0,
                "total_devices_with_signed_uaf": 0,
                "total_devices_without_signed_uaf": 0,
                "laptop_devices": 0,
                "laptops_assigned": 0,
                "laptops_without_signed_uaf": 0,
                "unassigned_laptops": 0,
            }
        return stats_by_centre[centre_id]

    for row in active_rows:
        stats = get_stats(row)
        assigned = is_assigned(row)
        missing_fields = collect_missing_fields(row)

        stats["total_devices"] += 1
        stats["total_active_devices"] += 1
        stats["total_assigned_devices"] += 1 if assigned else 0
        stats["total_unassigned_devices"] += 0 if assigned else 1

        stats["missing_device_name"] += 1 if "device_name" in missing_fields else 0
        stats["missing_serial_number"] += 1 if "serial_number" in missing_fields else 0
        stats["missing_system_model"] += 1 if "system_model" in missing_fields else 0
        stats["missing_category"] += 1 if "category" in missing_fields else 0
        stats["missing_department"] += 1 if "department" in missing_fields else 0

        if is_uaf_required(row):
            stats["laptop_devices"] += 1
            if assigned:
                stats["laptops_assigned"] += 1
            else:
                stats["unassigned_laptops"] += 1

            if is_signed_uaf(row, signed_uaf_map):
                stats["total_devices_with_signed_uaf"] += 1
            else:
                stats["total_devices_without_signed_uaf"] += 1
                stats["laptops_without_signed_uaf"] += 1

        if missing_fields:
            missing_rows_by_centre[row.get("centre_id")].append(
                [
                    int(row["id"]),
                    clean_text(row.get("serial_number"), "N/A"),
                    clean_text(row.get("device_name"), "N/A"),
                    clean_text(row.get("system_model"), "N/A"),
                    category_display_map.get(clean_text(row.get("category")), clean_text(row.get("category"), "N/A")),
                    clean_text(row.get("department__name"), "N/A"),
                    clean_text(row.get("centre__name"), "No Centre"),
                    assignee_display(row),
                    clean_text(row.get("status"), "N/A"),
                    ", ".join(missing_fields),
                ]
            )

    for centre_id, stats in stats_by_centre.items():
        stats["duplicate_serial_number_count"] = len(duplicate_ids_by_centre.get(centre_id, set()))

    workbook = Workbook()
    workbook.remove(workbook.active)

    # Sheet 1: Centre Summary (active/non-disposed only).
    summary_headers = [
        "Centre name",
        "Total devices",
        "Total active devices",
        "Total assigned devices",
        "Total unassigned devices",
        "Assigned Percentage",
        "Total devices with signed UAF",
        "Total devices without signed UAF",
        "UAF Compliance Percentage",
        "Number of laptop devices",
        "Number of laptops assigned",
        "Number of laptops without signed UAF",
        "Number of unassigned laptops",
    ]
    ws_summary = workbook.create_sheet("Centre Summary")
    summary_map = prepare_sheet(ws_summary, summary_headers)

    centre_stats_rows = sorted(stats_by_centre.values(), key=lambda entry: entry["centre_name"].casefold())
    for stats in centre_stats_rows:
        total_active = int(stats["total_active_devices"])
        laptop_total = int(stats["laptop_devices"])
        ws_summary.append(
            [
                stats["centre_name"],
                stats["total_devices"],
                stats["total_active_devices"],
                stats["total_assigned_devices"],
                stats["total_unassigned_devices"],
                percent_fraction(stats["total_assigned_devices"], total_active),
                stats["total_devices_with_signed_uaf"],
                stats["total_devices_without_signed_uaf"],
                percent_fraction(stats["total_devices_with_signed_uaf"], laptop_total),
                stats["laptop_devices"],
                stats["laptops_assigned"],
                stats["laptops_without_signed_uaf"],
                stats["unassigned_laptops"],
            ]
        )

    for header in ("Assigned Percentage", "UAF Compliance Percentage"):
        col = summary_map[header]
        for row_idx in range(2, ws_summary.max_row + 1):
            ws_summary.cell(row=row_idx, column=col).number_format = "0.00%"

    # Basic warning highlights.
    for row_idx in range(2, ws_summary.max_row + 1):
        uaf_cell = ws_summary.cell(row=row_idx, column=summary_map["UAF Compliance Percentage"])
        uaf_val = float(uaf_cell.value or 0)
        if uaf_val >= 0.95:
            uaf_cell.fill = GREEN_FILL
        elif uaf_val < 0.80:
            uaf_cell.fill = RED_FILL
        else:
            uaf_cell.fill = YELLOW_FILL

    autosize_columns(ws_summary)

    # Sheet 2: Centre Rankings.
    ws_rank = workbook.create_sheet("Centre Rankings")
    rank_headers = ["Ranking Metric", "Rank", "Centre", "Value", "Total active devices"]
    rank_map = prepare_sheet(ws_rank, rank_headers)

    def append_ranking(metric_label: str, rows: list[dict[str, Any]], value_key: str, *, percent_value: bool = False):
        for index, item in enumerate(rows, start=1):
            ws_rank.append(
                [
                    metric_label,
                    index,
                    item["centre_name"],
                    item[value_key],
                    item["total_active_devices"],
                ]
            )
            if percent_value:
                ws_rank.cell(row=ws_rank.max_row, column=rank_map["Value"]).number_format = "0.00%"

    limit = max(1, int(ranking_limit))
    append_ranking(
        "Highest unassigned devices",
        sorted(centre_stats_rows, key=lambda item: item["total_unassigned_devices"], reverse=True)[:limit],
        "total_unassigned_devices",
    )
    append_ranking(
        "Highest devices missing device_name",
        sorted(centre_stats_rows, key=lambda item: item["missing_device_name"], reverse=True)[:limit],
        "missing_device_name",
    )
    append_ranking(
        "Highest devices without signed UAF",
        sorted(centre_stats_rows, key=lambda item: item["total_devices_without_signed_uaf"], reverse=True)[:limit],
        "total_devices_without_signed_uaf",
    )

    uaf_candidates = [item for item in centre_stats_rows if int(item["laptop_devices"]) > 0]
    for item in uaf_candidates:
        item["uaf_compliance"] = percent_fraction(
            int(item["total_devices_with_signed_uaf"]), int(item["laptop_devices"])
        )
    append_ranking(
        "Lowest UAF compliance",
        sorted(uaf_candidates, key=lambda item: item["uaf_compliance"])[:limit],
        "uaf_compliance",
        percent_value=True,
    )
    append_ranking(
        "Highest duplicate serial numbers",
        sorted(centre_stats_rows, key=lambda item: item["duplicate_serial_number_count"], reverse=True)[:limit],
        "duplicate_serial_number_count",
    )
    append_ranking(
        "Highest missing serial numbers",
        sorted(centre_stats_rows, key=lambda item: item["missing_serial_number"], reverse=True)[:limit],
        "missing_serial_number",
    )
    autosize_columns(ws_rank)

    # Sheet 3+: Missing Data per centre (active/non-disposed only).
    missing_headers = [
        "Device ID",
        "Serial Number",
        "Device Name",
        "System Model",
        "Category",
        "Department",
        "Centre",
        "Assignee",
        "Status",
        "Missing Fields",
    ]
    used_sheet_names = {"Centre Summary", "Centre Rankings", "Disposed Devices"}
    for stats in centre_stats_rows:
        centre_id = stats["centre_id"]
        centre_name = stats["centre_name"]
        title = safe_sheet_title(f"Missing - {centre_name}", used_sheet_names)
        used_sheet_names.add(title)

        ws_missing = workbook.create_sheet(title)
        trainer_names = trainer_names_by_centre.get(centre_id, [])
        trainer_label = ", ".join(trainer_names) if trainer_names else "Not Assigned"

        span_end = get_column_letter(len(missing_headers))
        ws_missing.merge_cells(f"A1:{span_end}1")
        ws_missing.merge_cells(f"A2:{span_end}2")
        ws_missing["A1"] = f"Centre: {centre_name}"
        ws_missing["A2"] = f"Trainer: {trainer_label}"
        ws_missing["A1"].font = Font(bold=True, size=14, color="1F4E78")
        ws_missing["A2"].font = Font(bold=True, size=12)
        ws_missing["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws_missing["A2"].alignment = Alignment(horizontal="left", vertical="center")

        missing_map = prepare_sheet(ws_missing, missing_headers, header_row=4)
        rows_for_centre = missing_rows_by_centre.get(centre_id, [])
        for values in rows_for_centre:
            ws_missing.append(values)
            ws_missing.cell(row=ws_missing.max_row, column=missing_map["Missing Fields"]).fill = RED_FILL
        autosize_columns(ws_missing)

    # Final sheet: Disposed Devices listing only.
    disposed_headers = [
        "Centre",
        "Device ID",
        "Serial Number",
        "Device Name",
        "System Model",
        "Category",
        "Department",
        "Assignee",
        "Status",
        "Device Condition",
        "Disposal Reason",
        "Disposal Date (if available)",
    ]
    ws_disposed = workbook.create_sheet("Disposed Devices")
    prepare_sheet(ws_disposed, disposed_headers)
    for row in sorted(disposed_rows, key=lambda item: (clean_text(item.get("centre__name"), "No Centre"), int(item["id"]))):
        ws_disposed.append(
            [
                clean_text(row.get("centre__name"), "No Centre"),
                int(row["id"]),
                clean_text(row.get("serial_number"), "N/A"),
                clean_text(row.get("device_name"), "N/A"),
                clean_text(row.get("system_model"), "N/A"),
                category_display_map.get(clean_text(row.get("category")), clean_text(row.get("category"), "N/A")),
                clean_text(row.get("department__name"), "N/A"),
                assignee_display(row),
                clean_text(row.get("status"), "N/A"),
                clean_text(row.get("device_condition"), "N/A"),
                clean_text(row.get("disposal_reason"), "N/A"),
                row.get("date") or "",
            ]
        )
    autosize_columns(ws_disposed)

    return workbook


def export_inventory_operational_excel(
    *,
    output_path: str | Path | None = None,
    centre_codes: list[str] | None = None,
    include_no_centre: bool = True,
    ranking_limit: int = 10,
) -> Path:
    workbook = build_inventory_operational_workbook(
        centre_codes=centre_codes,
        include_no_centre=include_no_centre,
        ranking_limit=ranking_limit,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = Path(output_path) if output_path else Path("reports") / f"inventory_operational_report_{timestamp}.xlsx"
    destination = destination.expanduser()
    if not destination.is_absolute():
        destination = Path.cwd() / destination
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination
