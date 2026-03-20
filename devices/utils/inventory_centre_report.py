from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.db import connections

from devices.models import Import


HEADERS = [
    "Centre",
    "Department",
    "Category",
    "device_name",
    "System Model",
    "Serial Number",
    "Device Condition",
]

COLUMN_WIDTHS = {
    "A": 13,
    "B": 15,
    "C": 20,
    "D": 17,
    "E": 29,
    "F": 26,
    "G": 19,
}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def _clean_text(value: object, fallback: str = "N/A") -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    return text or fallback


def _sort_text(value: object) -> str:
    return _clean_text(value, fallback="").casefold()


def _centre_name(device: Import) -> str:
    return _clean_text(device.centre.name if device.centre else None, fallback="No Centre")


def _department_name(device: Import) -> str:
    return _clean_text(device.department.name if device.department else None)


def _category_name(device: Import) -> str:
    if device.category:
        return _clean_text(device.get_category_display())
    return "N/A"


def _device_sort_key(device: Import) -> tuple[object, ...]:
    centre_name = device.centre.name if device.centre else None
    department_name = device.department.name if device.department else None
    return (
        centre_name is None,
        _sort_text(centre_name),
        department_name is None,
        _sort_text(department_name),
        _sort_text(device.get_category_display() if device.category else ""),
        _sort_text(device.device_name),
        _sort_text(device.serial_number),
        device.pk or 0,
    )


def get_inventory_devices(include_disposed: bool = False, include_unapproved: bool = False) -> list[Import]:
    connection = connections["default"]
    if connection.settings_dict.get("ENGINE", "").endswith("sqlite3"):
        connection.settings_dict.setdefault("OPTIONS", {})["timeout"] = 30

    queryset = Import.objects.select_related("centre", "department")

    # Follow the same default export scope used on the device pages:
    # approved inventory that has not been disposed.
    if not include_disposed:
        queryset = queryset.filter(is_disposed=False)
    if not include_unapproved:
        queryset = queryset.filter(is_approved=True)

    return sorted(queryset, key=_device_sort_key)


def _merge_centre_cells(worksheet, start_row: int, end_row: int) -> None:
    if start_row < end_row:
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    worksheet.cell(row=start_row, column=1).alignment = BODY_ALIGNMENT


def build_inventory_workbook(devices: list[Import], merge_centres: bool = False) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "IT Inventory"
    worksheet.freeze_panes = "A2"

    worksheet.append(HEADERS)

    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    for column_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width

    current_centre = None
    centre_start_row = 2

    for row_number, device in enumerate(devices, start=2):
        centre_name = _centre_name(device)

        if merge_centres:
            if current_centre is None:
                current_centre = centre_name
                centre_start_row = row_number
            elif centre_name != current_centre:
                _merge_centre_cells(worksheet, centre_start_row, row_number - 1)
                current_centre = centre_name
                centre_start_row = row_number

        worksheet.append(
            [
                centre_name,
                _department_name(device),
                _category_name(device),
                _clean_text(device.device_name),
                _clean_text(device.system_model),
                _clean_text(device.serial_number),
                _clean_text(device.device_condition),
            ]
        )

        for cell in worksheet[row_number]:
            cell.alignment = BODY_ALIGNMENT

    if merge_centres and devices:
        _merge_centre_cells(worksheet, centre_start_row, worksheet.max_row)

    return workbook


def export_inventory_workbook(
    output_path: str | Path,
    include_disposed: bool = False,
    include_unapproved: bool = False,
    merge_centres: bool = False,
) -> tuple[Path, int]:
    devices = get_inventory_devices(
        include_disposed=include_disposed,
        include_unapproved=include_unapproved,
    )
    workbook = build_inventory_workbook(devices, merge_centres=merge_centres)

    destination = Path(output_path).expanduser()
    if not destination.is_absolute():
        destination = Path.cwd() / destination

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination, len(devices)
