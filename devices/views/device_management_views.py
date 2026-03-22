from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Q, F, Case, When, IntegerField, Count, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseRedirect
from django.template.loader import get_template
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta, datetime
from io import TextIOWrapper
from decimal import Decimal, InvalidOperation

# Models
from devices.models import (
    CustomUser,
    DeviceAgreement,
    DeviceDeletionRequest,
    DeviceLog,
    DeviceRepair,
    DeviceUserHistory,
    Employee,
    Import,
    Centre,
    Notification,
    PendingUpdate,
    Department,
    DeviceConfigurationType,
    DeviceConfiguration,
)
from devices.utils.devices_utils import generate_pdf_buffer
from devices.utils.emails import send_custom_email, send_custom_email, send_device_assignment_email
from devices.utils.device_access import (
    assignment_employee_queryset,
    can_access_inventory_lists,
    can_clear_device_users,
    can_delete_devices,
    can_manage_device_assignments,
    can_request_device_deletion,
    can_review_device_requests,
)
from it_operations.models import BackupRegistry, WorkPlan, IncidentReport, MissionCriticalAsset, WorkPlanTask
from devices.forms import ClearanceForm
from ppm.models import PPMTask, PPMPeriod, PPMActivity

# Third-party & Standard Library
import csv
import logging
import re
from io import BytesIO

# Excel (openpyxl)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# PDF (ReportLab)
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    Frame, PageTemplate
)

# Django Shortcuts and HTTP
from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404

# Logging
logger = logging.getLogger(__name__)
from django.http import JsonResponse
from django.views.decorators.http import require_POST


ASSET_TAG_RE = re.compile(r"^\s*(\d+)-([LDS])-\s*MOHI\s*$", re.IGNORECASE)


def _build_assignee_search_query(search_query):
    search_query = (search_query or "").strip()
    if not search_query:
        return Q()

    assignee_query = (
        Q(assignee__first_name__icontains=search_query) |
        Q(assignee__last_name__icontains=search_query) |
        Q(assignee__email__icontains=search_query) |
        Q(assignee__staff_number__icontains=search_query) |
        Q(assignee_cache__icontains=search_query) |
        Q(assignee_first_name__icontains=search_query) |
        Q(assignee_last_name__icontains=search_query) |
        Q(assignee_email_address__icontains=search_query)
    )

    terms = [term for term in re.split(r"\s+", search_query) if term]
    if len(terms) > 1:
        tokenized_name_query = Q()
        for term in terms:
            tokenized_name_query &= (
                Q(assignee__first_name__icontains=term) |
                Q(assignee__last_name__icontains=term) |
                Q(assignee__email__icontains=term) |
                Q(assignee__staff_number__icontains=term) |
                Q(assignee_cache__icontains=term) |
                Q(assignee_first_name__icontains=term) |
                Q(assignee_last_name__icontains=term) |
                Q(assignee_email_address__icontains=term)
            )
        assignee_query |= tokenized_name_query

    return assignee_query
APPROVAL_FIELD_LABELS = (
    ("category", "Category"),
    ("centre", "Centre"),
    ("department", "Department"),
    ("device_name", "Device Name"),
    ("system_model", "System Model"),
    ("processor", "Processor"),
    ("ram_gb", "RAM (GB)"),
    ("hdd_gb", "HDD/SSD (GB)"),
    ("serial_number", "Serial Number"),
    ("assignee", "Assignee"),
    ("device_condition", "Condition"),
    ("status", "Status"),
    ("date", "Date"),
)
CATEGORY_LABELS = dict(Import.CATEGORY_CHOICES)


def _serialize_device_value(field_name, value):
    if field_name == "category":
        return CATEGORY_LABELS.get(value, value or "N/A")
    if field_name == "centre":
        return getattr(value, "name", None) or "N/A"
    if field_name == "department":
        return getattr(value, "name", None) or "N/A"
    if field_name == "assignee":
        if value:
            parts = [value.full_name]
            if value.staff_number:
                parts.append(f"ID: {value.staff_number}")
            if value.email:
                parts.append(value.email)
            return " | ".join(parts)
        return "Unassigned"
    if field_name == "date":
        return value.strftime("%Y-%m-%d") if value else "N/A"
    return value or "N/A"


def _approval_field_value(source, field_name):
    if field_name == "assignee":
        assignee = getattr(source, "assignee", None)
        if assignee:
            return _serialize_device_value(field_name, assignee)

        legacy_parts = [
            (getattr(source, "assignee_first_name", "") or "").strip(),
            (getattr(source, "assignee_last_name", "") or "").strip(),
        ]
        legacy_name = " ".join(part for part in legacy_parts if part).strip()
        legacy_email = (getattr(source, "assignee_email_address", "") or "").strip()
        if legacy_name and legacy_email:
            return f"{legacy_name} | {legacy_email}"
        if legacy_name:
            return legacy_name
        return "Unassigned"

    return _serialize_device_value(field_name, getattr(source, field_name, None))


def _build_approval_preview(device, pending_update=None):
    rows = []
    changed_count = 0
    proposed_source = pending_update or device

    for field_name, label in APPROVAL_FIELD_LABELS:
        current_value = _approval_field_value(device, field_name)
        proposed_value = _approval_field_value(proposed_source, field_name)
        is_changed = bool(pending_update and current_value != proposed_value)
        if is_changed:
            changed_count += 1
        rows.append({
            "label": label,
            "current": current_value,
            "proposed": proposed_value,
            "is_changed": is_changed,
        })

    return {
        "rows": rows,
        "changed_count": changed_count,
        "has_changes": changed_count > 0,
        "is_update_request": pending_update is not None,
    }


def _validate_device_identity(*, serial_number, device_name, exclude_pk=None):
    normalized_serial = (serial_number or "").strip()
    normalized_device_name = (device_name or "").strip()

    if not normalized_serial:
        raise ValueError("Serial number is required.")
    if not normalized_device_name:
        raise ValueError("Device name is required.")

    serial_qs = Import.objects.filter(serial_number__iexact=normalized_serial)
    if exclude_pk is not None:
        serial_qs = serial_qs.exclude(pk=exclude_pk)
    if serial_qs.exists():
        raise ValueError(f"Serial number {normalized_serial} already exists.")

    return normalized_serial, normalized_device_name


def _can_trainer_reassign_device(user, device):
    return bool(
        getattr(user, "is_trainer", False)
        and getattr(user, "centre_id", None)
        and getattr(device, "centre_id", None) == getattr(user, "centre_id", None)
    )


def _sync_device_assignment_agreement(device, *, old_assignee, new_assignee, actor):
    if old_assignee == new_assignee:
        return

    if old_assignee:
        DeviceAgreement.objects.filter(
            device=device,
            employee=old_assignee,
            is_archived=False,
        ).update(is_archived=True)

    if new_assignee:
        DeviceAgreement.objects.get_or_create(
            device=device,
            employee=new_assignee,
            is_archived=False,
            defaults={"issuance_it_user": actor},
        )

    if device.uaf_signed:
        device.uaf_signed = False
        device.save(update_fields=["uaf_signed"])


def _device_asset_kind_from_record(device: Import):
    if (device.device_name or "").upper().endswith("-L-MOHI") or device.category == "laptop":
        return "laptop"
    if (device.device_name or "").upper().endswith("-S-MOHI"):
        return "server"
    if (device.device_name or "").upper().endswith("-D-MOHI") or device.category == "system_unit":
        if (device.system_model or "").lower().find("server") != -1:
            return "server"
        return "desktop"
    return None


def _asset_letter_for_kind(kind: str):
    return {"laptop": "L", "desktop": "D", "server": "S"}.get(kind)


def _next_asset_tag(kind: str):
    letter = _asset_letter_for_kind(kind)
    if not letter:
        return None

    existing = Import.objects.filter(device_name__iendswith=f"-{letter}-MOHI").values_list("device_name", flat=True)
    max_num = 0
    for name in existing:
        if not name:
            continue
        match = ASSET_TAG_RE.match(str(name))
        if not match:
            continue
        num_str, found_letter = match.group(1), match.group(2).upper()
        if found_letter != letter:
            continue
        try:
            max_num = max(max_num, int(num_str))
        except ValueError:
            continue

    # Laptop sequence must continue from 0565-L-MOHI -> next is 0566-L-MOHI (at minimum)
    if kind == "laptop":
        max_num = max(max_num, 565)

    next_num = max_num + 1
    # Always format as 4 digits (matches 0565-L-MOHI)
    while True:
        candidate = f"{next_num:04d}-{letter}-MOHI"
        if not Import.objects.filter(device_name__iexact=candidate).exists():
            return candidate
        next_num += 1


def _ensure_device_configuration_items(device: Import):
    kind = _device_asset_kind_from_record(device)
    if kind not in {"laptop", "desktop", "server"}:
        return

    types = DeviceConfigurationType.objects.filter(is_active=True).order_by("sort_order", "name")
    if kind == "laptop":
        types = types.filter(applies_to_laptop=True)
    elif kind == "desktop":
        types = types.filter(applies_to_desktop=True)
    else:
        types = types.filter(applies_to_server=True)

    DeviceConfiguration.objects.bulk_create(
        [DeviceConfiguration(device=device, config_type=t) for t in types],
        ignore_conflicts=True,
    )


def _block_actions_if_inactive(request, device: Import):
    if getattr(device, "is_active", True):
        return None
    messages.warning(request, "This device is currently under repair and is inactive. Actions are blocked until repairs are completed.")
    return redirect("device_detail", pk=device.pk)


def _notify_device_request_reviewers(*, device, message, related_object=None):
    reviewers = CustomUser.objects.filter(
        is_active=True,
        is_trainer=False,
    ).filter(
        Q(is_superuser=True) | Q(is_it_manager=True) | Q(is_senior_it_officer=True) | Q(is_staff=True)
    ).distinct()
    target_object = related_object or device
    content_type = ContentType.objects.get_for_model(target_object)
    for reviewer in reviewers:
        notification = Notification.objects.filter(
            user=reviewer,
            content_type=content_type,
            object_id=target_object.pk,
            is_read=False,
        ).first()
        if notification:
            notification.message = message
            notification.responded_by = None
            notification.save(update_fields=["message", "responded_by"])
            continue
        Notification.objects.create(
            user=reviewer,
            message=message,
            content_type=content_type,
            object_id=target_object.pk,
        )


def _perform_device_delete(*, import_instance, actor):
    serial_number = import_instance.serial_number
    device_id = import_instance.pk
    delete_request_id = (
        DeviceDeletionRequest.objects.filter(device_id=device_id)
        .values_list("pk", flat=True)
        .first()
    )
    device_summary = {
        'serial_number': import_instance.serial_number,
        'device_name': import_instance.device_name or 'N/A',
        'system_model': import_instance.system_model or 'N/A',
        'category': import_instance.get_category_display(),
        'centre': import_instance.centre.name if import_instance.centre else 'N/A',
        'department': import_instance.department.name if import_instance.department else 'N/A',
        'assignee': import_instance.assignee.full_name if import_instance.assignee else 'N/A',
        'status': import_instance.status or 'N/A',
        'date_added': import_instance.date.strftime('%Y-%m-%d') if import_instance.date else 'N/A',
    }

    device_content_type = ContentType.objects.get_for_model(Import)
    delete_request_content_type = ContentType.objects.get_for_model(DeviceDeletionRequest)
    notification_query = Q(content_type=device_content_type, object_id=device_id)
    if delete_request_id:
        notification_query |= Q(content_type=delete_request_content_type, object_id=delete_request_id)
    Notification.objects.filter(
        is_read=False,
    ).filter(notification_query).update(is_read=True, responded_by=actor)

    import_instance.delete()

    it_email = "it@mohiafrica.org"
    subject = f"Device Deleted: {serial_number}"
    message = f"""
Device Deleted Summary

Serial Number:      {device_summary['serial_number']}
Device Name:        {device_summary['device_name']}
Model:              {device_summary['system_model']}
Category:           {device_summary['category']}
Centre:             {device_summary['centre']}
Department:         {device_summary['department']}
Assignee:           {device_summary['assignee']}
Status:             {device_summary['status']}
Date Added:         {device_summary['date_added']}

Deleted by:         {actor.get_full_name() or actor.username} ({actor.email or 'N/A'})
Deleted on:         {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}

This device has been permanently removed from the inventory.
    """

    try:
        send_custom_email(
            subject=subject,
            message=message,
            recipient_list=[it_email]
        )
    except Exception as e:
        logger.error(f"Failed to send deletion email for {serial_number}: {str(e)}")

    return serial_number


@login_required
def import_delete(request, pk):
    import_instance = get_object_or_404(Import, pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            serial_number = import_instance.serial_number
            device_summary = {
                'serial_number': import_instance.serial_number,
                'device_name': import_instance.device_name or 'N/A',
                'system_model': import_instance.system_model or 'N/A',
                'category': import_instance.get_category_display(),
                'centre': import_instance.centre.name if import_instance.centre else 'N/A',
                'department': import_instance.department.name if import_instance.department else 'N/A',
                'assignee': import_instance.assignee.full_name if import_instance.assignee else 'N/A',
                'status': import_instance.status or 'N/A',
                'date_added': import_instance.date.strftime('%Y-%m-%d') if import_instance.date else 'N/A',
            }

            # Delete the device
            import_instance.delete()

            # Mark related notifications as read
            content_type = ContentType.objects.get_for_model(Import)
            Notification.objects.filter(
                content_type=content_type,
                object_id=pk,
                is_read=False
            ).update(is_read=True, responded_by=request.user)

            # Send email to IT about deletion
            it_email = "it@mohiafrica.org"
            subject = f"Device Deleted: {serial_number}"
            message = f"""
Device Deleted Summary

Serial Number:      {device_summary['serial_number']}
Device Name:        {device_summary['device_name']}
Model:              {device_summary['system_model']}
Category:           {device_summary['category']}
Centre:             {device_summary['centre']}
Department:         {device_summary['department']}
Assignee:           {device_summary['assignee']}
Status:             {device_summary['status']}
Date Added:         {device_summary['date_added']}

Deleted by:         {request.user.get_full_name() or request.user.username} ({request.user.email or 'N/A'})
Deleted on:         {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}

This device has been permanently removed from the inventory.
            """

            try:
                send_custom_email(
                    subject=subject,
                    message=message,
                    recipient_list=[it_email]
                )
            except Exception as e:
                logger.error(f"Failed to send deletion email for {serial_number}: {str(e)}")

            messages.success(request, f"Device {serial_number} deleted successfully. IT notified.")
        return redirect('display_approved_imports')

    # GET → redirect (no confirmation page shown)
    return redirect('display_approved_imports')


@require_POST
@login_required
def check_serial(request):
    serial = request.POST.get('serial_number', '').strip()
    pk = request.POST.get('pk', '')
    
    exists = Import.objects.filter(serial_number=serial).exclude(pk=pk).exists()
    return JsonResponse({
        'exists': exists,
        'pk': pk if pk else None
    })


@login_required
def start_device_configuration(request, pk):
    device = get_object_or_404(Import, pk=pk)
    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only configure devices from your own centre.")
        return redirect('display_approved_imports')

    blocked = _block_actions_if_inactive(request, device)
    if blocked:
        return blocked

    kind = _device_asset_kind_from_record(device)
    if kind not in {"laptop", "desktop", "server"}:
        messages.warning(request, "Configuration checklist is only available for Laptop/Desktop/Server devices.")
        return redirect("device_detail", pk=device.pk)

    _ensure_device_configuration_items(device)
    return redirect('device_configuration', pk=device.pk)


@login_required
def device_configuration(request, pk):
    device = get_object_or_404(Import, pk=pk)
    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only configure devices from your own centre.")
        return redirect('display_approved_imports')

    blocked = _block_actions_if_inactive(request, device)
    if blocked:
        return blocked

    kind = _device_asset_kind_from_record(device)
    if kind not in {"laptop", "desktop", "server"}:
        messages.warning(request, "Configuration checklist is only available for Laptop/Desktop/Server devices.")
        return redirect("device_detail", pk=device.pk)

    _ensure_device_configuration_items(device)
    configs = DeviceConfiguration.objects.filter(device=device).select_related('config_type', 'completed_by')

    if request.method == 'POST':
        selected = set(request.POST.getlist('completed'))
        now = timezone.now()
        for cfg in configs:
            want_completed = str(cfg.pk) in selected
            if want_completed and not cfg.is_completed:
                cfg.is_completed = True
                cfg.completed_by = request.user
                cfg.completed_at = now
                cfg.save(update_fields=['is_completed', 'completed_by', 'completed_at'])
            elif (not want_completed) and cfg.is_completed:
                cfg.is_completed = False
                cfg.completed_by = None
                cfg.completed_at = None
                cfg.save(update_fields=['is_completed', 'completed_by', 'completed_at'])

        messages.success(request, "Configuration checklist updated.")
        return redirect('device_detail', pk=device.pk)

    total = configs.count()
    completed = configs.filter(is_completed=True).count()
    context = {
        'device': device,
        'configs': configs,
        'total': total,
        'completed': completed,
        'done': (total > 0 and completed == total),
    }
    return render(request, 'import/device_configuration.html', context)


def _is_superuser(user):
    return user.is_superuser


@login_required
@user_passes_test(_is_superuser)
def manage_configuration_types(request):
    types = DeviceConfigurationType.objects.all().order_by('sort_order', 'name')
    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, "Name is required.")
            return redirect('manage_configuration_types')

        description = (request.POST.get('description') or '').strip()
        applies_to_laptop = bool(request.POST.get('applies_to_laptop'))
        applies_to_desktop = bool(request.POST.get('applies_to_desktop'))
        applies_to_server = bool(request.POST.get('applies_to_server'))
        is_active = 'is_active' in request.POST
        try:
            sort_order = int(request.POST.get('sort_order') or 0)
        except ValueError:
            sort_order = 0

        DeviceConfigurationType.objects.create(
            name=name,
            description=description or None,
            sort_order=sort_order,
            is_active=is_active,
            applies_to_laptop=applies_to_laptop,
            applies_to_desktop=applies_to_desktop,
            applies_to_server=applies_to_server,
        )
        messages.success(request, "Configuration type added.")
        return redirect('manage_configuration_types')

    return render(request, 'configuration/manage_types.html', {'types': types})


@login_required
@user_passes_test(_is_superuser)
def configuration_type_edit(request, type_id):
    t = get_object_or_404(DeviceConfigurationType, id=type_id)
    if request.method == 'POST':
        t.name = (request.POST.get('name') or '').strip()
        t.description = (request.POST.get('description') or '').strip() or None
        try:
            t.sort_order = int(request.POST.get('sort_order') or 0)
        except ValueError:
            t.sort_order = 0
        t.is_active = bool(request.POST.get('is_active'))
        t.applies_to_laptop = bool(request.POST.get('applies_to_laptop'))
        t.applies_to_desktop = bool(request.POST.get('applies_to_desktop'))
        t.applies_to_server = bool(request.POST.get('applies_to_server'))
        t.save()
        messages.success(request, "Configuration type updated.")
        return redirect('manage_configuration_types')
    return render(request, 'configuration/type_edit.html', {'t': t})


@login_required
@user_passes_test(_is_superuser)
@require_POST
def configuration_type_delete(request, type_id):
    t = get_object_or_404(DeviceConfigurationType, id=type_id)
    t.delete()
    messages.success(request, "Configuration type deleted.")
    return redirect('manage_configuration_types')

@login_required
def get_list_context(request, initial_queryset, view_name, is_disposed=False):
    data = initial_queryset.select_related('centre', 'department', 'assignee')

    # Filters
    centre_filter = request.GET.get('centre', '').strip()
    department_filter = request.GET.get('department', '').strip()
    search_query = request.GET.get('search', '').strip()
    show_duplicates = request.GET.get('show_duplicates', '').strip()

    if centre_filter:
        data = data.filter(centre__id=centre_filter)
    if department_filter:
        data = data.filter(department__id=department_filter)

    if show_duplicates == 'on':
        duplicate_serials = (
            data.values('serial_number')
            .annotate(serial_count=Count('serial_number'))
            .filter(serial_count__gt=1)
            .values_list('serial_number', flat=True)
        )
        data = data.filter(serial_number__in=duplicate_serials)

    if search_query:
        # Build search query - common fields for all views
        search_filter = (
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(device_name__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            _build_assignee_search_query(search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(reason_for_update__icontains=search_query) |
            Q(deletion_request__reason__icontains=search_query) |
            Q(deletion_request__requested_by__username__icontains=search_query) |
            Q(deletion_request__requested_by__first_name__icontains=search_query) |
            Q(deletion_request__requested_by__last_name__icontains=search_query)
        )
        
        # Add disposal_reason only for disposed view
        if is_disposed:
            search_filter |= Q(disposal_reason__icontains=search_query)
        
        data = data.filter(search_filter)

    data = data.order_by('-pk')

    # Pagination
    items_per_page = request.GET.get('items_per_page', '10')
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10

    paginator = Paginator(data, items_per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    page_ids = [obj.id for obj in page_obj]

    open_repairs_by_device_id = {}
    if page_ids:
        # Single query for "who put it under repair" on this page.
        # If multiple open repairs exist, we pick the most recent one per device.
        open_repairs = (
            DeviceRepair.objects.filter(
                device_id__in=page_ids,
                status=DeviceRepair.STATUS_IN_PROGRESS,
            )
            .select_related("created_by")
            .order_by("-created_at")
        )
        for repair in open_repairs:
            if repair.device_id not in open_repairs_by_device_id:
                open_repairs_by_device_id[repair.device_id] = repair

    latest_pending_by_import_id = {}
    deletion_request_by_device_id = {}
    if page_ids and not is_disposed:
        pending_updates = (
            PendingUpdate.objects.filter(import_record_id__in=page_ids)
            .select_related('updated_by', 'centre', 'department', 'assignee')
            .order_by('import_record_id', '-created_at')
        )
        for pending in pending_updates:
            latest_pending_by_import_id.setdefault(pending.import_record_id, pending)
        deletion_requests = (
            DeviceDeletionRequest.objects.filter(device_id__in=page_ids)
            .select_related('requested_by')
        )
        for deletion_request in deletion_requests:
            deletion_request_by_device_id[deletion_request.device_id] = deletion_request

    available_centres = Centre.objects.all().order_by('name')
    if request.user.is_trainer:
        available_centres = available_centres.filter(pk=request.user.centre_id) if request.user.centre_id else available_centres.none()

    # Pending updates
    data_with_pending = []
    for item in page_obj:
        pending = latest_pending_by_import_id.get(item.id)
        deletion_request = deletion_request_by_device_id.get(item.id)
        data_with_pending.append({
            'item': item,
            'pending_update': pending,
            'deletion_request': deletion_request,
            'open_repair': open_repairs_by_device_id.get(item.id),
            'approval_preview': _build_approval_preview(item, pending),
        })

    # Stats
    total_devices = initial_queryset.count()
    standard_pending_count = initial_queryset.filter(is_approved=False).count()
    deletion_request_count = initial_queryset.filter(deletion_request__isnull=False).count() if not is_disposed else 0
    unapproved_count = (
        initial_queryset.filter(Q(is_approved=False) | Q(deletion_request__isnull=False)).distinct().count()
        if not is_disposed
        else standard_pending_count
    )
    approved_imports = total_devices - unapproved_count
    this_month_count = initial_queryset.filter(
        date__year=timezone.now().year,
        date__month=timezone.now().month
    ).count() if is_disposed else 0

    category_choices = Import.CATEGORY_CHOICES

    # Configuration status for Laptop/Desktop/Server (shown as button in list)
    required_ids = set()
    needs_configuration_ids = set()
    configuration_progress = {}
    if page_ids:
        required_ids = set(
            Import.objects.filter(id__in=page_ids).filter(
                Q(category__in=['laptop', 'system_unit']) |
                Q(device_name__iendswith='-L-MOHI') |
                Q(device_name__iendswith='-D-MOHI') |
                Q(device_name__iendswith='-S-MOHI')
            ).values_list('id', flat=True)
        )
        if required_ids:
            progress_rows = DeviceConfiguration.objects.filter(
                device_id__in=required_ids
            ).values('device_id').annotate(
                total=Count('id'),
                completed=Count('id', filter=Q(is_completed=True)),
            )
            for row in progress_rows:
                device_id = row['device_id']
                total = row['total'] or 0
                completed = row['completed'] or 0
                configuration_progress[device_id] = {'completed': completed, 'total': total}

            for device_id in required_ids:
                progress = configuration_progress.get(device_id, {'completed': 0, 'total': 0})
                if progress['total'] == 0 or progress['completed'] < progress['total']:
                    needs_configuration_ids.add(device_id)

    return {
        'data_with_pending': data_with_pending,
        'paginator': paginator,
        'data': page_obj,
        'report_data': {
            'total_records': paginator.count,
            'search_query': search_query,
            'items_per_page': items_per_page,
        },
        'centres': available_centres,
        'departments': Department.objects.all().order_by('name'),
        'category_choices': category_choices,
        'centre_filter': centre_filter,
        'department_filter': department_filter,
        'show_duplicates': show_duplicates,
        'show_all_centres_option': not request.user.is_trainer,
        'items_per_page_options': [10, 25, 50, 100, 500],
        'unapproved_count': unapproved_count,
        'bulk_approvable_count': standard_pending_count,
        'total_devices': total_devices,
        'approved_imports': approved_imports,
        'view_name': view_name,
        'this_month_count': this_month_count,
        'needs_configuration_ids': needs_configuration_ids,
        'configuration_progress': configuration_progress,
        'can_delete_devices': can_delete_devices(request.user),
        'can_request_device_deletion': can_request_device_deletion(request.user),
        'can_review_device_requests': can_review_device_requests(request.user),
        'can_manage_assignments_list': can_manage_device_assignments(request.user),
    }

@login_required
def display_approved_imports(request):
    if request.user.is_trainer:
        initial_queryset = Import.objects.filter(
            centre=request.user.centre,
            is_approved=True,
            is_disposed=False,
            deletion_request__isnull=True,
        ) if request.user.centre else Import.objects.none()
    elif can_access_inventory_lists(request.user):
        initial_queryset = Import.objects.filter(
            is_approved=True,
            is_disposed=False,
            deletion_request__isnull=True,
        )
    else:
        initial_queryset = Import.objects.none()

    context = get_list_context(request, initial_queryset, 'display_approved_imports')
    return render(request, 'import/displaycsv_approved.html', context)

@login_required
def display_unapproved_imports(request):
    if request.user.is_trainer:
        initial_queryset = Import.objects.filter(
            centre=request.user.centre,
            is_disposed=False,
        ).filter(
            Q(is_approved=False) | Q(deletion_request__isnull=False)
        ) if request.user.centre else Import.objects.none()
    elif can_review_device_requests(request.user):
        initial_queryset = Import.objects.filter(is_disposed=False).filter(
            Q(is_approved=False) | Q(deletion_request__isnull=False)
        )
    else:
        initial_queryset = Import.objects.none()

    context = get_list_context(request, initial_queryset, 'display_unapproved_imports')
    return render(request, 'import/displaycsv_unapproved.html', context)

@login_required
def display_disposed_imports(request):
    if request.user.is_trainer:
        initial_queryset = Import.objects.filter(centre=request.user.centre, is_disposed=True) if request.user.centre else Import.objects.none()
    elif can_access_inventory_lists(request.user):
        initial_queryset = Import.objects.filter(is_disposed=True)
    else:
        initial_queryset = Import.objects.none()

    context = get_list_context(request, initial_queryset, 'display_disposed_imports', is_disposed=True)
    return render(request, 'import/displaycsv_disposed.html', context)


@login_required
def dispose_device(request, device_id):
    device = get_object_or_404(Import, id=device_id)
    if device.is_disposed:
        messages.error(request, "This device is already disposed.")
        return redirect('display_approved_imports')
    if request.method == 'POST':
        disposal_reason = request.POST.get('disposal_reason', '').strip()
        if not disposal_reason:
            messages.error(request, "Please provide a valid reason for disposal.")
            return redirect('display_approved_imports')
        with transaction.atomic():
            device.is_disposed = True
            device.disposal_reason = disposal_reason
            device.status = 'Disposed'
            device.reason_for_update = f"Device disposed by {request.user.username}: {disposal_reason}"
            device.save()
            messages.success(request, f"Device {device.serial_number} disposed successfully.")
            return redirect('display_disposed_imports')
    return render(request, 'import/dispose_device.html', {'device': device})




@login_required
def device_history(request, pk):
    device = get_object_or_404(
        Import.objects.select_related("assignee", "department", "centre", "added_by", "approved_by"),
        pk=pk,
    )

    # Permission check
    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only view devices from your own centre.")
        return redirect('display_approved_imports')

    # ===== Device Tracking Timeline (most recent first) =====
    historical_records = device.history.select_related(
        "assignee", "department", "centre", "history_user"
    ).order_by("history_date")  # oldest first for building
    timeline = []
    current_state = None
    start_date = None
    change_user = None

    for record in historical_records:
        assignee_display = (
            f"{record.assignee.first_name} {record.assignee.last_name} ({record.assignee.staff_number})"
            if record.assignee else "Unassigned"
        )
        department_display = record.department.name if record.department else "N/A"
        centre_display = record.centre.name if record.centre else "N/A"
        status_display = record.status or "N/A"
        condition_display = record.device_condition or "N/A"

        state = (
            assignee_display,
            department_display,
            centre_display,
            status_display,
            condition_display,
        )

        user = (
            record.history_user.get_full_name() or record.history_user.username
            if record.history_user else "System"
        )

        if current_state is None:
            current_state = state
            start_date = record.history_date
            change_user = user
            continue

        if state != current_state:
            timeline.append({
                'start_date': start_date,
                'end_date': record.history_date,
                'assignee': current_state[0],
                'department': current_state[1],
                'centre': current_state[2],
                'status': current_state[3],
                'condition': current_state[4],
                'changed_by': change_user,
            })
            current_state = state
            start_date = record.history_date
            change_user = user

    if current_state:
        timeline.append({
            'start_date': start_date,
            'end_date': 'Current',
            'assignee': current_state[0],
            'department': current_state[1],
            'centre': current_state[2],
            'status': current_state[3],
            'condition': current_state[4],
            'changed_by': 'N/A',
        })

    # Reverse to show most recent first
    timeline = timeline[::-1]

    # ===== Summarized Change History (newest first) =====
    history_records = list(
        device.history.select_related("history_user").order_by("-history_date")
    )  # list for indexing
    history_data = []

    field_names = {
        'centre': 'Centre',
        'department': 'Department',
        'device_name': 'Device Name',
        'system_model': 'System Model',
        'processor': 'Processor',
        'ram_gb': 'RAM (GB)',
        'hdd_gb': 'HDD (GB)',
        'serial_number': 'Serial Number',
        'assignee': 'Assignee',
        'device_condition': 'Device Condition',
        'status': 'Status',
        'added_by': 'Added By',
        'approved_by': 'Approved By',
        'is_approved': 'Is Approved',
        'reason_for_update': 'Reason for Update',
        'category': 'Category',
    }

    i = 0
    centre_name_cache: dict[int, str] = {}
    department_name_cache: dict[int, str] = {}
    user_username_cache: dict[int, str] = {}
    employee_display_cache: dict[int, str] = {}

    def _as_int(value):
        try:
            return int(value)
        except Exception:
            return None

    def _centre_name(value) -> str:
        pk_int = _as_int(value)
        if not pk_int:
            return "N/A"
        if pk_int in centre_name_cache:
            return centre_name_cache[pk_int]
        name = Centre.objects.filter(pk=pk_int).values_list("name", flat=True).first() or "N/A"
        centre_name_cache[pk_int] = name
        return name

    def _department_name(value) -> str:
        pk_int = _as_int(value)
        if not pk_int:
            return "N/A"
        if pk_int in department_name_cache:
            return department_name_cache[pk_int]
        name = Department.objects.filter(pk=pk_int).values_list("name", flat=True).first() or "N/A"
        department_name_cache[pk_int] = name
        return name

    def _username(value) -> str:
        pk_int = _as_int(value)
        if not pk_int:
            return "N/A"
        if pk_int in user_username_cache:
            return user_username_cache[pk_int]
        username = CustomUser.objects.filter(pk=pk_int).values_list("username", flat=True).first() or "N/A"
        user_username_cache[pk_int] = username
        return username

    def _employee_display(value) -> str:
        pk_int = _as_int(value)
        if not pk_int:
            return "N/A"
        if pk_int in employee_display_cache:
            return employee_display_cache[pk_int]
        row = Employee.objects.filter(pk=pk_int).values("first_name", "last_name", "staff_number").first()
        if not row:
            disp = "N/A"
        else:
            staff_no = row.get("staff_number") or "N/A"
            first = (row.get("first_name") or "").strip()
            last = (row.get("last_name") or "").strip()
            full = (f"{first} {last}").strip() or "N/A"
            disp = f"{full} ({staff_no})"
        employee_display_cache[pk_int] = disp
        return disp

    while i < len(history_records):
        record = history_records[i]
        user = (
            record.history_user.get_full_name() or record.history_user.username
            if record.history_user else "Unknown"
        )

        # Handle creation
        if record.history_type == '+':
            history_data.append({
                'date': record.history_date,
                'change_type': 'Created',
                'diff': {},
                'user': user,
                'is_multiple': False,
            })
            i += 1
            continue

        # Start group
        group = [{'record': record, 'user': user}]
        j = i + 1
        while j < len(history_records):
            next_record = history_records[j]
            next_user = (
                next_record.history_user.get_full_name() or next_record.history_user.username
                if next_record.history_user else "Unknown"
            )
            time_diff = record.history_date - next_record.history_date
            if next_user == user and time_diff <= timedelta(minutes=15):
                group.append({'record': next_record, 'user': next_user})
                j += 1
            else:
                break

        # Base record for diff (record just before the group, older than the group's oldest)
        prev = history_records[j] if j < len(history_records) else None

        # Latest record for date
        latest_record = group[0]['record']

        if prev is None:
            i = j
            continue

        changes = latest_record.diff_against(prev)
        diff = {}

        for change in changes.changes:
            if not hasattr(change, 'field'):
                continue

            field_name = field_names.get(change.field, change.field.replace('_', ' ').title())

            # Resolve values
            if change.field == 'centre':
                old_value = _centre_name(change.old)
                new_value = _centre_name(change.new)
            elif change.field == 'department':
                old_value = _department_name(change.old)
                new_value = _department_name(change.new)
            elif change.field in ['added_by', 'approved_by']:
                old_value = _username(change.old)
                new_value = _username(change.new)
            elif change.field == 'assignee':
                old_value = _employee_display(change.old)
                new_value = _employee_display(change.new)
            else:
                old_value = change.old if change.old is not None else 'N/A'
                new_value = change.new if change.new is not None else 'N/A'

            if str(old_value).strip() == str(new_value).strip():
                continue

            diff[field_name] = {'old': old_value, 'new': new_value}

        if diff or len(group) == 1:
            history_data.append({
                'date': latest_record.history_date,
                'change_type': 'Edited' if len(group) == 1 else 'Edited (multiple saves)',
                'diff': diff,
                'user': user,
                'is_multiple': len(group) > 1,
            })

        i = j

    # ===== Legacy User History =====
    user_history = device.user_history.all().order_by('assigned_date').values(
        'assignee_first_name', 'assignee_last_name', 'assignee_email_address',
        'assigned_by__username', 'assigned_date', 'cleared_date'
    )

    # IMPORTANT: your template uses entry.assignee_first_name etc, so keep raw dict keys
    user_history_data = [
        {
            'assignee_first_name': f"{entry['assignee_first_name'] or ''} {entry['assignee_last_name'] or ''}".strip() or 'N/A',
            'email': entry['assignee_email_address'] or 'N/A',
            'assigned_by': entry['assigned_by__username'] or 'N/A',
            'assigned_date': entry['assigned_date'],
            'cleared_date': entry['cleared_date']
        }
        for entry in user_history
    ]

    # ===== UAF Agreements =====
    past_agreements = DeviceAgreement.objects.filter(
        device=device, is_archived=True
    ).select_related('employee', 'issuance_it_user', 'clearance_it_user').order_by('-clearance_date')

    current_agreement = DeviceAgreement.objects.filter(
        device=device, is_archived=False
    ).select_related('employee', 'issuance_it_user').first()

    maintenance_logs = (
        device.logs.filter(ppm_task__isnull=False)
        .select_related("user", "ppm_task", "ppm_task__period")
        .prefetch_related("ppm_task__activities")
        .order_by("-created_at")
    )

    reason_re = re.compile(r"Reason:\s*(.*?)(?:\n|$)", re.IGNORECASE)
    notes_re = re.compile(r"Notes:\s*(.*?)(?:\n|$)", re.IGNORECASE)
    notes_pipe_re = re.compile(r"\|\s*Notes:\s*(.*)$", re.IGNORECASE)
    activities_done_re = re.compile(r"Activities performed:\s*(.*?)(?:\n|$)", re.IGNORECASE)
    activities_legacy_re = re.compile(r"Activities:\s*(.*?)(?:\n|$)", re.IGNORECASE)

    for log in maintenance_logs:
        msg = (log.message or "").replace("\r\n", "\n").replace("\r", "\n")
        msg_lower = msg.lower()

        log.ppm_is_not_done = ("no activity performed" in msg_lower) or ("not done" in msg_lower)

        log.ppm_reason = None
        log.ppm_notes = None
        log.ppm_activities = None

        reason_match = reason_re.search(msg)
        if reason_match:
            log.ppm_reason = reason_match.group(1).strip() or None

        notes_match = notes_re.search(msg) or notes_pipe_re.search(msg)
        if notes_match:
            log.ppm_notes = notes_match.group(1).strip() or None

        activities_match = activities_done_re.search(msg) or activities_legacy_re.search(msg)
        if activities_match:
            log.ppm_activities = activities_match.group(1).strip() or None

    return render(request, 'import/device_history.html', {
        'device': device,
        'timeline': timeline,
        'history': history_data,
        'user_history': user_history_data,
        'past_agreements': past_agreements,
        'current_agreement': current_agreement,
        'maintenance_logs': maintenance_logs,
    })


@login_required
def device_detail(request, pk):
    device = get_object_or_404(
        Import.objects.select_related("assignee", "department", "centre", "added_by", "approved_by"),
        pk=pk,
    )

    # Permission: trainers can only view their centre's devices
    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only view devices from your own centre.")
        return redirect('display_approved_imports')

    agreement_exists = DeviceAgreement.objects.filter(device=device, is_archived=False).exists()
    past_agreements_exist = DeviceAgreement.objects.filter(device=device, is_archived=True).exists()
    current_agreement = None
    if device.assignee:
        current_agreement = (
            DeviceAgreement.objects.filter(device=device, employee=device.assignee, is_archived=False)
            .order_by("-issuance_date", "-id")
            .first()
        )
    if not current_agreement:
        current_agreement = DeviceAgreement.objects.filter(device=device, is_archived=False).order_by("-issuance_date", "-id").first()

    clearance = device.clearances.order_by('-created_at').first()
    can_manage_assignments = can_manage_device_assignments(request.user)
    can_clear_user = can_clear_device_users(request.user)
    can_trainer_reassign = _can_trainer_reassign_device(request.user, device)

    config_kind = _device_asset_kind_from_record(device)
    supports_configuration = config_kind in {"laptop", "desktop", "server"}
    configuration_types_exist = False
    if supports_configuration:
        types = DeviceConfigurationType.objects.filter(is_active=True)
        if config_kind == "laptop":
            types = types.filter(applies_to_laptop=True)
        elif config_kind == "desktop":
            types = types.filter(applies_to_desktop=True)
        else:
            types = types.filter(applies_to_server=True)
        configuration_types_exist = types.exists()

    context = {
        'device': device,
        'agreement_exists': agreement_exists,
        'past_agreements_exist': past_agreements_exist,
        'current_agreement': current_agreement,
        'clearance': clearance,

        'can_edit': can_manage_assignments,
        'can_manage_assignments': can_manage_assignments,
        'can_clear_user': can_clear_user,
        'can_trainer_reassign': can_trainer_reassign,
        'can_approve': can_review_device_requests(request.user) and not request.user.is_trainer and not device.is_approved,
        'can_dispose': request.user.is_superuser and not request.user.is_trainer and not device.is_disposed,
        'can_delete': can_request_device_deletion(request.user),

        'supports_configuration': supports_configuration,
        'configuration_types_exist': configuration_types_exist,
    }

    return render(request, 'import/device_detail.html', context)


@login_required
def device_repairs(request, pk):
    device = get_object_or_404(
        Import.objects.select_related("assignee", "department", "centre"),
        pk=pk,
    )

    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only view devices from your own centre.")
        return redirect("display_approved_imports")

    repair_assignees = (
        CustomUser.objects.filter(is_active=True)
        .filter(Q(is_trainer=False) | Q(pk=request.user.pk) | Q(repairs_assigned__device=device))
        .distinct()
        .order_by("first_name", "last_name", "username")
    )
    repairs = DeviceRepair.objects.filter(device=device).select_related("created_by", "assigned_to").order_by("-created_at")
    open_repairs = repairs.filter(status=DeviceRepair.STATUS_IN_PROGRESS)

    return render(request, "import/device_repairs.html", {
        "device": device,
        "repairs": repairs,
        "open_repairs": open_repairs,
        "repair_assignees": repair_assignees,
    })


def _get_repair_assignee(assigned_to_id):
    if not assigned_to_id:
        return None
    return CustomUser.objects.filter(
        pk=assigned_to_id,
        is_active=True,
        is_trainer=False,
    ).first()


def _sync_device_repair_state(device):
    has_open_repairs = DeviceRepair.objects.filter(
        device=device,
        status=DeviceRepair.STATUS_IN_PROGRESS,
    ).exists()
    has_written_off_repair = DeviceRepair.objects.filter(
        device=device,
        status=DeviceRepair.STATUS_WRITTEN_OFF,
    ).exists()
    target_is_active = not has_open_repairs and not has_written_off_repair
    if device.is_active != target_is_active:
        device.is_active = target_is_active
        device.save(update_fields=["is_active"])


def _notify_repair_assignment(*, repair, assigned_by, is_reassignment):
    assigned_to = repair.assigned_to
    if not assigned_to or not assigned_to.id:
        return
    if assigned_by and assigned_to.pk == assigned_by.pk and not is_reassignment:
        return

    action = "reassigned" if is_reassignment else "assigned"
    sender_name = assigned_by.get_full_name() if assigned_by else "System"
    device_label = repair.device.serial_number or repair.device.system_model or repair.device.device_name or f"Device {repair.device_id}"
    notification_message = f"{sender_name} has {action} repair for {device_label} to you. Open the repair page to continue."

    Notification.objects.create(
        user=assigned_to,
        message=notification_message,
        content_type=ContentType.objects.get_for_model(DeviceRepair),
        object_id=repair.pk,
    )

    if assigned_to.email:
        subject = f"Repair {'Handover' if is_reassignment else 'Assignment'}: {device_label}"
        message = (
            f"Hello {assigned_to.get_full_name()},\n\n"
            f"{sender_name} has {action} a repair to you.\n\n"
            f"Device: {device_label}\n"
            f"Status: {dict(DeviceRepair.STATUS_CHOICES).get(repair.status, repair.status)}\n"
            f"Issue: {repair.issue_description}\n"
            f"Technician: {repair.technician_responsible or assigned_to.get_full_name()}\n\n"
            f"Open the device repairs page in the system to continue working on it.\n"
        )
        send_custom_email(
            subject,
            message,
            [assigned_to.email],
            also_notify=False,
            related_object=repair,
        )


def _build_handover_note(previous_assignee, new_assignee, actor):
    current_time = timezone.now()
    if timezone.is_aware(current_time):
        current_time = timezone.localtime(current_time)
    timestamp = current_time.strftime("%Y-%m-%d %H:%M")
    previous_name = previous_assignee.get_full_name() if previous_assignee else "Unassigned"
    new_name = new_assignee.get_full_name() if new_assignee else "Unassigned"
    actor_name = actor.get_full_name() if actor else "System"
    return f"[{timestamp}] Repair handed over from {previous_name} to {new_name} by {actor_name}."


@login_required
@require_POST
def device_repair_create(request, pk):
    device = get_object_or_404(
        Import.objects.select_related("assignee", "department", "centre"),
        pk=pk,
    )

    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only update records for your own centre.")
        return redirect("display_approved_imports")

    date_str = (request.POST.get("date_of_repair") or "").strip()
    date_of_repair = None
    if not date_str:
        messages.error(request, "Date of repair is required.")
        return redirect("device_detail", pk=device.pk)
    try:
        date_of_repair = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "Invalid repair date.")
        return redirect("device_detail", pk=device.pk)

    month = date_of_repair.strftime("%B")

    reported_by = (request.POST.get("reported_by") or "").strip()
    issue_description = (request.POST.get("issue_description") or "").strip()
    repair_action_taken = (request.POST.get("repair_action_taken") or "").strip()
    technician_responsible = (request.POST.get("technician_responsible") or "").strip()
    assigned_to = _get_repair_assignee(request.POST.get("assigned_to"))
    notes = (request.POST.get("notes") or "").strip()
    external_repair = (request.POST.get("external_repair") in {"on", "true", "1", True})
    status = (request.POST.get("status") or DeviceRepair.STATUS_IN_PROGRESS).strip()

    if not issue_description:
        messages.error(request, "Issue description is required.")
        return redirect("device_detail", pk=device.pk)

    cost_val = (request.POST.get("cost_kes") or "").strip()
    cost_kes = None
    if cost_val:
        try:
            cost_kes = Decimal(cost_val)
        except (InvalidOperation, ValueError):
            messages.error(request, "Invalid cost value.")
            return redirect("device_detail", pk=device.pk)

    centre_name = device.centre.name if device.centre else "N/A"
    dept_name = device.department.name if device.department else "N/A"
    centre_department = f"{centre_name} - {dept_name}"

    if device.assignee:
        owner = f"{device.assignee.full_name} ({device.assignee.staff_number or 'No ID'})"
    else:
        owner = "Unassigned"

    if assigned_to is None:
        assigned_to = request.user

    technician_name = technician_responsible or assigned_to.get_full_name()
    repair = DeviceRepair.objects.create(
        device=device,
        month=month or None,
        date_of_repair=date_of_repair,
        centre_department=centre_department,
        owner=owner,
        model=device.system_model or device.device_name or None,
        serial_number=device.serial_number or None,
        reported_by=reported_by or None,
        issue_description=issue_description,
        repair_action_taken=repair_action_taken or None,
        technician_responsible=technician_name or None,
        cost_kes=cost_kes,
        status=status if status in dict(DeviceRepair.STATUS_CHOICES) else DeviceRepair.STATUS_IN_PROGRESS,
        notes=notes or None,
        external_repair=bool(external_repair),
        created_by=request.user,
        assigned_to=assigned_to,
        completed_at=timezone.now() if status == DeviceRepair.STATUS_COMPLETED else None,
    )

    _sync_device_repair_state(device)
    if assigned_to and assigned_to.pk != request.user.pk:
        _notify_repair_assignment(repair=repair, assigned_by=request.user, is_reassignment=False)

    messages.success(request, "Repair record saved.")
    return redirect("device_repairs", pk=device.pk)


@login_required
@require_POST
def device_repair_edit(request, pk, repair_id):
    device = get_object_or_404(
        Import.objects.select_related("assignee", "department", "centre"),
        pk=pk,
    )
    repair = get_object_or_404(DeviceRepair.objects.select_related("assigned_to"), pk=repair_id, device_id=device.pk)

    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only update records for your own centre.")
        return redirect("display_approved_imports")

    if repair.status != DeviceRepair.STATUS_IN_PROGRESS:
        messages.error(request, "Only repairs in progress can be edited.")
        return redirect("device_repairs", pk=device.pk)

    reported_by = (request.POST.get("reported_by") or "").strip()
    issue_description = (request.POST.get("issue_description") or "").strip()
    repair_action_taken = (request.POST.get("repair_action_taken") or "").strip()
    technician_responsible = (request.POST.get("technician_responsible") or "").strip()
    notes = (request.POST.get("notes") or "").strip()
    external_repair = (request.POST.get("external_repair") in {"on", "true", "1", True})
    status = (request.POST.get("status") or DeviceRepair.STATUS_IN_PROGRESS).strip()
    assigned_to = _get_repair_assignee(request.POST.get("assigned_to")) or repair.assigned_to or request.user

    if not issue_description:
        messages.error(request, "Issue description is required.")
        return redirect("device_repairs", pk=device.pk)

    cost_val = (request.POST.get("cost_kes") or "").strip()
    cost_kes = None
    if cost_val:
        try:
            cost_kes = Decimal(cost_val)
        except (InvalidOperation, ValueError):
            messages.error(request, "Invalid cost value.")
            return redirect("device_repairs", pk=device.pk)

    previous_assignee = repair.assigned_to
    assignee_changed = (previous_assignee.pk if previous_assignee else None) != assigned_to.pk
    combined_notes = notes
    if assignee_changed:
        handover_note = _build_handover_note(previous_assignee, assigned_to, request.user)
        combined_notes = f"{combined_notes}\n{handover_note}".strip() if combined_notes else handover_note

    repair.reported_by = reported_by or None
    repair.issue_description = issue_description
    repair.repair_action_taken = repair_action_taken or None
    repair.technician_responsible = technician_responsible or assigned_to.get_full_name() or None
    repair.cost_kes = cost_kes
    repair.notes = combined_notes or None
    repair.external_repair = bool(external_repair)
    repair.assigned_to = assigned_to
    repair.status = status if status in dict(DeviceRepair.STATUS_CHOICES) else DeviceRepair.STATUS_IN_PROGRESS
    repair.completed_at = timezone.now() if repair.status == DeviceRepair.STATUS_COMPLETED else None
    repair.save(
        update_fields=[
            "reported_by",
            "issue_description",
            "repair_action_taken",
            "technician_responsible",
            "cost_kes",
            "notes",
            "external_repair",
            "assigned_to",
            "status",
            "completed_at",
            "updated_at",
        ]
    )

    _sync_device_repair_state(device)

    if assignee_changed and assigned_to and assigned_to.pk != request.user.pk:
        _notify_repair_assignment(repair=repair, assigned_by=request.user, is_reassignment=True)

    messages.success(request, "Repair record updated.")
    return redirect("device_repairs", pk=device.pk)


@login_required
@require_POST
def device_repair_close(request, pk, repair_id):
    device = get_object_or_404(Import, pk=pk)
    repair = get_object_or_404(DeviceRepair, pk=repair_id, device_id=device.pk)

    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only update records for your own centre.")
        return redirect("display_approved_imports")

    repair.status = DeviceRepair.STATUS_COMPLETED
    repair.completed_at = timezone.now()
    repair.save(update_fields=["status", "completed_at", "updated_at"])

    _sync_device_repair_state(device)

    messages.success(request, "Repair marked as completed.")
    return redirect("device_repairs", pk=device.pk)


@login_required
def repair_report(request):
    user = request.user
    qs = DeviceRepair.objects.select_related("device", "created_by", "device__centre", "device__department")

    if user.is_trainer and user.centre:
        qs = qs.filter(device__centre=user.centre)

    period = (request.GET.get("period") or "").strip()  # YYYY-MM
    if period:
        try:
            year_str, month_str = period.split("-", 1)
            qs = qs.filter(date_of_repair__year=int(year_str), date_of_repair__month=int(month_str))
        except Exception:
            period = ""

    status_filter = (request.GET.get("status") or "").strip()
    if status_filter in {DeviceRepair.STATUS_IN_PROGRESS, DeviceRepair.STATUS_COMPLETED, DeviceRepair.STATUS_WRITTEN_OFF}:
        qs = qs.filter(status=status_filter)
    else:
        status_filter = ""

    if request.GET.get("export") == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Repairs"
        ws.append([
            "MONTH",
            "Date of Repair",
            "Center/Department",
            "Owner",
            "Model",
            "Serial Number",
            "Reported By",
            "Issue Description",
            "Repair Action Taken",
            "Technician Responsible",
            "Cost (KES)",
            "Status",
            "Notes",
            "External repair",
        ])
        for r in qs.order_by("-date_of_repair", "-created_at"):
            ws.append([
                r.month or "",
                r.date_of_repair.isoformat() if r.date_of_repair else "",
                r.centre_department or "",
                r.owner or "",
                r.model or "",
                r.serial_number or "",
                r.reported_by or "",
                r.issue_description or "",
                r.repair_action_taken or "",
                r.technician_responsible or "",
                float(r.cost_kes) if r.cost_kes is not None else "",
                dict(DeviceRepair.STATUS_CHOICES).get(r.status, r.status),
                r.notes or "",
                bool(r.external_repair),
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        resp = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = "repair_report.xlsx" if not period else f"repair_report_{period}.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    periods = (
        DeviceRepair.objects.exclude(date_of_repair=None)
        .annotate(m=TruncMonth("date_of_repair"))
        .values_list("m", flat=True)
        .distinct()
        .order_by("-m")
    )
    period_options = [p.strftime("%Y-%m") for p in periods if p]

    total_repairs = qs.count()
    open_repairs = qs.filter(status=DeviceRepair.STATUS_IN_PROGRESS).count()
    completed_repairs = qs.filter(status=DeviceRepair.STATUS_COMPLETED).count()
    written_off_repairs = qs.filter(status=DeviceRepair.STATUS_WRITTEN_OFF).count()
    total_cost = qs.aggregate(total=Sum("cost_kes"))["total"] or 0
    external_count = qs.filter(external_repair=True).count()

    monthly = (
        qs.exclude(date_of_repair=None)
        .annotate(m=TruncMonth("date_of_repair"))
        .values("m")
        .annotate(count=Count("id"), cost=Sum("cost_kes"))
        .order_by("m")
    )
    chart_labels = [row["m"].strftime("%b %Y") for row in monthly if row["m"]]
    chart_counts = [row["count"] for row in monthly]
    chart_costs = [float(row["cost"] or 0) for row in monthly]

    return render(request, "repairs/repair_report.html", {
        "period": period,
        "status_filter": status_filter,
        "period_options": period_options,
        "total_repairs": total_repairs,
        "open_repairs": open_repairs,
        "completed_repairs": completed_repairs,
        "written_off_repairs": written_off_repairs,
        "external_count": external_count,
        "total_cost": total_cost,
        "chart_labels": chart_labels,
        "chart_counts": chart_counts,
        "chart_costs": chart_costs,
        "repairs": qs.order_by("-date_of_repair", "-created_at")[:200],
    })

def _build_device_history_payload(device: Import):
    """
    Builds history + agreements context used by the comprehensive device details page.
    """
    # ===== Device Tracking Timeline (most recent first) =====
    historical_records = device.history.order_by('history_date')  # oldest first for building
    timeline = []
    current_state = None
    start_date = None
    change_user = None

    for record in historical_records:
        assignee_display = (
            f"{record.assignee.first_name} {record.assignee.last_name} ({record.assignee.staff_number})"
            if getattr(record, 'assignee', None) else "Unassigned"
        )
        department_display = record.department.name if getattr(record, 'department', None) else "N/A"
        centre_display = record.centre.name if getattr(record, 'centre', None) else "N/A"
        status_display = record.status or "N/A"
        condition_display = record.device_condition or "N/A"

        state = (
            assignee_display,
            department_display,
            centre_display,
            status_display,
            condition_display,
        )

        user = (
            record.history_user.get_full_name() or record.history_user.username
            if getattr(record, 'history_user', None) else "System"
        )

        if current_state is None:
            current_state = state
            start_date = record.history_date
            change_user = user
            continue

        if state != current_state:
            timeline.append({
                'start_date': start_date,
                'end_date': record.history_date,
                'assignee': current_state[0],
                'department': current_state[1],
                'centre': current_state[2],
                'status': current_state[3],
                'condition': current_state[4],
                'changed_by': change_user,
            })
            current_state = state
            start_date = record.history_date
            change_user = user

    if current_state:
        timeline.append({
            'start_date': start_date,
            'end_date': 'Current',
            'assignee': current_state[0],
            'department': current_state[1],
            'centre': current_state[2],
            'status': current_state[3],
            'condition': current_state[4],
            'changed_by': 'N/A',
        })

    timeline = timeline[::-1]

    # ===== Summarized Change History (newest first) =====
    history_records = list(device.history.all().order_by('-history_date'))
    history_data = []

    field_names = {
        'centre': 'Centre',
        'department': 'Department',
        'device_name': 'Device Name',
        'system_model': 'System Model',
        'processor': 'Processor',
        'ram_gb': 'RAM (GB)',
        'hdd_gb': 'HDD (GB)',
        'serial_number': 'Serial Number',
        'assignee': 'Assignee',
        'device_condition': 'Device Condition',
        'status': 'Status',
        'added_by': 'Added By',
        'approved_by': 'Approved By',
        'is_approved': 'Is Approved',
        'reason_for_update': 'Reason for Update',
        'category': 'Category',
    }

    i = 0
    while i < len(history_records):
        record = history_records[i]
        user = (
            record.history_user.get_full_name() or record.history_user.username
            if getattr(record, 'history_user', None) else "Unknown"
        )

        if record.history_type == '+':
            history_data.append({
                'date': record.history_date,
                'change_type': 'Created',
                'diff': {},
                'user': user,
                'is_multiple': False,
            })
            i += 1
            continue

        group = [{'record': record, 'user': user}]
        j = i + 1
        while j < len(history_records):
            next_record = history_records[j]
            next_user = (
                next_record.history_user.get_full_name() or next_record.history_user.username
                if getattr(next_record, 'history_user', None) else "Unknown"
            )
            time_diff = record.history_date - next_record.history_date
            if next_user == user and time_diff <= timedelta(minutes=15):
                group.append({'record': next_record, 'user': next_user})
                j += 1
            else:
                break

        first_record_in_group = group[-1]['record']
        prev = first_record_in_group.prev_record
        latest_record = group[0]['record']

        if prev is None:
            i = j
            continue

        changes = latest_record.diff_against(prev)
        diff = {}

        for change in changes.changes:
            if not hasattr(change, 'field'):
                continue

            field_name = field_names.get(change.field, change.field.replace('_', ' ').title())

            if change.field == 'centre':
                old_value = Centre.objects.get(pk=change.old).name if change.old and Centre.objects.filter(pk=change.old).exists() else 'N/A'
                new_value = Centre.objects.get(pk=change.new).name if change.new and Centre.objects.filter(pk=change.new).exists() else 'N/A'
            elif change.field == 'department':
                old_value = Department.objects.get(pk=change.old).name if change.old and Department.objects.filter(pk=change.old).exists() else 'N/A'
                new_value = Department.objects.get(pk=change.new).name if change.new and Department.objects.filter(pk=change.new).exists() else 'N/A'
            elif change.field in ['added_by', 'approved_by']:
                old_value = CustomUser.objects.get(pk=change.old).username if change.old and CustomUser.objects.filter(pk=change.old).exists() else 'N/A'
                new_value = CustomUser.objects.get(pk=change.new).username if change.new and CustomUser.objects.filter(pk=change.new).exists() else 'N/A'
            elif change.field == 'assignee':
                old_value = f"{Employee.objects.get(pk=change.old).full_name} ({Employee.objects.get(pk=change.old).staff_number})" if change.old and Employee.objects.filter(pk=change.old).exists() else 'N/A'
                new_value = f"{Employee.objects.get(pk=change.new).full_name} ({Employee.objects.get(pk=change.new).staff_number})" if change.new and Employee.objects.filter(pk=change.new).exists() else 'N/A'
            else:
                old_value = change.old if change.old is not None else 'N/A'
                new_value = change.new if change.new is not None else 'N/A'

            if str(old_value).strip() == str(new_value).strip():
                continue

            diff[field_name] = {'old': old_value, 'new': new_value}

        if diff or len(group) == 1:
            history_data.append({
                'date': latest_record.history_date,
                'change_type': 'Edited' if len(group) == 1 else 'Edited (multiple saves)',
                'diff': diff,
                'user': user,
                'is_multiple': len(group) > 1,
            })

        i = j

    # ===== Legacy User History =====
    user_history = device.user_history.all().order_by('assigned_date').values(
        'assignee_first_name', 'assignee_last_name', 'assignee_email_address',
        'assigned_by__username', 'assigned_date', 'cleared_date'
    )
    user_history_data = [
        {
            'assignee_name': f"{entry['assignee_first_name'] or ''} {entry['assignee_last_name'] or ''}".strip() or 'N/A',
            'email': entry['assignee_email_address'] or 'N/A',
            'assigned_by': entry['assigned_by__username'] or 'N/A',
            'assigned_date': entry['assigned_date'],
            'cleared_date': entry['cleared_date']
        }
        for entry in user_history
    ]

    past_agreements = DeviceAgreement.objects.filter(
        device=device, is_archived=True
    ).select_related('employee', 'issuance_it_user', 'clearance_it_user').order_by('-clearance_date')

    current_agreement = DeviceAgreement.objects.filter(
        device=device, is_archived=False
    ).select_related('employee', 'issuance_it_user', 'clearance_it_user').first()

    return {
        'timeline': timeline,
        'history': history_data,
        'user_history': user_history_data,
        'past_agreements': past_agreements,
        'current_agreement': current_agreement,
    }



@login_required
def export_to_excel(request):
    # === GET PARAMETERS ===
    scope = request.GET.get('scope', 'page')
    search_query = request.GET.get('search', '').strip()
    page_number = request.GET.get('page', '1')
    items_per_page = request.GET.get('items_per_page', '10')
    view_context = request.GET.get('view_context', 'display_approved_imports')

    # === MISSING FILTERS ADDED ===
    centre_filter = request.GET.get('centre', '').strip()
    department_filter = request.GET.get('department', '').strip()
    show_duplicates = request.GET.get('show_duplicates')

    # ---- pagination / validation -------------------------------------------------
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10

    try:
        page_number = int(page_number) if page_number else 1
    except ValueError:
        page_number = 1

    # ---- base queryset ----------------------------------------------------------
    if request.user.is_superuser:
        base_qs = Import.objects.all()
    elif request.user.is_trainer:
        base_qs = Import.objects.filter(centre=request.user.centre)
    else:
        base_qs = Import.objects.none()

    # Apply view context filtering
    if view_context == 'display_unapproved_imports':
        data = base_qs.filter(is_approved=False, is_disposed=False).order_by('-pk')
    elif view_context == 'display_disposed_imports':
        data = base_qs.filter(is_disposed=True).order_by('-pk')
    else:  # default: approved
        data = base_qs.filter(is_approved=True, is_disposed=False).order_by('-pk')

    # === APPLY FILTERS (same logic as display views) ===
    if centre_filter:
        data = data.filter(centre__id=centre_filter)
    if department_filter:
        data = data.filter(department__id=department_filter)

    # Duplicate filter
    if show_duplicates == 'on':
        duplicate_serials = (
            data.values('serial_number')
                .annotate(serial_count=Count('serial_number'))
                .filter(serial_count__gt=1)
                .values_list('serial_number', flat=True)
        )
        data = data.filter(serial_number__in=duplicate_serials)

    # Search filter
    if search_query:
        search_q = (
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(device_name__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            _build_assignee_search_query(search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(reason_for_update__icontains=search_query)
        )
        # Add disposal_reason only for disposed view
        if view_context == 'display_disposed_imports':
            search_q |= Q(disposal_reason__icontains=search_query)
        data = data.filter(search_q)

    # === PAGINATION FOR "PAGE" SCOPE ===
    final_data = data
    if scope == 'page':
        paginator = Paginator(data, items_per_page)
        try:
            page_obj = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.page(1)
        final_data = page_obj.object_list  # Only items on current page
    else:
        final_data = list(data)  # All filtered items

    # ---- workbook ---------------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IT Inventory"

    # ---- headers ----------------------------------------------------------------
    headers = [
        'Centre', 'Department', 'Category', 'device_name', 'System Model',
        'Processor', 'RAM (GB)', 'HDD (GB)', 'Serial Number',
        'Assignee First Name', 'Assignee Last Name', 'Assignee Email',
        'Device Condition', 'Status', 'Date', 'Added By',
        'Approved By', 'Is Approved', 'Disposal Reason'
    ]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- data rows --------------------------------------------------------------
    for item in final_data:
        row = [
            item.centre.name if item.centre else 'N/A',
            item.department.name if item.department else 'N/A',
            item.get_category_display() or 'N/A',
            item.device_name or 'N/A',
            item.system_model or 'N/A',
            item.processor or 'N/A',
            item.ram_gb or 'N/A',
            item.hdd_gb or 'N/A',
            item.serial_number or 'N/A',
            item.assignee_first_name or 'N/A',
            item.assignee_last_name or 'N/A',
            item.assignee_email_address or 'N/A',
            item.device_condition or 'N/A',
            item.status or 'N/A',
            item.date.strftime('%Y-%m-%d') if item.date else 'N/A',
            item.added_by.username if item.added_by else 'N/A',
            item.approved_by.username if item.approved_by else 'N/A',
            'Yes' if item.is_approved else 'No',
            item.disposal_reason or 'N/A',
        ]
        ws.append(row)

    # Wrap text & auto-adjust column widths
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap_align

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # ---- response ---------------------------------------------------------------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"IT_Inventory_{'All' if scope == 'all' else 'Page'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_to_pdf(request):
    # === GET PARAMETERS ===
    scope = request.GET.get('scope', 'page')
    search_query = request.GET.get('search', '').strip()
    page_number = request.GET.get('page', '1')
    items_per_page = request.GET.get('items_per_page', '10')
    view_context = request.GET.get('view_context', 'display_approved_imports')

    # === MISSING FILTERS ADDED ===
    centre_filter = request.GET.get('centre', '').strip()
    department_filter = request.GET.get('department', '').strip()
    show_duplicates = request.GET.get('show_duplicates')

    # --- Pagination validation ---
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10
    try:
        page_number = int(page_number) if page_number else 1
    except ValueError:
        page_number = 1

    # --- Base Queryset ---
    if request.user.is_superuser:
        base_qs = Import.objects.select_related('centre', 'department', 'added_by', 'approved_by')
    elif request.user.is_trainer:
        base_qs = Import.objects.filter(centre=request.user.centre)\
                             .select_related('centre', 'department', 'added_by', 'approved_by')
    else:
        base_qs = Import.objects.none()

    # Apply view context
    if view_context == 'display_unapproved_imports':
        qs = base_qs.filter(is_approved=False, is_disposed=False).order_by('-pk')
    elif view_context == 'display_disposed_imports':
        qs = base_qs.filter(is_disposed=True).order_by('-pk')
    else:
        qs = base_qs.filter(is_approved=True, is_disposed=False).order_by('-pk')

    # === APPLY FILTERS ===
    if centre_filter:
        qs = qs.filter(centre__id=centre_filter)
    if department_filter:
        qs = qs.filter(department__id=department_filter)

    if show_duplicates == 'on':
        duplicate_serials = (
            qs.values('serial_number')
                .annotate(serial_count=Count('serial_number'))
                .filter(serial_count__gt=1)
                .values_list('serial_number', flat=True)
        )
        qs = qs.filter(serial_number__in=duplicate_serials)

    if search_query:
        search_q = (
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(device_name__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            _build_assignee_search_query(search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(reason_for_update__icontains=search_query)
        )
        if view_context == 'display_disposed_imports':
            search_q |= Q(disposal_reason__icontains=search_query)
        qs = qs.filter(search_q)

    # === FINAL DATA FOR EXPORT ===
    if scope == 'page':
        paginator = Paginator(qs, items_per_page)
        try:
            page_obj = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.page(1)
        data = list(page_obj.object_list)
    else:
        data = list(qs.iterator())

    # --- Response ---
    response = HttpResponse(content_type='application/pdf')
    filename = f"IT_Inventory_{'All' if scope == 'all' else 'Page'}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # --- Document Setup ---
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=10*mm, leftMargin=10*mm,
        topMargin=15*mm, bottomMargin=18*mm
    )
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    styles.add(ParagraphStyle(name='ReportTitle', fontSize=18, leading=22,
                              textColor=colors.HexColor('#143C50'), alignment=TA_CENTER, spaceAfter=6))
    styles.add(ParagraphStyle(name='SubTitle', fontSize=11, leading=13,
                              textColor=colors.HexColor('#143C50'), alignment=TA_CENTER, spaceAfter=10))
    styles.add(ParagraphStyle(name='Cell', fontSize=7.5, leading=9,
                              alignment=TA_LEFT))

    # Title
    title = 'MOHO IT Inventory Report' if request.user.is_superuser else f'{request.user.centre.name} IT Inventory Report'
    elements.append(Paragraph(title, styles['ReportTitle']))
    elements.append(Paragraph(
        f"Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles['SubTitle']
    ))
    elements.append(Spacer(1, 6*mm))

    # Table setup
    col_widths = [65, 65, 60, 65, 80, 95, 120, 60, 80, 80]
    headers = [
        'Centre', 'Department', 'Category', 'device_name', 'Model',
        'Specs & serial number', 'Assignee',
        'Condition', 'Status & Date', 'Disposal Reason'
    ]

    table_data = [headers]
    cell_style = styles['Cell']

    def safe(v):
        return str(v or 'N/A')

    for item in data:
        specs = (
            f"<b>RAM:</b> {safe(item.ram_gb)} GB<br/>"
            f"<b>HDD:</b> {safe(item.hdd_gb)} GB<br/>"
            f"<b>Serial:</b> {safe(item.serial_number)}"
        )
        assignee = (
            f"{safe(item.assignee_first_name)} {safe(item.assignee_last_name)}"
            f"<br/><font size=6>{safe(item.assignee_email_address)}</font>"
        )
        status_date = (
            f"<b>Status:</b> {safe(item.status)}<br/>"
            f"<b>Date:</b> {safe(item.date.strftime('%Y-%m-%d') if item.date else '')}"
        )

        row = [
            Paragraph(safe(item.centre.name if item.centre else ''), cell_style),
            Paragraph(safe(item.department.name if item.department else ''), cell_style),
            Paragraph(safe(item.get_category_display()), cell_style),
            Paragraph(safe(item.device_name), cell_style),
            Paragraph(safe(item.system_model), cell_style),
            Paragraph(specs, cell_style),
            Paragraph(assignee, cell_style),
            Paragraph(safe(item.device_condition), cell_style),
            Paragraph(status_date, cell_style),
            Paragraph(safe(item.disposal_reason), cell_style),
        ]
        table_data.append(row)

    if len(table_data) == 1:
        table_data.append([Paragraph('No records found.', cell_style)] * len(headers))

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#143C50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('FONTSIZE', (0,1), (-1,-1), 7.5),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,1), (-1,-1), 3),
        ('BOTTOMPADDING', (0,1), (-1,-1), 3),
    ]))
    elements.append(table)

    # Page numbering and watermark
    def add_page_elements(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(148.5 * mm, 6 * mm, f"Page {doc.page}")
        canvas.restoreState()
        canvas.saveState()
        canvas.setFont("Helvetica", 60)
        canvas.setFillGray(0.9, 0.15)
        canvas.rotate(45)
        canvas.drawCentredString(400, 100, "MOHI IT")
        canvas.restoreState()

    try:
        doc.build(
            elements,
            onFirstPage=add_page_elements,
            onLaterPages=add_page_elements
        )
    except Exception as e:
        logger.error(f"PDF generation failed: {e}")
        return HttpResponse(f"Error generating PDF: {e}", status=500)

    return response
    




@login_required
def import_add(request):
    user = request.user

    if request.method == 'POST':

        # ── New employee creation from modal ─────────────────────────────
        if request.POST.get('new_employee_submit') == '1':
            try:
                first_name     = (request.POST.get('new_first_name')     or '').strip()
                last_name      = (request.POST.get('new_last_name')      or '').strip()
                email          = (request.POST.get('new_email')          or '').strip().lower()
                staff_number   = (request.POST.get('new_staff_number')   or '').strip()
                designation    = (request.POST.get('new_designation')    or '').strip()
                department_id  = request.POST.get('new_department')
                centre_id      = request.POST.get('new_centre')

                if not first_name or not last_name:
                    messages.error(request, "First name and last name are required to create an employee.")
                    return redirect('import_add')

                # Check for existing employee by email (unique)
                if email:
                    if Employee.objects.filter(email__iexact=email).exists():
                        messages.warning(request, f"An employee with email {email} already exists.")
                        return redirect('import_add')

                # Check for existing by name (case-insensitive)
                if Employee.objects.filter(
                    first_name__iexact=first_name,
                    last_name__iexact=last_name
                ).exists():
                    messages.info(request, f"An employee named {first_name} {last_name} already exists.")
                    return redirect('import_add')

                # Optional relations
                department = None
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        pass

                centre = None
                if user.is_trainer:
                    if not user.centre:
                        messages.error(request, "Your account has no centre assigned.")
                        return redirect('import_add')
                    centre = user.centre
                elif centre_id:
                    try:
                        centre = Centre.objects.get(id=centre_id)
                    except Centre.DoesNotExist:
                        pass

                # Create
                employee = Employee.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    email=email or None,
                    staff_number=staff_number or None,
                    designation=designation or None,
                    department=department,
                    centre=centre,
                    is_active=True,
                )

                messages.success(request, f"Employee created successfully: {employee.full_name}")
                return redirect('import_add')

            except Exception as e:
                logger.exception("Failed to create new employee from modal")
                messages.error(request, f"Could not create employee: {str(e)}")
                return redirect('import_add')

        # ── Bulk CSV upload ──────────────────────────────────────────────
        if 'file' in request.FILES:
            file = request.FILES['file']
            if not file.name.lower().endswith('.csv'):
                messages.error(request, "Only CSV files are accepted.")
                return redirect('import_add')

            try:
                centre_id     = request.POST.get('bulk_centre')
                department_id = request.POST.get('bulk_department')
                category      = request.POST.get('bulk_category')

                if not department_id:
                    messages.error(request, "Please select a department.")
                    return redirect('import_add')

                if not category:
                    messages.error(request, "Please select a device category.")
                    return redirect('import_add')

                if user.is_trainer:
                    if not user.centre:
                        messages.error(request, "Your account has no centre assigned.")
                        return redirect('import_add')
                    centre = user.centre
                else:
                    if not centre_id:
                        messages.error(request, "Please select a centre.")
                        return redirect('import_add')
                    centre = Centre.objects.get(id=centre_id)

                department = Department.objects.get(id=department_id)

                stats = handle_uploaded_file(file, user, centre, department, category)

                approval_note = " (pending approval)" if user.is_trainer else ""
                messages.success(
                    request,
                    f"Imported {stats['created_count']} devices to {centre} – {department} ({category}){approval_note}"
                )

                # Combined summary email for assigned devices
                assigned_summary = ""
                assigned_count = 0
                for dev in Import.objects.filter(serial_number__in=stats['created_serials'], assignee__isnull=False):
                    assigned_summary += f"- SN: {dev.serial_number} ({dev.category}) assigned to {dev.assignee.full_name}\n"
                    assigned_count += 1

                if assigned_count > 0:
                    message = f"Bulk upload summary: {assigned_count} devices assigned.\n\n{assigned_summary}"
                    send_custom_email(
                        "Bulk Device Assignments Summary",
                        message,
                        ["it@mohiafrica.org"]
                    )
                    messages.info(request, "Summary email sent to IT for assigned devices.")

                if stats['skipped_existing']:
                    messages.warning(request, f"Skipped {stats['skipped_existing']} existing serial numbers.")
                if stats['skipped_validation']:
                    messages.warning(request, f"Skipped {stats['skipped_validation']} invalid rows.")

                return redirect('display_approved_imports')

            except Exception as e:
                logger.exception("Bulk upload failed")
                messages.error(request, f"Error processing upload: {str(e)}")
                return redirect('import_add')

        # ── Single device add ────────────────────────────────────────────
        else:
            try:
                with transaction.atomic():
                    centre_id     = request.POST.get('centre')
                    department_id = request.POST.get('department')
                    category      = request.POST.get('category')
                    serial_number = (request.POST.get('serial_number') or '').strip()
                    provided_device_name = (request.POST.get('device_name') or '').strip()
                    is_server = str(request.POST.get('is_server') or '').strip().lower() in {'1', 'true', 'yes', 'y', 'on'}

                    if not all([centre_id, department_id, category, serial_number]):
                        messages.error(request, "All required fields must be filled.")
                        return redirect('import_add')

                    centre = Centre.objects.get(id=centre_id) if not user.is_trainer else user.centre
                    if user.is_trainer and not centre:
                        messages.error(request, "Your account has no centre assigned.")
                        return redirect('import_add')
                    department = Department.objects.get(id=department_id)

                    assignee_id = request.POST.get('assignee')
                    assignee = None
                    if assignee_id and assignee_id.strip():
                        assignee = assignment_employee_queryset(user).filter(id=assignee_id).first()
                        if assignee is None:
                            messages.error(request, "Invalid assignee selected.")
                            return redirect('import_add')

                    kind = None
                    if category == "laptop":
                        kind = "laptop"
                    elif category == "system_unit":
                        kind = "server" if is_server else "desktop"

                    generated_device_name = _next_asset_tag(kind) if kind else None
                    final_device_name = generated_device_name or provided_device_name
                    serial_number, final_device_name = _validate_device_identity(
                        serial_number=serial_number,
                        device_name=final_device_name,
                    )

                    device = Import(
                        added_by=user,
                        centre=centre,
                        department=department,
                        category=category,
                        device_name=final_device_name,
                        system_model=request.POST.get('system_model'),
                        processor=request.POST.get('processor'),
                        ram_gb=request.POST.get('ram_gb'),
                        hdd_gb=request.POST.get('hdd_gb'),
                        serial_number=serial_number,
                        assignee=assignee,
                        device_condition=request.POST.get('device_condition'),
                        status=request.POST.get('status'),
                        date=timezone.now().date(),
                        is_approved=not user.is_trainer,
                        approved_by=user if not user.is_trainer and user.is_superuser else None
                    )
                    device.save()

                    if assignee:
                        # Create agreement
                        DeviceAgreement.objects.create(
                            device=device,
                            employee=assignee,
                        )

                    if kind in {"laptop", "desktop", "server"}:
                        _ensure_device_configuration_items(device)
                        messages.success(request, f"Device {serial_number} added. Start configuration checks to finish setup.")
                        return redirect('device_detail', pk=device.pk)

                    messages.success(request, f"Device {serial_number} added successfully.")
                    return redirect('display_approved_imports')

            except Exception as e:
                logger.exception("Single device add failed")
                messages.error(request, f"Error: {str(e)}")
                return redirect('import_add')

    # GET – show form
    centres = Centre.objects.all().order_by('name')
    if user.is_trainer:
        centres = centres.filter(pk=user.centre_id) if user.centre_id else centres.none()
    departments = Department.objects.all().order_by('name')
    employees = assignment_employee_queryset(user)

    context = {
        'centres': centres,
        'departments': departments,
        'employees': employees,
        'category_choices': Import.CATEGORY_CHOICES,
    }
    return render(request, 'import/add.html', context)


def handle_uploaded_file(file, user, centre, department, category):
    stats = {
        'total_rows': 0,
        'created_count': 0,
        'assigned_count': 0,
        'skipped_existing': 0,
        'skipped_validation': 0,
        'created_serials': [],  # for summary email
    }

    header_mapping = {
        'device_name': 'device_name',
        'system_model': 'system_model',
        'processor': 'processor',
        'ram_gb': 'ram_gb',
        'hdd_gb': 'hdd_gb',
        'serial_number': 'serial_number',
        'assignee_first_name': 'assignee_first_name',
        'assignee_last_name': 'assignee_last_name',
        'assignee_email_address': 'assignee_email_address',
        'device_condition': 'device_condition',
        'status': 'status',
        'date': 'date',
    }

    try:
        file.seek(0)
        decoded = TextIOWrapper(file, encoding='utf-8-sig')
        reader = csv.reader(decoded)
        headers = [h.lower().strip() for h in next(reader, [])]

        missing_headers = [
            required_header
            for required_header in ('serial_number', 'device_name')
            if required_header not in headers
        ]
        if missing_headers:
            raise ValueError(
                "CSV missing required column(s): " + ", ".join(missing_headers)
            )

        sn_idx = headers.index('serial_number')
        device_name_idx = headers.index('device_name')

        devices_to_create = []
        seen_serials = set()
        admins = CustomUser.objects.filter(
            is_trainer=False
        ).filter(
            Q(is_superuser=True) | Q(is_it_manager=True) | Q(is_senior_it_officer=True)
        )

        for row in reader:
            if not any(row):
                continue
            stats['total_rows'] += 1

            sn = (row[sn_idx] or '').strip()
            device_name = (row[device_name_idx] or '').strip() if device_name_idx < len(row) else ''
            if not sn or not device_name:
                stats['skipped_validation'] += 1
                continue

            normalized_serial = sn.casefold()
            if normalized_serial in seen_serials:
                stats['skipped_existing'] += 1
                continue

            if Import.objects.filter(serial_number__iexact=sn).exists():
                stats['skipped_existing'] += 1
                continue
            seen_serials.add(normalized_serial)

            device = Import(
                added_by=user,
                centre=centre,
                department=department,
                category=category,
                serial_number=sn,
                device_name=device_name,
                is_approved=not user.is_trainer,
                approved_by=user if not user.is_trainer and user.is_superuser else None,
                date=timezone.now().date(),
            )

            for h, value in zip(headers, row):
                value = (value or '').strip()
                field = header_mapping.get(h)
                if field and field not in {'serial_number', 'device_name'}:
                    if field == 'date' and value:
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                            try:
                                device.date = datetime.strptime(value, fmt).date()
                                break
                            except ValueError:
                                continue
                    else:
                        setattr(device, field, value or None)

            # Employee assignment logic
            first = (getattr(device, 'assignee_first_name', '') or '').strip()
            last = (getattr(device, 'assignee_last_name', '') or '').strip()
            email = (getattr(device, 'assignee_email_address', '') or '').strip().lower()

            employee = None
            if email:
                employee = Employee.objects.filter(email__iexact=email).first()

            if not employee and first and last:
                employee = Employee.objects.filter(
                    first_name__iexact=first,
                    last_name__iexact=last
                ).first()

            if not employee and first and last:
                employee = Employee.objects.create(
                    first_name=first,
                    last_name=last,
                    email=email or None,
                )

            if employee:
                device.assignee = employee
                stats['assigned_count'] = stats.get('assigned_count', 0) + 1
                stats['created_serials'].append(sn)

            devices_to_create.append(device)

        if devices_to_create:
            with transaction.atomic():
                created = Import.objects.bulk_create(devices_to_create, batch_size=400)
                stats['created_count'] = len(created)

                created_devices = list(
                    Import.objects.filter(id__in=[d.id for d in created if d.id]).select_related('assignee')
                )

                for dev in created_devices:
                    if dev.assignee:
                        DeviceAgreement.objects.get_or_create(
                            device=dev,
                            employee=dev.assignee,
                            is_archived=False,
                            defaults={"issuance_it_user": user},
                        )

                # Send emails
                for dev in created_devices:
                    if dev.assignee and dev.assignee.email:
                        send_device_assignment_email(dev, action='assigned')

                # Trainer notifications
                if user.is_trainer:
                    for dev in created_devices:
                        for admin in admins:
                            Notification.objects.create(
                                user=admin,
                                message=f"Bulk upload – new device {dev.serial_number} awaiting approval.",
                                content_type=ContentType.objects.get_for_model(Import),
                                object_id=dev.pk
                            )

        return stats

    except Exception as e:
        logger.exception("CSV processing failed")
        raise ValueError(f"CSV processing error: {str(e)}")

@login_required
def download_csv_template(request):
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="device_import_template.csv"'},
    )
    writer = csv.writer(response)
    writer.writerow([
        'serial_number',
        'device_name',
        'system_model',
        'processor',
        'ram_gb',
        'hdd_gb',
        'assignee_first_name',
        'assignee_last_name',
        'assignee_email_address',
        'device_condition',
        'status',
        'date',
    ])
    return response

def _apply_pending_update_to_import(import_instance, pending_update, approved_by):
    old_assignee = import_instance.assignee

    if pending_update.centre is not None:
        import_instance.centre = pending_update.centre
    if pending_update.department is not None:
        import_instance.department = pending_update.department
    if getattr(pending_update, "category", None):
        import_instance.category = pending_update.category
    if pending_update.device_name is not None:
        import_instance.device_name = pending_update.device_name
    if pending_update.system_model is not None:
        import_instance.system_model = pending_update.system_model
    if pending_update.processor is not None:
        import_instance.processor = pending_update.processor
    if pending_update.ram_gb is not None:
        import_instance.ram_gb = pending_update.ram_gb
    if pending_update.hdd_gb is not None:
        import_instance.hdd_gb = pending_update.hdd_gb
    if getattr(pending_update, "serial_number", None):
        import_instance.serial_number = pending_update.serial_number
    if getattr(pending_update, "assignee", None) is not None:
        import_instance.assignee = pending_update.assignee
    if pending_update.assignee_first_name is not None:
        import_instance.assignee_first_name = pending_update.assignee_first_name
    if pending_update.assignee_last_name is not None:
        import_instance.assignee_last_name = pending_update.assignee_last_name
    if pending_update.assignee_email_address is not None:
        import_instance.assignee_email_address = pending_update.assignee_email_address
    if pending_update.device_condition is not None:
        import_instance.device_condition = pending_update.device_condition
    if pending_update.status is not None:
        import_instance.status = pending_update.status
    if pending_update.date is not None:
        import_instance.date = pending_update.date
    if pending_update.reason_for_update is not None:
        import_instance.reason_for_update = pending_update.reason_for_update

    if (
        import_instance.serial_number
        and Import.objects.filter(serial_number__iexact=import_instance.serial_number)
        .exclude(pk=import_instance.pk)
        .exists()
    ):
        raise ValueError(f"Serial number {import_instance.serial_number} is already used by another device.")

    import_instance.is_approved = True
    import_instance.approved_by = approved_by
    import_instance.save()
    _sync_device_assignment_agreement(
        import_instance,
        old_assignee=old_assignee,
        new_assignee=import_instance.assignee,
        actor=approved_by,
    )
    return old_assignee


@login_required
@user_passes_test(lambda u: u.is_superuser and not u.is_trainer)
def import_approve(request, pk):
    import_instance = get_object_or_404(Import, pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            pending_update = PendingUpdate.objects.filter(import_record=import_instance).order_by('-created_at').first()

            def _mark_related_notifications_read(*, pending_update_id=None, trainer_user_id=None):
                q = Q(
                    content_type=ContentType.objects.get_for_model(Import),
                    object_id=import_instance.pk,
                )
                if pending_update_id:
                    q |= Q(
                        content_type=ContentType.objects.get_for_model(PendingUpdate),
                        object_id=pending_update_id,
                    )

                qs = Notification.objects.filter(q, is_read=False).filter(
                    Q(user__is_superuser=True) | Q(user__is_it_manager=True) | Q(user__is_senior_it_officer=True)
                )
                if trainer_user_id:
                    qs = qs.exclude(user_id=trainer_user_id)
                qs.update(is_read=True, responded_by=request.user)

            if pending_update:
                # Save pk before deleting
                pending_update_id = pending_update.pk
                trainer_user_id = pending_update.updated_by_id
                try:
                    old_assignee = _apply_pending_update_to_import(
                        import_instance=import_instance,
                        pending_update=pending_update,
                        approved_by=request.user,
                    )
                except ValueError as exc:
                    messages.error(request, str(exc))
                    return redirect('display_unapproved_imports')
                # Check if assignee changed
                if import_instance.assignee != old_assignee:
                    # Email to old assignee (cleared)
                    if old_assignee and old_assignee.email:
                        send_device_assignment_email(import_instance, action='cleared', cleared_by=request.user)
                    # Email to new assignee (issued)
                    if import_instance.assignee and import_instance.assignee.email:
                        send_device_assignment_email(import_instance, action='assigned')
                    # Email to IT (transfer summary)
                    if old_assignee and import_instance.assignee:
                        send_device_assignment_email(import_instance, action='transferred', cleared_by=request.user)
                # Delete pending update after saving
                pending_update.delete()
                _mark_related_notifications_read(
                    pending_update_id=pending_update_id,
                    trainer_user_id=trainer_user_id,
                )
                messages.success(request, f"Device {import_instance.serial_number} update approved.")
            else:
                import_instance.is_approved = True
                import_instance.approved_by = request.user
                import_instance.save()
                _mark_related_notifications_read()
                messages.success(request, f"Device {import_instance.serial_number} approved.")
            return redirect('display_unapproved_imports')
    return redirect('display_unapproved_imports')

@login_required
@user_passes_test(lambda u: u.is_superuser and not u.is_trainer)
def import_reject(request, pk):
    import_instance = get_object_or_404(Import, pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            pending_update = PendingUpdate.objects.filter(import_record=import_instance).order_by('-created_at').first()

            def _mark_related_notifications_read(*, pending_update_id=None, trainer_user_id=None):
                q = Q(
                    content_type=ContentType.objects.get_for_model(Import),
                    object_id=import_instance.pk,
                )
                if pending_update_id:
                    q |= Q(
                        content_type=ContentType.objects.get_for_model(PendingUpdate),
                        object_id=pending_update_id,
                    )

                qs = Notification.objects.filter(q, is_read=False).filter(
                    Q(user__is_superuser=True) | Q(user__is_it_manager=True) | Q(user__is_senior_it_officer=True)
                )
                if trainer_user_id:
                    qs = qs.exclude(user_id=trainer_user_id)
                qs.update(is_read=True, responded_by=request.user)

            if pending_update:
                pending_update.pending_clarification = True
                pending_update.save()
                # Notify trainer for clarification (single notification)
                trainer = pending_update.updated_by
                if trainer:
                    content_type = ContentType.objects.get_for_model(PendingUpdate)
                    if not Notification.objects.filter(user=trainer, content_type=content_type, object_id=pending_update.pk).exists():
                        Notification.objects.create(
                            user=trainer,
                            message=f"Your update for device {pending_update.serial_number} was rejected. Please provide clarification.",
                            content_type=content_type,
                            object_id=pending_update.pk
                        )
                _mark_related_notifications_read(
                    pending_update_id=pending_update.pk,
                    trainer_user_id=getattr(trainer, "id", None),
                )
                messages.success(request, f"Update for device {pending_update.serial_number} sent back for clarification.")
            else:
                # For new import requests, mark as pending clarification
                import_instance.pending_clarification = True
                import_instance.save()
                # Notify trainer for clarification (single notification)
                trainer = import_instance.added_by
                if trainer:
                    content_type = ContentType.objects.get_for_model(Import)
                    if not Notification.objects.filter(user=trainer, content_type=content_type, object_id=import_instance.pk).exists():
                        Notification.objects.create(
                            user=trainer,
                            message=f"Your import request for device {import_instance.serial_number} was rejected. Please provide clarification.",
                            content_type=content_type,
                            object_id=import_instance.pk
                        )
                _mark_related_notifications_read(trainer_user_id=getattr(trainer, "id", None))
                messages.success(request, f"Import request for device {import_instance.serial_number} sent back for clarification.")
            return redirect('display_unapproved_imports')
    return redirect('display_unapproved_imports')

@login_required
@user_passes_test(lambda u: u.is_superuser and not u.is_trainer)
def import_approve_all(request):
    if request.method == 'POST':
        page_number = request.GET.get('page', '1')
        items_per_page = request.GET.get('items_per_page', '10')
        search_query = request.GET.get('search', '')
       
        try:
            items_per_page = int(items_per_page)
            if items_per_page not in [10, 25, 50, 100, 500]:
                items_per_page = 10
        except ValueError:
            items_per_page = 10
        try:
            page_number = int(page_number) if page_number else 1
        except ValueError:
            page_number = 1
        data = Import.objects.filter(is_approved=False, is_disposed=False).order_by('-pk')
        if search_query:
            query = (
                Q(centre__name__icontains=search_query) |
                Q(centre__centre_code__icontains=search_query) |
                Q(department__name__icontains=search_query) |
                Q(device_name__icontains=search_query) |
                Q(system_model__icontains=search_query) |
                Q(processor__icontains=search_query) |
                Q(ram_gb__icontains=search_query) |
                Q(hdd_gb__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(assignee__first_name__icontains=search_query) |
                Q(assignee__last_name__icontains=search_query) |
                Q(assignee__email__icontains=search_query) |
                Q(assignee__staff_number__icontains=search_query) |
                Q(device_condition__icontains=search_query) |
                Q(status__icontains=search_query) |
                Q(date__icontains=search_query) |
                Q(reason_for_update__icontains=search_query)
            )
            data = data.filter(query)
        paginator = Paginator(data, items_per_page)
        try:
            data_on_page = paginator.page(page_number)
        except PageNotAnInteger:
            data_on_page = paginator.page(1)
        except EmptyPage:
            data_on_page = paginator.page(paginator.num_pages)
        approved_count = 0
        skipped_conflicts = []
        with transaction.atomic():
            for item in data_on_page:
                pending_update = PendingUpdate.objects.filter(import_record=item).order_by('-created_at').first()
                if pending_update:
                    try:
                        old_assignee = _apply_pending_update_to_import(
                            import_instance=item,
                            pending_update=pending_update,
                            approved_by=request.user,
                        )
                    except ValueError as exc:
                        skipped_conflicts.append(str(exc))
                        continue
                    # Check if assignee changed
                    if pending_update.assignee != old_assignee:
                        # Email to old assignee (cleared)
                        if old_assignee and old_assignee.email:
                            send_device_assignment_email(item, action='cleared', cleared_by=request.user)
                        # Email to new assignee (issued)
                        if item.assignee and item.assignee.email:
                            send_device_assignment_email(item, action='assigned')
                        # Email to IT (transfer summary)
                        if old_assignee and item.assignee:
                            send_device_assignment_email(item, action='transferred', cleared_by=request.user)
                    pending_update.delete()
                    content_type = ContentType.objects.get_for_model(PendingUpdate)
                    Notification.objects.filter(
                        content_type=content_type,
                        object_id=pending_update.pk,
                        user__is_trainer=False,
                        is_read=False
                    ).filter(
                        Q(user__is_superuser=True) | Q(user__is_it_manager=True) | Q(user__is_senior_it_officer=True)
                    ).update(is_read=True, responded_by=request.user)
                    approved_count += 1
                elif not item.is_approved:
                    item.is_approved = True
                    item.approved_by = request.user
                    item.save()
                    content_type = ContentType.objects.get_for_model(Import)
                    Notification.objects.filter(
                        content_type=content_type,
                        object_id=item.pk,
                        user__is_trainer=False,
                        is_read=False
                    ).filter(
                        Q(user__is_superuser=True) | Q(user__is_it_manager=True) | Q(user__is_senior_it_officer=True)
                    ).update(is_read=True, responded_by=request.user)
                    approved_count += 1
        if approved_count > 0:
            messages.success(request, f"{approved_count} device(s) approved successfully.")
        else:
            messages.info(request, "No unapproved devices to approve on this page.")
        if skipped_conflicts:
            messages.warning(request, skipped_conflicts[0])
       
        redirect_url = reverse('display_unapproved_imports')
        query_params = [f"page={page_number}", f"items_per_page={items_per_page}"]
        if search_query:
            query_params.append(f"search={search_query}")
        redirect_url += "?" + "&".join(query_params)
        return redirect(redirect_url)
    return redirect('display_unapproved_imports')
def _can_delete(user):
    """Only IT Manager or Senior IT Officer can delete."""
    return user.is_it_manager or user.is_senior_it_officer


@login_required
def import_delete(request, pk):
    import_instance = get_object_or_404(Import, pk=pk)

    if request.user.is_trainer and import_instance.centre != request.user.centre:
        messages.error(request, "You can only manage devices from your own centre.")
        return redirect('display_approved_imports')

    if request.method != 'POST':
        return redirect('display_approved_imports')

    if request.user.is_trainer:
        delete_reason = (request.POST.get('delete_reason') or '').strip()
        if not delete_reason:
            messages.error(request, "Deletion reason is required before requesting approval.")
            return redirect('display_approved_imports')
        if len(delete_reason) < 10:
            messages.error(request, "Deletion reason must be at least 10 characters.")
            return redirect('display_approved_imports')

        deletion_request, _ = DeviceDeletionRequest.objects.update_or_create(
            device=import_instance,
            defaults={
                "reason": delete_reason,
                "requested_by": request.user,
            },
        )
        _notify_device_request_reviewers(
            device=import_instance,
            message=(
                f"Delete request for device {import_instance.serial_number} by "
                f"{request.user.get_full_name() or request.user.username} awaiting approval."
            ),
            related_object=deletion_request,
        )
        messages.success(request, f"Delete request for {import_instance.serial_number} submitted for approval.")
        return redirect('display_unapproved_imports')

    if not can_delete_devices(request.user):
        messages.error(request, "You do not have permission to delete devices.")
        return redirect('display_approved_imports')

    with transaction.atomic():
        serial_number = _perform_device_delete(import_instance=import_instance, actor=request.user)
        messages.success(request, f"Device {serial_number} deleted successfully. IT notified.")
    return redirect('display_approved_imports')


@login_required
def import_update(request, pk):
    """Update existing device - assignee can only be set if currently None"""
    device = get_object_or_404(Import, pk=pk)
    if not can_manage_device_assignments(request.user):
        messages.error(request, "You do not have permission to update device assignments.")
        return redirect('device_detail', pk=device.pk)
    # Permission: trainers can only edit their centre's devices
    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only update records for your own centre.")
        return redirect('display_approved_imports')
    blocked = _block_actions_if_inactive(request, device)
    if blocked:
        return blocked
    if request.method == 'POST':
        # Handle new employee creation from modal
        if request.POST.get('new_employee_submit') == '1':
            try:
                first_name = (request.POST.get('new_first_name') or '').strip()
                last_name = (request.POST.get('new_last_name') or '').strip()
                email = (request.POST.get('new_email') or '').strip().lower()
                staff_number = (request.POST.get('new_staff_number') or '').strip()
                department_id = request.POST.get('new_department')
                centre_id = request.POST.get('new_centre')
                if not first_name or not last_name:
                    messages.error(request, "First name and last name are required.")
                    return redirect('import_update', pk=pk)
                # Email uniqueness
                if email and Employee.objects.filter(email__iexact=email).exists():
                    messages.warning(request, f"Email {email} is already used by another employee.")
                    return redirect('import_update', pk=pk)
                # Name uniqueness (soft check)
                if Employee.objects.filter(first_name__iexact=first_name, last_name__iexact=last_name).exists():
                    messages.info(request, f"An employee named {first_name} {last_name} already exists.")
                    return redirect('import_update', pk=pk)
                department = None
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        pass
                centre = None
                if request.user.is_trainer:
                    if not request.user.centre:
                        messages.error(request, "Your account has no centre assigned.")
                        return redirect('import_update', pk=pk)
                    centre = request.user.centre
                elif centre_id:
                    try:
                        centre = Centre.objects.get(id=centre_id)
                    except Centre.DoesNotExist:
                        pass
                new_employee = Employee.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    email=email or None,
                    staff_number=staff_number or None,
                    department=department,
                    centre=centre,
                    is_active=True,
                )
                messages.success(request, f"New employee created: {new_employee.full_name}")
                # Auto-select the new one after reload
                return redirect(f"{reverse('import_update', kwargs={'pk': pk})}?new_employee={new_employee.id}")
            except Exception as e:
                logger.exception("Failed to create employee")
                messages.error(request, f"Could not create employee: {str(e)}")
                return redirect('import_update', pk=pk)
        # Normal update
        try:
            with transaction.atomic():
                department_id = request.POST.get('department')
                category = request.POST.get('category')
                serial_number = (request.POST.get('serial_number') or '').strip()
                # Centre handling
                centre = device.centre
                if not request.user.is_trainer:
                    centre_id = request.POST.get('centre')
                    if centre_id:
                        try:
                            centre = Centre.objects.get(id=centre_id)
                        except Centre.DoesNotExist:
                            messages.error(request, "Invalid centre selected.")
                            return redirect('import_update', pk=pk)
                # Required fields
                department = device.department
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        messages.error(request, "Invalid department selected.")
                        return redirect('import_update', pk=pk)
                if not category:
                    messages.error(request, "Category is required.")
                    return redirect('import_update', pk=pk)
                if serial_number and Import.objects.filter(serial_number=serial_number).exclude(pk=pk).exists():
                    messages.error(request, f"Serial number {serial_number} is already used by another device.")
                    return redirect('import_update', pk=pk)
                # Date
                date_value = device.date
                date_str = request.POST.get('date', '').strip()
                if date_str:
                    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                        try:
                            date_value = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                # Assignee logic - only allow setting if currently None
                new_assignee = device.assignee
                assignee_id = request.POST.get('assignee', '').strip()
                if assignee_id and not device.assignee:  # only allow assignment if empty
                    try:
                        new_assignee = Employee.objects.get(id=assignee_id)
                    except (Employee.DoesNotExist, ValueError):
                        pass
                elif assignee_id and device.assignee:
                    # Attempt to change existing assignee → block
                    if assignee_id != str(device.assignee.id):
                        messages.error(request, "Cannot change assignee. Clear the current user first.")
                        return redirect('import_update', pk=pk)
                # Collect changes
                fields_to_update = {}
                form_data = {
                    'centre': centre,
                    'department': department,
                    'category': category,
                    'device_name': request.POST.get('device_name', '').strip(),
                    'system_model': request.POST.get('system_model', '').strip(),
                    'processor': request.POST.get('processor', '').strip(),
                    'ram_gb': request.POST.get('ram_gb', '').strip(),
                    'hdd_gb': request.POST.get('hdd_gb', '').strip(),
                    'serial_number': serial_number,
                    'device_condition': request.POST.get('device_condition', '').strip(),
                    'status': request.POST.get('status', '').strip(),
                    'reason_for_update': request.POST.get('reason_for_update', '').strip(),
                    'date': date_value,
                    'assignee': new_assignee,
                }
                for field, new_val in form_data.items():
                    old_val = getattr(device, field)
                    if new_val != old_val:
                        fields_to_update[field] = new_val
                if not fields_to_update:
                    messages.info(request, "No changes detected.")
                    return redirect('import_update', pk=pk)
                # Trainer → pending approval
                if request.user.is_trainer:
                    reason = request.POST.get('reason_for_update', '').strip()
                    if not reason:
                        messages.error(request, "Reason for update is required for trainers.")
                        return redirect('import_update', pk=pk)
                    if len(reason) < 10:
                        messages.error(request, "Reason for update must be at least 10 characters for trainers.")
                        return redirect('import_update', pk=pk)
                    PendingUpdate.objects.create(
                        import_record=device,
                        **fields_to_update,
                        updated_by=request.user
                    )
                    device.is_approved = False
                    device.approved_by = None
                    device.save()
                    # Notify admins
                    admins = CustomUser.objects.filter(is_trainer=False).filter(
                        Q(is_superuser=True) | Q(is_it_manager=True) | Q(is_senior_it_officer=True)
                    )
                    for admin in admins:
                        Notification.objects.create(
                            user=admin,
                            message=f"Update request for device {device.serial_number} by {request.user} — awaiting approval.",
                            content_type=ContentType.objects.get_for_model(Import),
                            object_id=device.pk
                        )
                    messages.success(request, "Update request submitted for approval.")
                    return redirect('notifications_view')
                # Admin / superuser → direct update
                old_assignee = device.assignee
                for field, value in fields_to_update.items():
                    setattr(device, field, value)
                device.is_approved = True if request.user.is_superuser else device.is_approved
                device.approved_by = request.user if request.user.is_superuser else device.approved_by
                device.save()
                # Send assignment email only if we actually assigned someone
                if 'assignee' in fields_to_update and not old_assignee and device.assignee:
                    if device.assignee.email:
                        send_device_assignment_email(device, action='assigned')
                messages.success(request, "Device updated successfully.")
                return redirect('import_update', pk=pk)
        except Exception as e:
            logger.exception(f"Update failed for device {pk}")
            messages.error(request, f"Update failed: {str(e)}")
            return redirect('import_update', pk=pk)
    # GET - show form
    employees = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
    centres = Centre.objects.all().order_by('name')
    departments = Department.objects.all().order_by('name')
    # Auto-select newly created employee if redirected from modal
    new_employee_id = request.GET.get('new_employee')
    pre_selected_assignee = None
    if new_employee_id:
        try:
            pre_selected_assignee = Employee.objects.get(id=new_employee_id)
        except Employee.DoesNotExist:
            pass
    context = {
        'import_instance': device,
        'centres': centres,
        'departments': departments,
        'employees': employees,
        'category_choices': Import.CATEGORY_CHOICES,
        'pre_selected_assignee': pre_selected_assignee,
    }
    return render(request, 'import/update.html', context)

@login_required
def import_update(request, pk):
    """Update an existing device and allow trainer reassignment within the trainer's centre."""
    device = get_object_or_404(Import, pk=pk)
    if not can_manage_device_assignments(request.user):
        messages.error(request, "You do not have permission to update device assignments.")
        return redirect('device_detail', pk=device.pk)

    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only update records for your own centre.")
        return redirect('display_approved_imports')

    blocked = _block_actions_if_inactive(request, device)
    if blocked:
        return blocked

    can_trainer_reassign = _can_trainer_reassign_device(request.user, device)

    if request.method == 'POST':
        if request.POST.get('new_employee_submit') == '1':
            try:
                first_name = (request.POST.get('new_first_name') or '').strip()
                last_name = (request.POST.get('new_last_name') or '').strip()
                email = (request.POST.get('new_email') or '').strip().lower()
                staff_number = (request.POST.get('new_staff_number') or '').strip()
                department_id = request.POST.get('new_department')
                centre_id = request.POST.get('new_centre')

                if not first_name or not last_name:
                    messages.error(request, "First name and last name are required.")
                    return redirect('import_update', pk=pk)
                if email and Employee.objects.filter(email__iexact=email).exists():
                    messages.warning(request, f"Email {email} is already used by another employee.")
                    return redirect('import_update', pk=pk)
                if Employee.objects.filter(first_name__iexact=first_name, last_name__iexact=last_name).exists():
                    messages.info(request, f"An employee named {first_name} {last_name} already exists.")
                    return redirect('import_update', pk=pk)

                department = None
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        pass

                centre = None
                if request.user.is_trainer:
                    if not request.user.centre:
                        messages.error(request, "Your account has no centre assigned.")
                        return redirect('import_update', pk=pk)
                    centre = request.user.centre
                elif centre_id:
                    try:
                        centre = Centre.objects.get(id=centre_id)
                    except Centre.DoesNotExist:
                        pass

                new_employee = Employee.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    email=email or None,
                    staff_number=staff_number or None,
                    department=department,
                    centre=centre,
                    is_active=True,
                )
                messages.success(request, f"New employee created: {new_employee.full_name}")
                return redirect(f"{reverse('import_update', kwargs={'pk': pk})}?new_employee={new_employee.id}")
            except Exception as e:
                logger.exception("Failed to create employee")
                messages.error(request, f"Could not create employee: {str(e)}")
                return redirect('import_update', pk=pk)

        try:
            with transaction.atomic():
                department_id = request.POST.get('department')
                category = request.POST.get('category')
                serial_number = (request.POST.get('serial_number') or '').strip()

                centre = device.centre
                if not request.user.is_trainer:
                    centre_id = request.POST.get('centre')
                    if centre_id:
                        try:
                            centre = Centre.objects.get(id=centre_id)
                        except Centre.DoesNotExist:
                            messages.error(request, "Invalid centre selected.")
                            return redirect('import_update', pk=pk)

                department = device.department
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        messages.error(request, "Invalid department selected.")
                        return redirect('import_update', pk=pk)

                if not category:
                    messages.error(request, "Category is required.")
                    return redirect('import_update', pk=pk)
                if serial_number and Import.objects.filter(serial_number__iexact=serial_number).exclude(pk=pk).exists():
                    messages.error(request, f"Serial number {serial_number} is already used by another device.")
                    return redirect('import_update', pk=pk)

                date_value = device.date
                date_str = request.POST.get('date', '').strip()
                if date_str:
                    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                        try:
                            date_value = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue

                new_assignee = device.assignee
                assignee_id = request.POST.get('assignee', '').strip()
                if assignee_id:
                    candidate_assignee = assignment_employee_queryset(request.user).filter(id=assignee_id).first()
                    if candidate_assignee is None:
                        messages.error(request, "Invalid assignee selected.")
                        return redirect('import_update', pk=pk)

                    if not device.assignee:
                        new_assignee = candidate_assignee
                    elif assignee_id == str(device.assignee.id):
                        new_assignee = device.assignee
                    elif can_trainer_reassign:
                        new_assignee = candidate_assignee
                    else:
                        messages.error(request, "Cannot change assignee. Clear the current user first.")
                        return redirect('import_update', pk=pk)

                fields_to_update = {}
                form_data = {
                    'centre': centre,
                    'department': department,
                    'category': category,
                    'device_name': request.POST.get('device_name', '').strip(),
                    'system_model': request.POST.get('system_model', '').strip(),
                    'processor': request.POST.get('processor', '').strip(),
                    'ram_gb': request.POST.get('ram_gb', '').strip(),
                    'hdd_gb': request.POST.get('hdd_gb', '').strip(),
                    'serial_number': serial_number,
                    'device_condition': request.POST.get('device_condition', '').strip(),
                    'status': request.POST.get('status', '').strip(),
                    'reason_for_update': request.POST.get('reason_for_update', '').strip(),
                    'date': date_value,
                    'assignee': new_assignee,
                }
                for field, new_val in form_data.items():
                    old_val = getattr(device, field)
                    if new_val != old_val:
                        fields_to_update[field] = new_val

                if not fields_to_update:
                    messages.info(request, "No changes detected.")
                    return redirect('import_update', pk=pk)

                if request.user.is_trainer:
                    reason = request.POST.get('reason_for_update', '').strip()
                    if not reason:
                        messages.error(request, "Reason for update is required for trainers.")
                        return redirect('import_update', pk=pk)
                    if len(reason) < 10:
                        messages.error(request, "Reason for update must be at least 10 characters for trainers.")
                        return redirect('import_update', pk=pk)

                    PendingUpdate.objects.create(
                        import_record=device,
                        **fields_to_update,
                        updated_by=request.user
                    )
                    device.is_approved = False
                    device.approved_by = None
                    device.save(update_fields=['is_approved', 'approved_by'])

                    admins = []
                    for admin in admins:
                        Notification.objects.create(
                            user=admin,
                            message=f"Update request for device {device.serial_number} by {request.user} â€” awaiting approval.",
                            content_type=ContentType.objects.get_for_model(Import),
                            object_id=device.pk
                        )
                    messages.success(request, "Update request submitted for approval.")
                    return redirect('notifications_view')

                old_assignee = device.assignee
                for field, value in fields_to_update.items():
                    setattr(device, field, value)
                device.is_approved = True if request.user.is_superuser else device.is_approved
                device.approved_by = request.user if request.user.is_superuser else device.approved_by
                device.save()

                _sync_device_assignment_agreement(
                    device,
                    old_assignee=old_assignee,
                    new_assignee=device.assignee,
                    actor=request.user,
                )

                if 'assignee' in fields_to_update and old_assignee != device.assignee:
                    if old_assignee and old_assignee.email:
                        send_device_assignment_email(device, action='cleared', cleared_by=request.user)
                    if device.assignee and device.assignee.email:
                        send_device_assignment_email(device, action='assigned')
                    if old_assignee and device.assignee:
                        send_device_assignment_email(device, action='transferred', cleared_by=request.user)

                messages.success(request, "Device updated successfully.")
                return redirect('import_update', pk=pk)
        except Exception as e:
            logger.exception(f"Update failed for device {pk}")
            messages.error(request, f"Update failed: {str(e)}")
            return redirect('import_update', pk=pk)

    employees = assignment_employee_queryset(request.user)
    centres = Centre.objects.all().order_by('name')
    if request.user.is_trainer:
        centres = centres.filter(pk=request.user.centre_id) if request.user.centre_id else centres.none()
    departments = Department.objects.all().order_by('name')

    new_employee_id = request.GET.get('new_employee')
    pre_selected_assignee = None
    if new_employee_id:
        pre_selected_assignee = employees.filter(id=new_employee_id).first()

    context = {
        'import_instance': device,
        'centres': centres,
        'departments': departments,
        'employees': employees,
        'category_choices': Import.CATEGORY_CHOICES,
        'pre_selected_assignee': pre_selected_assignee,
        'can_clear_user': can_clear_device_users(request.user),
        'can_trainer_reassign': can_trainer_reassign,
    }
    return render(request, 'import/update.html', context)


@login_required
def import_approve(request, pk):
    if not can_review_device_requests(request.user):
        messages.error(request, "You do not have permission to approve device requests.")
        return redirect('display_unapproved_imports')

    import_instance = get_object_or_404(Import, pk=pk)
    if request.method != 'POST':
        return redirect('display_unapproved_imports')

    reviewer_filter = (
        Q(user__is_superuser=True)
        | Q(user__is_it_manager=True)
        | Q(user__is_senior_it_officer=True)
        | (Q(user__is_staff=True) & Q(user__is_trainer=False))
    )

    with transaction.atomic():
        deletion_request = DeviceDeletionRequest.objects.filter(device=import_instance).select_related('requested_by').first()
        pending_update = PendingUpdate.objects.filter(import_record=import_instance).order_by('-created_at').first()

        if deletion_request:
            requester = deletion_request.requested_by
            deletion_request_id = deletion_request.pk
            serial_number = _perform_device_delete(import_instance=import_instance, actor=request.user)
            if requester:
                Notification.objects.create(
                    user=requester,
                    message=f"Your delete request for device {serial_number} was approved.",
                    content_type=ContentType.objects.get_for_model(Import),
                    object_id=pk,
                )
            Notification.objects.filter(
                Q(content_type=ContentType.objects.get_for_model(Import), object_id=pk)
                | Q(content_type=ContentType.objects.get_for_model(DeviceDeletionRequest), object_id=deletion_request_id),
                is_read=False,
            ).filter(reviewer_filter).update(is_read=True, responded_by=request.user)
            messages.success(request, f"Delete request approved. Device {serial_number} was deleted.")
            return redirect('display_unapproved_imports')

        if pending_update:
            pending_update_id = pending_update.pk
            trainer_user_id = pending_update.updated_by_id
            try:
                old_assignee = _apply_pending_update_to_import(
                    import_instance=import_instance,
                    pending_update=pending_update,
                    approved_by=request.user,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect('display_unapproved_imports')

            if import_instance.assignee != old_assignee:
                if old_assignee and old_assignee.email:
                    send_device_assignment_email(import_instance, action='cleared', cleared_by=request.user)
                if import_instance.assignee and import_instance.assignee.email:
                    send_device_assignment_email(import_instance, action='assigned')
                if old_assignee and import_instance.assignee:
                    send_device_assignment_email(import_instance, action='transferred', cleared_by=request.user)

            pending_update.delete()
            Notification.objects.filter(
                Q(content_type=ContentType.objects.get_for_model(Import), object_id=import_instance.pk)
                | Q(content_type=ContentType.objects.get_for_model(PendingUpdate), object_id=pending_update_id),
                is_read=False,
            ).filter(reviewer_filter).exclude(user_id=trainer_user_id).update(
                is_read=True,
                responded_by=request.user,
            )
            messages.success(request, f"Device {import_instance.serial_number} update approved.")
            return redirect('display_unapproved_imports')

        import_instance.is_approved = True
        import_instance.approved_by = request.user
        import_instance.save()
        Notification.objects.filter(
            content_type=ContentType.objects.get_for_model(Import),
            object_id=import_instance.pk,
            is_read=False,
        ).filter(reviewer_filter).update(is_read=True, responded_by=request.user)
        messages.success(request, f"Device {import_instance.serial_number} approved.")
        return redirect('display_unapproved_imports')


@login_required
def import_reject(request, pk):
    if not can_review_device_requests(request.user):
        messages.error(request, "You do not have permission to review device requests.")
        return redirect('display_unapproved_imports')

    import_instance = get_object_or_404(Import, pk=pk)
    if request.method != 'POST':
        return redirect('display_unapproved_imports')

    reviewer_filter = (
        Q(user__is_superuser=True)
        | Q(user__is_it_manager=True)
        | Q(user__is_senior_it_officer=True)
        | (Q(user__is_staff=True) & Q(user__is_trainer=False))
    )

    with transaction.atomic():
        deletion_request = DeviceDeletionRequest.objects.filter(device=import_instance).select_related('requested_by').first()
        pending_update = PendingUpdate.objects.filter(import_record=import_instance).order_by('-created_at').first()

        if deletion_request:
            requester = deletion_request.requested_by
            deletion_request_id = deletion_request.pk
            deletion_request.delete()
            if requester:
                Notification.objects.create(
                    user=requester,
                    message=f"Your delete request for device {import_instance.serial_number} was rejected.",
                    content_type=ContentType.objects.get_for_model(Import),
                    object_id=import_instance.pk,
                )
            Notification.objects.filter(
                Q(content_type=ContentType.objects.get_for_model(Import), object_id=import_instance.pk)
                | Q(content_type=ContentType.objects.get_for_model(DeviceDeletionRequest), object_id=deletion_request_id),
                is_read=False,
            ).filter(reviewer_filter).update(is_read=True, responded_by=request.user)
            messages.success(request, f"Delete request for device {import_instance.serial_number} was rejected.")
            return redirect('display_unapproved_imports')

        if pending_update:
            pending_update.pending_clarification = True
            pending_update.save()
            trainer = pending_update.updated_by
            if trainer:
                content_type = ContentType.objects.get_for_model(PendingUpdate)
                if not Notification.objects.filter(user=trainer, content_type=content_type, object_id=pending_update.pk).exists():
                    Notification.objects.create(
                        user=trainer,
                        message=f"Your update for device {pending_update.serial_number} was rejected. Please provide clarification.",
                        content_type=content_type,
                        object_id=pending_update.pk
                    )
            Notification.objects.filter(
                Q(content_type=ContentType.objects.get_for_model(Import), object_id=import_instance.pk)
                | Q(content_type=ContentType.objects.get_for_model(PendingUpdate), object_id=pending_update.pk),
                is_read=False,
            ).filter(reviewer_filter).exclude(user_id=getattr(trainer, "id", None)).update(
                is_read=True,
                responded_by=request.user,
            )
            messages.success(request, f"Update for device {pending_update.serial_number} sent back for clarification.")
            return redirect('display_unapproved_imports')

        import_instance.pending_clarification = True
        import_instance.save()
        trainer = import_instance.added_by
        if trainer:
            content_type = ContentType.objects.get_for_model(Import)
            if not Notification.objects.filter(user=trainer, content_type=content_type, object_id=import_instance.pk).exists():
                Notification.objects.create(
                    user=trainer,
                    message=f"Your import request for device {import_instance.serial_number} was rejected. Please provide clarification.",
                    content_type=content_type,
                    object_id=import_instance.pk
                )
        Notification.objects.filter(
            content_type=ContentType.objects.get_for_model(Import),
            object_id=import_instance.pk,
            is_read=False,
        ).filter(reviewer_filter).exclude(user_id=getattr(trainer, "id", None)).update(
            is_read=True,
            responded_by=request.user,
        )
        messages.success(request, f"Import request for device {import_instance.serial_number} sent back for clarification.")
        return redirect('display_unapproved_imports')


@login_required
def import_approve_all(request):
    if not can_review_device_requests(request.user):
        messages.error(request, "You do not have permission to approve device requests.")
        return redirect('display_unapproved_imports')

    if request.method != 'POST':
        return redirect('display_unapproved_imports')

    page_number = request.GET.get('page', '1')
    items_per_page = request.GET.get('items_per_page', '10')
    search_query = request.GET.get('search', '')

    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100, 500]:
            items_per_page = 10
    except ValueError:
        items_per_page = 10

    try:
        page_number = int(page_number) if page_number else 1
    except ValueError:
        page_number = 1

    data = Import.objects.filter(is_approved=False, is_disposed=False, deletion_request__isnull=True).order_by('-pk')
    if search_query:
        data = data.filter(
            Q(centre__name__icontains=search_query) |
            Q(centre__centre_code__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(device_name__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            _build_assignee_search_query(search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(reason_for_update__icontains=search_query)
        )

    paginator = Paginator(data, items_per_page)
    try:
        data_on_page = paginator.page(page_number)
    except PageNotAnInteger:
        data_on_page = paginator.page(1)
    except EmptyPage:
        data_on_page = paginator.page(paginator.num_pages)

    approved_count = 0
    skipped_conflicts = []
    with transaction.atomic():
        for item in data_on_page:
            pending_update = PendingUpdate.objects.filter(import_record=item).order_by('-created_at').first()
            if pending_update:
                try:
                    old_assignee = _apply_pending_update_to_import(
                        import_instance=item,
                        pending_update=pending_update,
                        approved_by=request.user,
                    )
                except ValueError:
                    skipped_conflicts.append(
                        f"Skipped {item.serial_number}: the updated serial number conflicts with another device."
                    )
                    continue
                if item.assignee != old_assignee:
                    if old_assignee and old_assignee.email:
                        send_device_assignment_email(item, action='cleared', cleared_by=request.user)
                    if item.assignee and item.assignee.email:
                        send_device_assignment_email(item, action='assigned')
                    if old_assignee and item.assignee:
                        send_device_assignment_email(item, action='transferred', cleared_by=request.user)
                pending_update.delete()
                Notification.objects.filter(
                    content_type=ContentType.objects.get_for_model(PendingUpdate),
                    object_id=pending_update.pk,
                    is_read=False,
                ).exclude(user__is_trainer=True).update(is_read=True, responded_by=request.user)
                approved_count += 1
            elif not item.is_approved:
                item.is_approved = True
                item.approved_by = request.user
                item.save()
                Notification.objects.filter(
                    content_type=ContentType.objects.get_for_model(Import),
                    object_id=item.pk,
                    is_read=False,
                ).exclude(user__is_trainer=True).update(is_read=True, responded_by=request.user)
                approved_count += 1

    if approved_count > 0:
        messages.success(request, f"{approved_count} device(s) approved successfully.")
    else:
        messages.info(request, "No unapproved devices to approve on this page.")
    if skipped_conflicts:
        messages.warning(request, skipped_conflicts[0])

    redirect_url = reverse('display_unapproved_imports')
    query_params = [f"page={page_number}", f"items_per_page={items_per_page}"]
    if search_query:
        query_params.append(f"search={search_query}")
    redirect_url += "?" + "&".join(query_params)
    return redirect(redirect_url)


@login_required
def clear_user(request, device_id):
    device = get_object_or_404(Import, id=device_id)
    if not can_clear_device_users(request.user):
        messages.error(request, "Only IT staff, IT managers, and senior IT officers can clear assigned devices.")
        return redirect('device_detail', pk=device.pk)
    if request.user.is_trainer and device.centre != request.user.centre:
        messages.error(request, "You can only clear devices from your centre.")
        return redirect('display_approved_imports')
    blocked = _block_actions_if_inactive(request, device)
    if blocked:
        return blocked
    if request.method == 'POST':
        form = ClearanceForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                previous_assignee = device.assignee
                # Save history
                if previous_assignee:
                    DeviceUserHistory.objects.create(
                        device=device,
                        assignee_first_name=previous_assignee.first_name,
                        assignee_last_name=previous_assignee.last_name,
                        assignee_email_address=previous_assignee.email,
                        assigned_by=device.added_by or request.user,
                        cleared_date=timezone.now()
                    )
                # Create clearance record
                clearance = form.save(commit=False)
                clearance.device = device
                clearance.cleared_by = request.user
                clearance.remarks = form.cleaned_data.get('remarks', 'Device cleared')
                clearance.save()
                # Clear assignee
                device.assignee = None
                device.assignee_cache = ''
                device.status = 'Available'
                device.department_id = 1 # default / unassigned dept
                device.reason_for_update = f"Cleared by {request.user.username} on {timezone.now().date()}"
                device.uaf_signed = False  # Reset UAF signed flag
                device.save()
                # Archive current agreement (if exists)
                agreement = device.agreements.filter(is_archived=False).first()
                if agreement:
                    agreement.is_archived = True
                    agreement.save(update_fields=['is_archived'])
                # Send clearance email with PDF attachment to previous user
                if previous_assignee and previous_assignee.email:
                    send_clearance_email_with_pdf(device, previous_assignee, request.user)
                messages.success(request, f"Device {device.serial_number} cleared successfully.")
                return redirect('display_approved_imports')
    else:
        form = ClearanceForm()
    return render(request, 'import/clear_user.html', {
        'form': form,
        'device': device
    })

def send_clearance_email_with_pdf(device, assignee, cleared_by):
    """Sends clearance email with PDF form attached"""
    subject = f"Device Clearance Confirmation - {device.serial_number}"
    message = f"""
Dear {assignee.full_name},

This is to confirm that device {device.serial_number} ({device.device_name or device.system_model or 'Unknown'}) has been cleared from your use.

Details:
- Cleared by: {cleared_by.get_full_name() or cleared_by.username}
- Date: {timezone.now().date()}
- Centre: {device.centre.name if device.centre else 'N/A'}
- Department: {device.department.name if device.department else 'N/A'}

Please find the official clearance form attached.

Best regards,
MoH IAfrica IT Team
"""

    # Generate PDF
    pdf_buffer = BytesIO()
    pdf_buffer.write(generate_pdf_buffer(device).read())  
    pdf_buffer.seek(0)

    attachment = ('clearance_form.pdf', pdf_buffer.read(), 'application/pdf')

    send_custom_email(subject, message, [assignee.email], attachment)


@login_required
def download_clearance_form(request, device_id):
    device = get_object_or_404(Import, id=device_id)
    clearance = device.clearances.order_by('-created_at').first()
    if not clearance:
        messages.error(request, "No clearance record found for this device.")
        return redirect('display_approved_imports')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="clearance_form_{device.serial_number}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    normal_style = styles['Normal']
    footer_style = ParagraphStyle(
        name='FooterStyle',
        parent=normal_style,
        fontSize=10,
        alignment=1
    )
    remarks_style = ParagraphStyle(
        name='RemarksStyle',
        parent=normal_style,
        fontSize=10,
        wordWrap='CJK',
        leading=12,
        alignment=0
    )
    elements.append(Paragraph(f'Clearance Form for Device {device.serial_number} - MOHI IT Inventory', title_style))
    elements.append(Spacer(1, 12))
    data = [
        ['Field', 'Value'],
        ['Device Serial Number', device.serial_number or 'N/A'],
        ['Device Name', device.device_name or 'N/A'],
        ['Centre', device.centre.name if device.centre else 'N/A'],
        ['Department', device.department.name if device.department else 'N/A'],
        ['Status', device.status or 'N/A'],
        ['Date', device.date.strftime("%Y-%m-%d") if device.date else 'N/A'],
        ['Cleared By', clearance.cleared_by.username],
        ['Clearance Date', clearance.created_at.strftime("%Y-%m-%d")],
        ['Approved By', device.approved_by.username if device.approved_by else 'N/A'],
    ]
    remarks = device.reason_for_update or clearance.remarks or 'N/A'
    data.append(['Remarks', Paragraph(remarks, remarks_style)])
    table = Table(data, colWidths=[100*mm, 100*mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))
    user_history = device.user_history.all().order_by('assigned_date')
    if user_history.exists():
        history_data = [['Assignee Name', 'Email', 'Assigned By', 'Assigned Date', 'Cleared Date']]
        for history in user_history:
            assignee_name = f"{history.assignee_first_name or ''} {history.assignee_last_name or ''}".strip() or 'N/A'
            history_data.append([
                assignee_name,
                history.assignee_email_address or 'N/A',
                history.assigned_by.username if history.assigned_by else 'N/A',
                history.assigned_date.strftime("%Y-%m-%d") if history.assigned_date else 'N/A',
                history.cleared_date.strftime("%Y-%m-%d") if history.cleared_date else 'N/A',
            ])
        history_table = Table(history_data, colWidths=[40*mm, 40*mm, 40*mm, 40*mm, 40*mm])
        history_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(Paragraph('User History', normal_style))
        elements.append(Spacer(1, 6))
        elements.append(history_table)
        elements.append(Spacer(1, 12))
    elements.append(Paragraph('Signature Section', footer_style))
    elements.append(Spacer(1, 6))
    signature_data = [
        ['Cleared By Signature:', ''],
        ['Date:', ''],
        ['Approved By Name & Signature:', ''],
        ['Date:', ''],
    ]
    signature_table = Table(signature_data, colWidths=[80*mm, 120*mm])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(signature_table)
    import random
    def add_watermark(canvas, doc):
        watermark_text = "MOHI IT"
        canvas.saveState()
        canvas.setFont("Helvetica", 20)
        canvas.setFillGray(0.95, 0.95)
        page_width, page_height = doc.pagesize
        grid_size = 80
        placed_positions = []
        for x in range(0, int(page_width), grid_size):
            for y in range(0, int(page_height), grid_size):
                offset_x = random.randint(-40, 40)
                offset_y = random.randint(-40, 40)
                adjusted_x = x + offset_x
                adjusted_y = y + offset_y
                if (10 <= adjusted_x <= page_width - 10 and
                    10 <= adjusted_y <= page_height - 10 and
                    not any(abs(adjusted_x - px) < 50 or abs(adjusted_y - py) < 50 for px, py in placed_positions)):
                    canvas.rotate(45)
                    canvas.drawString(adjusted_x, adjusted_y, watermark_text)
                    canvas.rotate(-45)
                    placed_positions.append((adjusted_x, adjusted_y))
        canvas.restoreState()
    doc.build(elements, onFirstPage=add_watermark, onLaterPages=add_watermark)
    return response

