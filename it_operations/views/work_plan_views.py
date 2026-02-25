import io
import csv
import os
from datetime import datetime, date, timedelta
import calendar

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.conf import settings
from django.core.mail import send_mail
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.db.models import Q
from django.contrib.auth import get_user_model

# ReportLab Imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Internal Imports
from ..models import PublicHoliday, WorkPlan, WorkPlanTask
from devices.models import Centre, Department, CustomUser
from ..utils import get_kenyan_holidays, notify_collaborator

# Excel exports
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
import logging
from itertools import chain
from django.db import transaction
from datetime import datetime, time, timedelta
from django.utils import timezone

logger = logging.getLogger(__name__)

User = get_user_model()

# ============ PERMISSION HELPERS ============

def is_manager_of(user, target_user):
    """
    Determines if 'user' is a manager of 'target_user' based on hierarchy.
    """
    if user == target_user:
        return True
    # IT Manager manages IT Staff (non-trainers)
    if getattr(user, 'is_it_manager', False) and target_user.is_staff and not getattr(target_user, 'is_trainer', False):
        return True
    # Senior IT manages Trainers
    if getattr(user, 'is_senior_it_officer', False) and getattr(target_user, 'is_trainer', False):
        return True
    return False

def _get_next_working_day(from_date):
    """
    Returns the next working day (Mon-Sat), skipping Sundays and public holidays.
    """
    current = from_date + timedelta(days=1)
    while True:
        if current.weekday() == 6:  # Sunday
            current += timedelta(days=1)
            continue
        if PublicHoliday.objects.filter(date=current).exists():
            current += timedelta(days=1)
            continue
        return current


def _week_add_deadline(week_start, now=None):
    """Monday 10:00 AM deadline for a given week start, timezone-consistent."""
    now = now or timezone.now()
    deadline_dt = datetime.combine(week_start, time(10, 0))
    if timezone.is_aware(now) and timezone.is_naive(deadline_dt):
        return timezone.make_aware(deadline_dt, timezone.get_current_timezone())
    if timezone.is_naive(now) and timezone.is_aware(deadline_dt):
        return timezone.make_naive(deadline_dt, timezone.get_current_timezone())
    return deadline_dt


def _send_workplan_reopened_email(work_plan, reopened_by):
    """Notify the plan owner that current-week task creation was reopened."""
    recipient = getattr(work_plan.user, "email", None)
    if not recipient:
        return False, "User has no email address."

    owner_name = work_plan.user.get_full_name() or work_plan.user.username
    manager_name = reopened_by.get_full_name() or reopened_by.username
    week_start = work_plan.week_start_date
    week_end = work_plan.week_end_date

    subject = f"Work Plan Reopened for {week_start.strftime('%d %b %Y')}"
    body = (
        f"Dear {owner_name},\n\n"
        "Your weekly work plan has been reopened by the IT Manager so you can add tasks.\n\n"
        f"Week: {week_start.strftime('%A, %d %b %Y')} - {week_end.strftime('%A, %d %b %Y')}\n"
        f"Reopened by: {manager_name}\n\n"
        "Please log in and complete your work plan as soon as possible.\n\n"
        "Regards,\n"
        "IT Operations System\n"
    )

    try:
        send_mail(
            subject=subject,
            message=body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
            recipient_list=[recipient],
            fail_silently=False,
        )
        return True, None
    except Exception as exc:
        logger.exception("Failed to send work plan reopened email for work_plan=%s", work_plan.id)
        return False, str(exc)


# ============ VIEWS ============

@login_required
def work_plan_list(request):
    user = request.user
    today = timezone.now().date()
    
    target_user = user 
    is_manager = getattr(user, 'is_it_manager', False)
    is_senior = getattr(user, 'is_senior_it_officer', False)
    users_to_filter = None

    if is_manager:
        users_to_filter = User.objects.filter(is_staff=True, is_active=True).exclude(is_trainer=True).order_by('first_name')
    elif is_senior:
        users_to_filter = User.objects.filter(is_trainer=True, is_active=True).order_by('first_name')

    filter_id = request.GET.get('user_filter')
    if filter_id and (is_manager or is_senior):
        try:
            target_user = User.objects.get(pk=filter_id)
        except User.DoesNotExist:
            pass
            
    current_week_start = today - timedelta(days=today.weekday())
    current_week_end = current_week_start + timedelta(days=6)
    
    current_tasks = WorkPlanTask.objects.filter(
        work_plan__user=target_user,
        date__range=[current_week_start, current_week_end]
    ).order_by('date')

    collab_tasks = WorkPlanTask.objects.filter(
        collaborators=target_user,
        date__gte=today
    ).order_by('date')

    plans = WorkPlan.objects.filter(user=target_user).order_by('-week_start_date')

    context = {
        'target_user': target_user,
        'users_to_filter': users_to_filter,
        'current_tasks': current_tasks,
        'collab_tasks': collab_tasks,
        'plans': plans,
        'is_viewing_others': user != target_user
    }
    return render(request, 'work_plan/workplan_list.html', context)


@login_required
def get_task_details_json(request, pk):
    """
    API endpoint for Edit Modal - returns task data in correct format.
    """
    task = get_object_or_404(WorkPlanTask, pk=pk)
    
    user = request.user
    is_owner = user == task.work_plan.user
    is_collab = user in task.collaborators.all()
    
    if not (is_owner or is_collab):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    data = {
        'id': task.id,
        'task_name': task.task_name,
        'date': task.date.strftime('%Y-%m-%d'),
        'centre': task.centre.id if task.centre else '',
        'department': task.department.id if task.department else '',
        'target': task.target or '',
        'resources_needed': task.resources_needed or '',
        'other_parties': task.other_parties or '',
        'is_leave': task.is_leave,
        'status': task.status,
        'collaborators': list(task.collaborators.values_list('id', flat=True)),
    }
    return JsonResponse(data)




@login_required
def work_plan_detail(request, pk):
    """
    Consolidated Detailed view of the user's own Work Plan.
    - Shows owned tasks from this plan.
    - Appends collaborative tasks from other plans that fall on the same week dates.
    - Visual distinction and limited permissions for collaborative tasks.
    - Managers can view any user's plan (if pk belongs to someone they manage).
    """
    work_plan = get_object_or_404(WorkPlan, pk=pk)
    user = request.user
    today = timezone.now().date()

    # Access Check
    is_owner = (user == work_plan.user)
    is_manager = is_manager_of(user, work_plan.user)
    is_it_manager_user = bool(getattr(user, 'is_it_manager', False))
    is_collab = work_plan.tasks.filter(collaborators=user).exists()

    if not (is_owner or is_manager or is_collab):
        messages.error(request, "Access Denied. You do not have permission to view this work plan.")
        return redirect('work_plan_list')

    # Global "Can Add Task" Logic (Only the plan owner)
    can_add_global = bool(getattr(work_plan, "can_add_tasks", False) and is_owner)

    # Helper: week bounds
    week_start = work_plan.week_start_date
    week_end = work_plan.week_end_date
    manager_can_toggle_creation = bool(
        is_it_manager_user
        and is_manager
        and work_plan.is_current_week
        and work_plan.deadline_passed_for_adding
    )

    # Handle Add Task (POST) - only owner can add to their plan
    if request.method == 'POST' and request.POST.get('add_task'):
        if not can_add_global:
            messages.error(request, "Adding tasks is locked for this week.")
            return redirect('work_plan_detail', pk=pk)

        # --- Read & validate inputs (match template field names) ---
        date_str = (request.POST.get('date') or '').strip()
        task_name = (request.POST.get('task_name') or '').strip()
        resources_needed = (request.POST.get('resources_needed') or '').strip()
        target = (request.POST.get('target') or '').strip()
        other_parties = (request.POST.get('other_parties') or '').strip()

        is_leave = request.POST.get('is_leave') in ('on', 'true', '1', 'yes')

        centre_id = (request.POST.get('centre') or '').strip()
        department_id = (request.POST.get('department') or '').strip()
        collaborator_ids = request.POST.getlist('collaborators')  # multi-select

        if not date_str:
            messages.error(request, "Date is required.")
            return redirect('work_plan_detail', pk=pk)

        try:
            task_date = timezone.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Invalid date.")
            return redirect('work_plan_detail', pk=pk)

        # Enforce only dates within this work plan week
        if task_date < week_start or task_date > week_end:
            messages.error(request, "Selected date must be within this work plan week.")
            return redirect('work_plan_detail', pk=pk)

        # If not leave, task_name must exist
        if not is_leave and not task_name:
            messages.error(request, "Task Name is required.")
            return redirect('work_plan_detail', pk=pk)

        # Optional FKs
        centre = Centre.objects.filter(id=centre_id).first() if centre_id else None
        department = Department.objects.filter(id=department_id).first() if department_id else None

        # Collaborators must be valid users (and never include owner)
        collab_qs = User.objects.filter(is_active=True, id__in=collaborator_ids).exclude(id=work_plan.user.id)

        # Prevent adding tasks on a day already blocked by leave (optional rule)
        if WorkPlanTask.objects.filter(work_plan=work_plan, date=task_date, is_leave=True).exists() and not is_leave:
            messages.error(request, "That day is marked as On Leave. Pick another date.")
            return redirect('work_plan_detail', pk=pk)

        try:
            with transaction.atomic():
                new_task = WorkPlanTask.objects.create(
                    work_plan=work_plan,
                    created_by=user,  # ✅ FIX: required field
                    date=task_date,
                    task_name=("On Leave" if is_leave else task_name),
                    is_leave=is_leave,
                    centre=centre,
                    department=department,
                    resources_needed=resources_needed,
                    target=target,
                    other_parties=other_parties,
                )

                if collab_qs.exists():
                    new_task.collaborators.set(collab_qs)

            messages.success(request, "Task created successfully ✅")
        except Exception as e:
            logger.exception("Failed to create task for work_plan=%s", work_plan.id)
            messages.error(request, f"Failed to create task: {str(e)}")

        return redirect('work_plan_detail', pk=pk)

    # Fetch owned tasks (from this plan)
    owned_tasks = work_plan.tasks.all().select_related('centre', 'department').prefetch_related('collaborators')

    # Fetch collaborative tasks from OTHER plans on the same week dates
    collaborative_tasks = WorkPlanTask.objects.filter(
        collaborators=user,
        date__gte=week_start,
        date__lte=week_end
    ).exclude(
        work_plan=work_plan
    ).select_related('work_plan__user', 'centre', 'department').prefetch_related('collaborators')

    # Combine and sort by date
    all_tasks = list(chain(owned_tasks, collaborative_tasks))
    all_tasks.sort(key=lambda t: t.date)

    # Process tasks with flags and permissions
    processed_tasks = []
    for t in all_tasks:
        is_task_owner = (t.work_plan.user == user)
        is_task_collab = user in t.collaborators.all()

        t.is_owned_task = is_task_owner
        t.is_collaborative_task = is_task_collab and not is_task_owner

        t.can_edit = is_task_owner or (t.is_collaborative_task and t.date >= today)
        t.can_delete = is_task_owner or is_manager
        t.can_reschedule = is_task_owner or is_manager or (t.is_collaborative_task and t.date >= today)
        t.can_change_status = (is_task_owner or is_manager or t.is_collaborative_task) and (t.date <= today)
        t.can_comment = True

        processed_tasks.append(t)

    # Leave dates - only from owned plan (for blocking add form)
    leave_dates = [
        task.date.strftime('%Y-%m-%d')
        for task in owned_tasks.filter(is_leave=True)
    ]

    context = {
        'work_plan': work_plan,
        'tasks': processed_tasks,
        'can_add_tasks': can_add_global,
        'today_date': today,
        'centres': Centre.objects.all(),
        'departments': Department.objects.all(),
        'potential_collaborators': User.objects.filter(is_active=True).exclude(id=work_plan.user.id).order_by('first_name'),
        'week_days': [work_plan.week_start_date + timedelta(days=i) for i in range(7)],
        'all_users': User.objects.filter(is_active=True).order_by('first_name'),
        'leave_dates': leave_dates,
        'is_owner': is_owner,
        'is_manager': is_manager,
        'manager_can_toggle_creation': manager_can_toggle_creation,
        'manager_override_creation_open': work_plan.manager_task_creation_override_open,
        'is_current_week_plan': work_plan.is_current_week,
        'creation_deadline_passed': work_plan.deadline_passed_for_adding,
        'has_collaborative_tasks': collaborative_tasks.exists(),
        'is_collaborator': is_collab,
    }
    return render(request, 'work_plan/workplan_detail.html', context)


@login_required
@require_POST
def work_plan_toggle_creation_override(request, pk):
    work_plan = get_object_or_404(WorkPlan, pk=pk)

    if not getattr(request.user, 'is_it_manager', False):
        messages.error(request, "Permission denied.")
        return redirect('work_plan_detail', pk=pk)

    if not is_manager_of(request.user, work_plan.user):
        messages.error(request, "Permission denied.")
        return redirect('work_plan_detail', pk=pk)

    if not work_plan.is_current_week:
        messages.error(request, "This action only applies to the current week.")
        return redirect('work_plan_detail', pk=pk)

    if not work_plan.deadline_passed_for_adding:
        messages.error(request, "Manager override is only available after Monday 10:00 AM.")
        return redirect('work_plan_detail', pk=pk)

    action = (request.POST.get('action') or '').strip().lower()
    if action not in {'open', 'close'}:
        action = 'close' if work_plan.manager_task_creation_override_open else 'open'

    work_plan.manager_task_creation_override_open = (action == 'open')
    work_plan.save(update_fields=['manager_task_creation_override_open', 'updated_at'])

    if work_plan.manager_task_creation_override_open:
        email_sent, email_error = _send_workplan_reopened_email(work_plan, request.user)
        messages.success(request, f"Task creation reopened for {work_plan.user.get_full_name() or work_plan.user.username} (current week).")
        if email_sent:
            messages.info(request, f"Email notification sent to {work_plan.user.email}.")
        elif email_error:
            messages.warning(request, f"Reopened, but email notification failed: {email_error}")
    else:
        messages.success(request, f"Task creation closed again for {work_plan.user.get_full_name() or work_plan.user.username} (current week).")

    return redirect('work_plan_detail', pk=pk)


@login_required
@require_POST
def work_plan_task_status_update(request, pk):
    task = get_object_or_404(WorkPlanTask, pk=pk)
    
    # 1. Date Security Check (Server-side enforcement)
    today = timezone.now().date()
    if task.date > today:
        messages.error(request, "Cannot mark future tasks as completed.")
        return redirect(request.META.get('HTTP_REFERER'))
        
    # 2. Permission Check
    user = request.user
    is_owner = (user == task.work_plan.user)
    is_manager = is_manager_of(user, task.work_plan.user)
    is_collab = user in task.collaborators.all()
    
    if is_owner or is_manager or is_collab:
        if task.status != 'Completed':
            task.status = 'Completed'
        else:
            task.status = 'Pending'
        task.save()
        messages.success(request, f"Task status updated.")
    else:
        messages.error(request, "Permission denied.")
        
    return redirect(request.META.get('HTTP_REFERER'))


@login_required
@require_POST
def work_plan_task_edit(request, pk):
    task = get_object_or_404(WorkPlanTask, pk=pk)
    user = request.user
    
    # Permission check
    is_owner = (user == task.work_plan.user)
    is_collab = user in task.collaborators.all()
    
    if not (is_owner or is_collab):
        messages.error(request, "Permission denied.")
        return redirect(request.META.get('HTTP_REFERER', 'work_plan_list'))

    try:
        new_is_leave = request.POST.get('is_leave') == 'on'
        current_is_leave = task.is_leave

        # CASE: Marking the day as "On Leave"
        if new_is_leave:
            # Get all tasks on this date
            all_tasks_on_date = WorkPlanTask.objects.filter(
                work_plan=task.work_plan,
                date=task.date
            )
            
            # Identify regular (non-leave) tasks to move
            regular_tasks = all_tasks_on_date.filter(is_leave=False)
            
            moved_count = 0
            if regular_tasks.exists():
                next_day = _get_next_working_day(task.date)
                new_week_start = next_day - timedelta(days=next_day.weekday())
                new_plan, _ = WorkPlan.objects.get_or_create(
                    user=task.work_plan.user,
                    week_start_date=new_week_start
                )
                
                # Copy each regular task to next working day
                for old_task in regular_tasks:
                    new_task = WorkPlanTask.objects.create(
                        work_plan=new_plan,
                        date=next_day,
                        task_name=old_task.task_name,
                        centre=old_task.centre,
                        department=old_task.department,
                        resources_needed=old_task.resources_needed,
                        target=old_task.target,
                        other_parties=old_task.other_parties,
                        comments=f"Auto-rescheduled from {old_task.date} due to leave" +
                                 (f"\n{old_task.comments}" if old_task.comments else ""),
                        is_leave=False,
                        created_by=user,
                        status='Pending'
                    )
                    new_task.collaborators.set(old_task.collaborators.all())
                    moved_count += 1
                
                messages.info(request, f"{moved_count} task(s) successfully moved to {next_day}.")
                
                # CRITICAL: DELETE the original regular tasks
                regular_tasks.delete()
            
            # Also delete any existing "On Leave" task to avoid duplicates
            all_tasks_on_date.filter(is_leave=True).delete()
            
            # Create single clean "On Leave" placeholder
            WorkPlanTask.objects.create(
                work_plan=task.work_plan,
                date=task.date,
                task_name="On Leave",
                is_leave=True,
                created_by=user,
                status='Pending',
                resources_needed="N/A",
                target="N/A"
            )
            
            messages.success(request, f"{task.date} is now marked as On Leave. All previous tasks have been moved.")
            return redirect(request.META.get('HTTP_REFERER', 'work_plan_list'))

        # CASE: Removing "On Leave" (converting back to normal task)
        elif current_is_leave and not new_is_leave:
            task.is_leave = False
            task.task_name = request.POST.get('task_name')
            task.target = request.POST.get('target')
            task.resources_needed = request.POST.get('resources_needed')
            task.other_parties = request.POST.get('other_parties')
            task.centre_id = request.POST.get('centre') or None
            task.department_id = request.POST.get('department') or None
            
            collab_ids = request.POST.getlist('collaborators')
            task.collaborators.set(collab_ids)
            
            task.save()
            messages.success(request, "Task updated — no longer on leave.")
            return redirect(request.META.get('HTTP_REFERER', 'work_plan_list'))

        # CASE: Normal edit (no change to leave status)
        else:
            task.task_name = request.POST.get('task_name')
            task.target = request.POST.get('target')
            task.resources_needed = request.POST.get('resources_needed')
            task.other_parties = request.POST.get('other_parties')
            task.centre_id = request.POST.get('centre') or None
            task.department_id = request.POST.get('department') or None
            
            collab_ids = request.POST.getlist('collaborators')
            task.collaborators.set(collab_ids)
            
            task.save()
            messages.success(request, "Task updated successfully.")
    
    except Exception as e:
        messages.error(request, f"Error updating task: {str(e)}")
    
    return redirect(request.META.get('HTTP_REFERER', 'work_plan_list'))

@login_required
@require_POST
def work_plan_task_reschedule(request, pk):
    task = get_object_or_404(WorkPlanTask, pk=pk)
    user = request.user
    
    # Permission: Owner OR Manager OR Collaborator
    is_owner = (user == task.work_plan.user)
    is_manager = is_manager_of(user, task.work_plan.user)
    is_collab = user in task.collaborators.all()
    
    if not (is_owner or is_manager or is_collab):
        messages.error(request, "Permission denied.")
        return redirect(request.META.get('HTTP_REFERER'))

    new_date_str = request.POST.get('reschedule_date')
    reason = request.POST.get('reschedule_reason', '').strip()
    
    if not reason:
        messages.error(request, "Reschedule reason is required.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    try:
        new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        
        # 1. Find or Create Plan for the new week
        new_week_start = new_date - timedelta(days=new_date.weekday())
        new_plan, _ = WorkPlan.objects.get_or_create(user=task.work_plan.user, week_start_date=new_week_start)
        
        # 2. Create New Task with reason
        new_task = WorkPlanTask.objects.create(
            work_plan=new_plan,
            date=new_date,
            task_name=task.task_name,
            centre=task.centre,
            department=task.department,
            other_parties=task.other_parties,
            resources_needed=task.resources_needed,
            target=task.target,
            comments=f"Rescheduled from {task.date}: {reason}" + 
                     (f"\n{task.comments}" if task.comments else ""),
            reschedule_reason=reason,  # Save the reason
            created_by=user,
            status='Pending',
            is_leave=task.is_leave
        )
        new_task.collaborators.set(task.collaborators.all())
        
        # 3. Mark old task as Rescheduled
        task.status = 'Rescheduled'
        task.reschedule_reason = reason  # Also save on old task for audit
        task.save()
        
        messages.success(request, f"Task rescheduled to {new_date}. Reason: {reason}")
    except Exception as e:
        messages.error(request, f"Error: {e}")

    return redirect(request.META.get('HTTP_REFERER'))


@login_required
@require_POST
def work_plan_task_comment_add(request, pk):
    """
    UPDATED: Now sends notifications to owner and all collaborators
    """
    task = get_object_or_404(WorkPlanTask, pk=pk)
    
    # Permission check
    if not (request.user == task.work_plan.user or 
            is_manager_of(request.user, task.work_plan.user) or 
            request.user in task.collaborators.all()):
        messages.error(request, "Permission denied.")
        return redirect('work_plan_list')

    new_comment = request.POST.get('new_comment')
    if new_comment:
        formatted = f"\n[{request.user.first_name}]: {new_comment}"
        task.comments = (task.comments or "") + formatted
        task.save()
        
        # CORRECTED IMPORT: utils is in it_operations app
        from it_operations.utils import notify_comment_added
        notify_comment_added(task, new_comment, request.user)
        
        messages.success(request, "Comment added and notifications sent.")
    
    return redirect('work_plan_detail', pk=task.work_plan.pk)


@login_required
@require_POST
def work_plan_task_delete(request, pk):
    task = get_object_or_404(WorkPlanTask, pk=pk)
    user = request.user
    
    # Permission: Owner OR Manager only. Collaborators CANNOT delete.
    is_owner = (user == task.work_plan.user)
    is_manager = is_manager_of(user, task.work_plan.user)
    
    if is_owner or is_manager:
        task.delete()
        messages.success(request, "Task deleted.")
    else:
        messages.error(request, "Permission denied.")
    return redirect('work_plan_detail', pk=task.work_plan.pk)


@login_required
def work_plan_create(request):
    if request.method == 'POST':
        date_str = request.POST.get('week_start_date')
        week_start = datetime.strptime(date_str, '%Y-%m-%d').date()
        if week_start.weekday() != 0:
            messages.error(request, "Must start on Monday.")
            return redirect('work_plan_create')
            
        plan, created = WorkPlan.objects.get_or_create(user=request.user, week_start_date=week_start)
        return redirect('work_plan_detail', pk=plan.pk)
    return render(request, 'work_plan/workplan_create.html')

@login_required
def work_plan_calendar(request):
    today = timezone.now().date()
    now = timezone.now()
    
    # 1. Date Navigation
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except ValueError:
        year = today.year
        month = today.month
    
    # 2. Target User (Manager Logic)
    target_user = request.user
    filter_id = request.GET.get('user_filter')
    if filter_id and (getattr(request.user, 'is_it_manager', False) or 
                     getattr(request.user, 'is_senior_it_officer', False) or 
                     request.user.is_superuser):
        try:
            target_user = User.objects.get(pk=filter_id)
        except User.DoesNotExist:
            pass

    holidays = get_kenyan_holidays(year)
    
    # 3. Fetch Tasks for the Month
    tasks = WorkPlanTask.objects.filter(
        Q(work_plan__user=target_user) | Q(collaborators=target_user),
        date__year=year, 
        date__month=month
    ).distinct().select_related('work_plan')

    events = []
    for t in tasks:
        color_class = 'border-blue-500 bg-blue-100 text-blue-800'
        status_code = 'active'
        
        if t.is_leave: 
            color_class = 'border-yellow-500 bg-yellow-100 text-yellow-800'
            status_code = 'leave'
        elif t.status == 'Completed': 
            color_class = 'border-green-500 bg-green-100 text-green-800'
            status_code = 'completed'
        elif t.status == 'Not Done': 
            color_class = 'border-red-500 bg-red-100 text-red-800'
            status_code = 'not_done'
        elif target_user in t.collaborators.all(): 
            color_class = 'border-purple-500 bg-purple-100 text-purple-800'
            status_code = 'collab'
            
        events.append({
            'id': t.id, 
            'title': t.task_name, 
            'date': t.date, 
            'color': color_class,
            'status_code': status_code,
            'work_plan_id': t.work_plan.id
        })

    # 4. Build Calendar Grid
    cal = calendar.Calendar(firstweekday=0)
    month_days = cal.monthdatescalendar(year, month)
    
    calendar_grid = []
    for week in month_days:
        week_data = []
        week_start = week[0]
        


        deadline = _week_add_deadline(week_start, now=now)
        existing_plan = WorkPlan.objects.filter(user=target_user, week_start_date=week_start).first()
        override_open = bool(
            existing_plan
            and existing_plan.manager_task_creation_override_open
            and existing_plan.is_current_week
        )
        can_add_to_week = bool((target_user == request.user) and ((now <= deadline) or override_open))
        for day in week:
            day_events = [e for e in events if e['date'] == day]
            
            # Flags
            has_leave_task = any(e['status_code'] == 'leave' for e in day_events)
            all_pending = (len(day_events) > 0 and 
                          all(e['status_code'] == 'active' for e in day_events))  # 'active' = Pending
            
            week_data.append({
                'date': day, 
                'day_num': day.day, 
                'is_current_month': day.month == month,
                'is_today': day == today,
                'is_holiday': day in holidays, 
                'events': day_events,
                'can_add_task': can_add_to_week and (day >= today),
                'has_leave_task': has_leave_task,
                'all_pending': all_pending,
                'css_class': "bg-gray-50 text-gray-400" if (day.weekday() == 6 or day in holidays) else "bg-white"
            })
        calendar_grid.append(week_data)

    context = {
        'calendar_grid': calendar_grid,
        'month_name': calendar.month_name[month],
        'year': year, 'month': month,
        'prev_year': (date(year, month, 1) - timedelta(days=1)).year,
        'prev_month': (date(year, month, 1) - timedelta(days=1)).month,
        'next_year': (date(year, month, 28) + timedelta(days=10)).year,
        'next_month': (date(year, month, 28) + timedelta(days=10)).month,
        'target_user': target_user,
        'centres': Centre.objects.all(), 
        'departments': Department.objects.all(),
        'potential_collaborators': User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('first_name'),
    }
    return render(request, 'work_plan/workplan_calendar.html', context)




@login_required
@require_POST
def work_plan_create_task_from_calendar(request):
    try:
        date_str = request.POST.get('date')
        task_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        week_start = task_date - timedelta(days=task_date.weekday())
        
        # Security: Check deadline (allow manager override on existing current-week plan)
        now = timezone.now()
        deadline = _week_add_deadline(week_start, now=now)
        existing_plan = WorkPlan.objects.filter(user=request.user, week_start_date=week_start).first()
        override_open = bool(
            existing_plan
            and existing_plan.manager_task_creation_override_open
            and existing_plan.is_current_week
        )
        if now > deadline and not override_open:
            messages.error(request, "Deadline passed for this week.")
            return redirect('work_plan_calendar')

        plan, _ = WorkPlan.objects.get_or_create(user=request.user, week_start_date=week_start)
        is_leave = request.POST.get('is_leave') == 'on'
        
        # CASE: Adding "On Leave"
        if is_leave:
            # Get all tasks on this date
            all_tasks_on_date = WorkPlanTask.objects.filter(
                work_plan=plan,
                date=task_date
            )
            
            regular_tasks = all_tasks_on_date.filter(is_leave=False)
            moved_count = 0
            
            if regular_tasks.exists():
                next_day = _get_next_working_day(task_date)
                new_week_start = next_day - timedelta(days=next_day.weekday())
                new_plan, _ = WorkPlan.objects.get_or_create(
                    user=request.user,
                    week_start_date=new_week_start
                )
                
                for old_task in regular_tasks:
                    new_task = WorkPlanTask.objects.create(
                        work_plan=new_plan,
                        date=next_day,
                        task_name=old_task.task_name,
                        centre=old_task.centre,
                        department=old_task.department,
                        resources_needed=old_task.resources_needed,
                        target=old_task.target,
                        other_parties=old_task.other_parties,
                        comments=f"Auto-rescheduled from {old_task.date} due to leave" +
                                 (f"\n{old_task.comments}" if old_task.comments else ""),
                        is_leave=False,
                        created_by=request.user,
                        status='Pending'
                    )
                    new_task.collaborators.set(old_task.collaborators.all())
                    moved_count += 1
                
                messages.info(request, f"{moved_count} task(s) moved to {next_day} due to leave.")
                regular_tasks.delete()
            
            # Remove old leave task
            all_tasks_on_date.filter(is_leave=True).delete()
            
            # Create new "On Leave"
            WorkPlanTask.objects.create(
                work_plan=plan,
                date=task_date,
                task_name="On Leave",
                is_leave=True,
                created_by=request.user,
                status='Pending',
                resources_needed="N/A",
                target="N/A"
            )
            messages.success(request, f"{task_date} marked as On Leave.")
        
        # CASE: Normal task
        else:
            if WorkPlanTask.objects.filter(work_plan=plan, date=task_date, is_leave=True).exists():
                messages.error(request, f"Cannot add task on {task_date}. This day is marked as 'On Leave'.")
                return redirect('work_plan_calendar')
            
            task = WorkPlanTask.objects.create(
                work_plan=plan,
                date=task_date,
                task_name=request.POST.get('task_name'),
                centre_id=request.POST.get('centre') or None,
                department_id=request.POST.get('department') or None,
                other_parties=request.POST.get('other_parties'),
                resources_needed=request.POST.get('resources_needed'),
                target=request.POST.get('target'),
                created_by=request.user,
                status='Pending'
            )
            
            collab_ids = request.POST.getlist('collaborators')
            if collab_ids:
                task.collaborators.set(collab_ids)
                for c_id in collab_ids:
                    try:
                        notify_collaborator(task, User.objects.get(pk=c_id))
                    except:
                        pass
            
            messages.success(request, "Task added successfully.")
    
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect('work_plan_calendar')


# ============================================
# 1. download_excel_report function
# ============================================

@login_required
def download_excel_report(request):
    """
    NOW INCLUDES: Tasks where user is owner OR collaborator
    """
    filter_id = request.GET.get('user_filter')
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    report_type = request.GET.get('report_type', 'weekly')

    target_user = request.user
    if filter_id and (getattr(request.user, 'is_it_manager', False) or 
                     getattr(request.user, 'is_senior_it_officer', False) or 
                     request.user.is_superuser):
        try:
            target_user = User.objects.get(pk=filter_id)
        except:
            pass
    
    # UPDATED: Include tasks where user is owner OR collaborator
    base_qs = WorkPlanTask.objects.filter(
        Q(work_plan__user=target_user) | Q(collaborators=target_user)
    ).distinct()

    if report_type == 'monthly': 
        tasks = base_qs.filter(date__year=year, date__month=month).order_by('date')
        filename = f"WorkPlan_{target_user.username}_{month}_{year}.csv"
    elif report_type == 'annual': 
        tasks = base_qs.filter(date__year=year).order_by('date')
        filename = f"WorkPlan_{target_user.username}_{year}_Annual.csv"
    else: 
        today = timezone.now().date()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        tasks = base_qs.filter(date__range=[start, end]).order_by('date')
        filename = f"WorkPlan_{target_user.username}_CurrentWeek.csv"

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    
    # UPDATED: Added "Task Owner" and "Role" columns
    writer.writerow([
        'Date', 
        'Task', 
        'Task Owner',  # NEW
        'Role',        # NEW: "Owner" or "Collaborator"
        'Centre', 
        'Dept', 
        'Collaborators', 
        'Other Parties', 
        'Status', 
        'Target', 
        'Resources', 
        'Comments (incl. Reschedule Reason)'
    ])
    
    for t in tasks:
        collabs = ", ".join([u.first_name for u in t.collaborators.all()])
        
        # Determine role
        role = "Owner" if t.work_plan.user == target_user else "Collaborator"
        
        # Combine comments + reschedule reason
        comments_parts = []
        if t.comments:
            comments_parts.append(t.comments.strip())
        if t.status == 'Rescheduled' and t.reschedule_reason:
            comments_parts.append(f"[Rescheduled Reason]: {t.reschedule_reason.strip()}")
        comments_display = " | ".join(comments_parts) if comments_parts else ""
        
        writer.writerow([
            t.date, 
            t.task_name,
            t.work_plan.user.get_full_name(),  # NEW
            role,                               # NEW
            t.centre.name if t.centre else '', 
            t.department.name if t.department else '',
            collabs,
            t.other_parties or '',
            t.status, 
            t.target or '',
            t.resources_needed or '',
            comments_display
        ])
    return response


# ============================================
# 2. _build_workplan_pdf function
# ============================================

def _build_workplan_pdf(work_plan_qs, user, title="Work Plan Report", report_type="weekly", period_str=None, target_user=None):
    """
    UPDATED: Now includes collaboration tasks
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=1.2*inch,
        bottomMargin=0.8*inch,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch
    )
    
    DARK_BLUE = colors.HexColor('#143C50')
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='ReportTitle', fontSize=18, fontName='Helvetica-Bold', textColor=DARK_BLUE, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='SubHeader', fontSize=12, fontName='Helvetica', textColor=colors.grey, alignment=TA_CENTER, spaceAfter=20))
    styles.add(ParagraphStyle(name='CellText', fontSize=8, leading=10, alignment=TA_LEFT))
    styles.add(ParagraphStyle(name='CellHeader', fontSize=9, leading=11, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_CENTER))

    story = []

    # Header Image
    header_img_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'document_title_1.png')
    if os.path.exists(header_img_path):
        header_img = Image(header_img_path, width=19.5*cm, height=1.4*cm)
        header_img.hAlign = 'CENTER'
        story.append(header_img)
        story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph(f"IT Department – {target_user}  Work Plan Report", styles['ReportTitle']))
    if period_str:
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(period_str, styles['SubHeader']))
    story.append(Spacer(1, 0.3*cm))

    # UPDATED: Headers include "Task Owner" and "Role"
    headers = [
        'Date', 
        'Task / Activity', 
        'Owner',        # NEW
        'Role',         # NEW
        'Centre / Dept', 
        'Collaborators', 
        'Other Parties', 
        'Comments', 
        'Target', 
        'Status'
    ]
    header_row = [Paragraph(h, styles['CellHeader']) for h in headers]
    data = [header_row]

    # UPDATED: Query includes collaboration tasks
    if target_user:
        tasks = WorkPlanTask.objects.filter(
            Q(work_plan__in=work_plan_qs) | Q(collaborators=target_user, date__range=[
                work_plan_qs[0].week_start_date if work_plan_qs else timezone.now().date(),
                work_plan_qs[0].week_end_date if work_plan_qs else timezone.now().date()
            ])
        ).distinct().order_by('date')
    else:
        tasks = WorkPlanTask.objects.filter(work_plan__in=work_plan_qs).order_by('date')

    for t in tasks:
        c_name = t.centre.name if t.centre else "N/A"
        d_name = t.department.name if t.department else "N/A"
        loc_str = f"<b>{c_name}</b><br/><i>{d_name}</i>"
        collabs = ", ".join([u.first_name for u in t.collaborators.all()]) if t.collaborators.exists() else "-"
        
        # Determine role
        role = "Owner" if (target_user and t.work_plan.user == target_user) else "Collaborator"
        role_color = "blue" if role == "Owner" else "purple"
        role_str = f"<font color='{role_color}'><b>{role}</b></font>"
        
        # Comments + Reschedule Reason
        comments_parts = []
        if t.comments:
            comments_parts.append(t.comments.strip())
        if t.status == 'Rescheduled' and t.reschedule_reason:
            comments_parts.append(f"[Rescheduled]: {t.reschedule_reason.strip()}")
        comments_display = "<br/>".join(comments_parts) if comments_parts else "-"

        status_color = "black"
        if t.status == 'Completed': status_color = "green"
        elif t.status == 'Not Done': status_color = "red"
        elif t.status == 'Rescheduled': status_color = "orange"
        status_str = f"<font color='{status_color}'>{t.status}</font>"

        task_name = f"<b>[ON LEAVE]</b> {t.task_name}" if t.is_leave else t.task_name

        row = [
            Paragraph(t.date.strftime('%d-%b'), styles['CellText']),
            Paragraph(task_name, styles['CellText']),
            Paragraph(t.work_plan.user.first_name, styles['CellText']),  # NEW
            Paragraph(role_str, styles['CellText']),                     # NEW
            Paragraph(loc_str, styles['CellText']),
            Paragraph(collabs, styles['CellText']),
            Paragraph(t.other_parties or '-', styles['CellText']),
            Paragraph(comments_display, styles['CellText']),
            Paragraph(t.target or '-', styles['CellText']),
            Paragraph(status_str, styles['CellText'])
        ]
        data.append(row)

    # UPDATED: Column widths adjusted for new columns
    col_widths = [1.8*cm, 4.2*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2.2*cm, 2.2*cm, 4.5*cm, 2.8*cm, 1.8*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), DARK_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    def add_text_watermark(canvas_obj, doc):
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica', 28)
        canvas_obj.setFillColor(colors.HexColor('#143C50'))
        canvas_obj.setFillAlpha(0.05) 
        
        x_step = 3 * inch
        y_step = 1.5 * inch
        header_cutoff = 6.5 * inch

        for x in range(-2, 14, 4):
            for y in range(-2, 10, 3):
                current_y = y * inch
                if current_y < header_cutoff:
                    canvas_obj.saveState()
                    canvas_obj.translate(x * inch, current_y)
                    canvas_obj.rotate(45)
                    canvas_obj.drawCentredString(0, 0, "MOHI IT")
                    canvas_obj.restoreState()
        
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=add_text_watermark, onLaterPages=add_text_watermark)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================
# 3. download_bulk_pdf_report function
# ============================================

@login_required
def download_bulk_pdf_report(request):
    """
    UPDATED: Now includes collaboration tasks in reports
    """
    filter_id = request.GET.get('user_filter')
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    report_type = request.GET.get('report_type', 'weekly')

    target_user = request.user
    if filter_id and (getattr(request.user, 'is_it_manager', False) or 
                     getattr(request.user, 'is_senior_it_officer', False) or 
                     request.user.is_superuser):
        try:
            target_user = User.objects.get(pk=filter_id)
        except:
            pass

    base_qs = WorkPlan.objects.filter(user=target_user)

    if report_type == 'monthly':
        work_plans = base_qs.filter(week_start_date__year=year, week_start_date__month=month)
        period_str = f"{calendar.month_name[month]} {year}"
        filename = f"WorkPlan_{target_user.username}_{month}_{year}_Report.pdf"
    elif report_type == 'annual':
        work_plans = base_qs.filter(week_start_date__year=year)
        period_str = f"Annual Report {year}"
        filename = f"WorkPlan_{target_user.username}_{year}_Annual.pdf"
    else:  # weekly
        today = timezone.now().date()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        work_plans = base_qs.filter(week_start_date=start)
        period_str = f"Week {start.strftime('%d %b')} - {end.strftime('%d %b %Y')}"
        filename = f"WorkPlan_{target_user.username}_Week_{start.strftime('%Y%m%d')}.pdf"

    if not work_plans.exists():
        messages.error(request, "No data found for the selected period.")
        return redirect('work_plan_calendar')

    # UPDATED: Pass target_user to include collaboration tasks
    pdf = _build_workplan_pdf(
        list(work_plans),
        request.user,
        title=f"Work Plan Report - {target_user.get_full_name()}",
        report_type=report_type,
        period_str=period_str,
        target_user=target_user  # NEW
    )

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



# ============================================
# 4. download_workplan_pdf function
# ============================================

@login_required
def download_workplan_pdf(request, pk):
    """
    UPDATED: Includes collaboration tasks for the work plan owner
    """
    work_plan = get_object_or_404(WorkPlan, pk=pk)
    is_owner = (request.user == work_plan.user)
    is_manager = is_manager_of(request.user, work_plan.user)
    is_collab = work_plan.tasks.filter(collaborators=request.user).exists()
    
    if not (is_owner or is_manager or is_collab):
        messages.error(request, "Access denied.")
        return redirect('work_plan_list')

    period_str = f"{work_plan.week_start_date.strftime('%d %B %Y')} - {work_plan.week_end_date.strftime('%d %B %Y')}"

    # UPDATED: Pass work_plan.user as target_user
    pdf = _build_workplan_pdf(
        [work_plan],
        request.user,
        title=f"Work Plan: {work_plan.user.get_full_name()}",
        report_type="weekly",
        period_str=period_str,
        target_user=work_plan.user  # NEW
    )
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="WorkPlan_{work_plan.week_start_date}.pdf"'
    return response


# ============================================
# 5. download_bulk_excel_report function (NEW)
# ============================================

@login_required
def download_bulk_excel_report(request):
    """
    NEW FUNCTION: Bulk Excel report similar to bulk PDF
    Includes tasks where user is owner OR collaborator
    """
    filter_id = request.GET.get('user_filter')
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    report_type = request.GET.get('report_type', 'weekly')

    target_user = request.user
    if filter_id and (getattr(request.user, 'is_it_manager', False) or 
                     getattr(request.user, 'is_senior_it_officer', False) or 
                     request.user.is_superuser):
        try:
            target_user = User.objects.get(pk=filter_id)
        except:
            pass
    
    # UPDATED: Include tasks where user is owner OR collaborator
    base_qs = WorkPlanTask.objects.filter(
        Q(work_plan__user=target_user) | Q(collaborators=target_user)
    ).distinct()

    if report_type == 'monthly': 
        tasks = base_qs.filter(date__year=year, date__month=month).order_by('date')
        period_str = f"{calendar.month_name[month]} {year}"
        filename = f"WorkPlan_{target_user.username}_{month}_{year}_Report.xlsx"
    elif report_type == 'annual': 
        tasks = base_qs.filter(date__year=year).order_by('date')
        period_str = f"Annual Report {year}"
        filename = f"WorkPlan_{target_user.username}_{year}_Annual_Report.xlsx"
    else:  # weekly
        today = timezone.now().date()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        tasks = base_qs.filter(date__range=[start, end]).order_by('date')
        period_str = f"Week {start.strftime('%d %b')} - {end.strftime('%d %b %Y')}"
        filename = f"WorkPlan_{target_user.username}_Week_{start.strftime('%Y%m%d')}_Report.xlsx"

    if not tasks.exists():
        messages.error(request, "No data found for the selected period.")
        return redirect('work_plan_calendar')

    # Create Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Work Plan Report"

    # === STYLING ===
    header_fill = PatternFill(start_color="143C50", end_color="143C50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # === TITLE ===
    worksheet.merge_cells('A1:L1')
    title_cell = worksheet['A1']
    title_cell.value = f"IT Department – Work Plan Report"
    title_cell.font = Font(bold=True, size=16, color="143C50")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # === PERIOD ===
    worksheet.merge_cells('A2:L2')
    period_cell = worksheet['A2']
    period_cell.value = period_str
    period_cell.font = Font(size=12, color="666666")
    period_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # === USER INFO ===
    worksheet.merge_cells('A3:L3')
    user_cell = worksheet['A3']
    user_cell.value = f"Report for: {target_user.get_full_name()}"
    user_cell.font = Font(size=11, color="333333")
    user_cell.alignment = Alignment(horizontal="center", vertical="center")

    # === HEADERS (Row 5) ===
    headers = [
        'Date', 
        'Task / Activity', 
        'Task Owner',
        'Role',
        'Centre', 
        'Department',
        'Collaborators', 
        'Other Parties', 
        'Status', 
        'Target', 
        'Resources', 
        'Comments (incl. Reschedule Reason)'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # === DATA ROWS ===
    row_num = 6
    for task in tasks:
        # Determine role
        role = "Owner" if task.work_plan.user == target_user else "Collaborator"
        
        # Collaborators
        collabs = ", ".join([u.get_full_name() for u in task.collaborators.all()]) if task.collaborators.exists() else "-"
        
        # Comments + Reschedule Reason
        comments_parts = []
        if task.comments:
            comments_parts.append(task.comments.strip())
        if task.status == 'Rescheduled' and task.reschedule_reason:
            comments_parts.append(f"[Rescheduled Reason]: {task.reschedule_reason.strip()}")
        comments_display = " | ".join(comments_parts) if comments_parts else ""
        
        # Task name (with leave indicator)
        task_name = f"[ON LEAVE] {task.task_name}" if task.is_leave else task.task_name
        
        row_data = [
            task.date.strftime('%d-%b-%Y'),
            task_name,
            task.work_plan.user.get_full_name(),
            role,
            task.centre.name if task.centre else 'N/A',
            task.department.name if task.department else 'N/A',
            collabs,
            task.other_parties or '-',
            task.status,
            task.target or '-',
            task.resources_needed or '-',
            comments_display
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = border
            
            # Status color coding
            if col_num == 9:  # Status column
                if task.status == 'Completed':
                    cell.font = Font(color="008000", bold=True)  # Green
                elif task.status == 'Not Done':
                    cell.font = Font(color="FF0000", bold=True)  # Red
                elif task.status == 'Rescheduled':
                    cell.font = Font(color="FF8C00", bold=True)  # Orange
            
            # Role color coding
            if col_num == 4:  # Role column
                if role == "Owner":
                    cell.font = Font(color="0000FF", bold=True)  # Blue
                else:
                    cell.font = Font(color="800080", bold=True)  # Purple
        
        row_num += 1

    # === COLUMN WIDTHS ===
    column_widths = {
        'A': 12,   # Date
        'B': 30,   # Task
        'C': 18,   # Task Owner
        'D': 12,   # Role
        'E': 18,   # Centre
        'F': 18,   # Department
        'G': 20,   # Collaborators
        'H': 18,   # Other Parties
        'I': 12,   # Status
        'J': 15,   # Target
        'K': 20,   # Resources
        'L': 35    # Comments
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # === ROW HEIGHTS ===
    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 20
    worksheet.row_dimensions[3].height = 18
    worksheet.row_dimensions[5].height = 30

    # === FREEZE PANES (Header row) ===
    worksheet.freeze_panes = 'A6'

    # === SAVE TO RESPONSE ===
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
